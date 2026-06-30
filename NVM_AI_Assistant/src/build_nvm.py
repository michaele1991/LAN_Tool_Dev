"""
GBE NVM Image Builder — pure Python, no Excel required
=======================================================
Reads the 'full nvm map' sheet from an .xlsm file, assembles
16-bit NVM words from bit-field rows, applies checksum at word 0x3F,
and writes .bin + .txt output files.

Usage (CLI):
    python build_nvm.py --platform Nahum13_ptl_pcd_p_h --step A0 \
                        --version 1.4 --variant V --output ./output

Requirements: openpyxl  (pip install openpyxl)
"""

import argparse
import struct
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet",
                           "--index-url", "https://pypi.org/simple/", "openpyxl"])
    from openpyxl import load_workbook

# ── constants ──────────────────────────────────────────────────────────────
CSUM_WORD  = 0x3F
CSUM_MAGIC = 0xBABA
NVM_SHEET  = "full nvm map"
GEN_SHEET  = "general variable"
DATA_ROW   = 7        # 1-based first data row in full nvm map

# Column indices (0-based, from iter_rows values_only=True)
COL_OFFSET = 0   # A - LAN Word Offset
COL_BITS   = 1   # B - Bits (e.g. "15:0", "7", "11:8")
COL_RTL    = 2   # C - RTL name
COL_CSPEC  = 3   # D - C-Spec name
COL_DESC   = 5   # F - Description
COL_V      = 6   # G - V value  (LAN SW / Consumer)
COL_LM     = 7   # H - LM value (Non-LAN SW / Corporate)

VARIANTS   = ("V", "LM")
VAR_COL    = {"V": COL_V, "LM": COL_LM}
VAR_LABEL  = {"V": "Consumer (V)",  "LM": "Corporate (LM)"}
VAR_SUFFIX = {"V": "Cons_Prod_NA",  "LM": "Corp_Prod_NA"}


# ── helpers ────────────────────────────────────────────────────────────────

def _find_platforms_root() -> Path | None:
    """Locate GBE_Image (or Eng_GBE_Image) folder relative to this script."""
    here = Path(__file__).resolve().parent
    for ancestor in [here, here.parent, here.parent.parent]:
        for name in ("Eng_GBE_Image", "GBE_Image_Creator/GBE_Image",
                     "GBE_Image", "platforms"):
            candidate = ancestor / name
            if candidate.is_dir():
                return candidate
    return None


def list_platforms(platforms_root: Path) -> list[Path]:
    """Return sorted list of platform subfolders that contain an .xlsm file."""
    return sorted(
        f for f in platforms_root.iterdir()
        if f.is_dir() and not f.name.startswith(".") and list(f.glob("*.xlsm"))
    )


def find_xlsm(platform_folder: Path) -> Path:
    xlsms = list(platform_folder.glob("*.xlsm"))
    if not xlsms:
        raise FileNotFoundError(f"No .xlsm found in {platform_folder}")
    return xlsms[0]


def _parse_offset(raw) -> int | None:
    """Convert '0ah', '0x0A', '10' etc. → int."""
    s = str(raw).strip().lower()
    s = s.rstrip("h")
    try:
        return int(s, 16) if (len(s) > 1 and any(c in "abcdef" for c in s)) or s.startswith("0") and len(s) > 1 else int(s, 16)
    except ValueError:
        try:
            return int(s)
        except ValueError:
            return None


def _parse_bits(raw) -> tuple[int, int] | None:
    """'15:0' → (15, 0),  '7' → (7, 7),  '' → None."""
    s = str(raw).strip()
    if not s or s.lower() == "none":
        return None
    if ":" in s:
        parts = s.split(":")
        try:
            return (int(parts[0]), int(parts[1]))
        except ValueError:
            return None
    try:
        b = int(s)
        return (b, b)
    except ValueError:
        return None


def _parse_value(raw) -> int | None:
    """'0x1', '0x0', '1', 'N/A' → int or None."""
    s = str(raw).strip()
    if not s or s.upper() in ("N/A", "NONE", ""):
        return None
    try:
        return int(s, 0)
    except ValueError:
        return None


# ── core assembly ──────────────────────────────────────────────────────────

def read_nvm_fields(xlsm_path: Path) -> dict[int, list[dict]]:
    """
    Read 'full nvm map' and return:
        { offset_int: [ {bits_str, hi, lo, v_val, lm_val, desc}, ... ] }
    """
    wb = load_workbook(str(xlsm_path), read_only=True, data_only=True, keep_vba=False)
    if NVM_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{NVM_SHEET}' not found in {xlsm_path.name}")

    ws = wb[NVM_SHEET]
    fields: dict[int, list[dict]] = {}

    for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
        if row[COL_OFFSET] is None:
            continue
        off = _parse_offset(row[COL_OFFSET])
        if off is None:
            continue

        bits_range = _parse_bits(row[COL_BITS])
        v_raw   = row[COL_V]   if len(row) > COL_V   else None
        lm_raw  = row[COL_LM]  if len(row) > COL_LM  else None

        entry = {
            "bits_str": str(row[COL_BITS] or "").strip(),
            "hi":  bits_range[0] if bits_range else None,
            "lo":  bits_range[1] if bits_range else None,
            "v":   _parse_value(v_raw),
            "lm":  _parse_value(lm_raw),
            "desc": str(row[COL_CSPEC] or row[COL_RTL] or "").strip(),
        }
        fields.setdefault(off, []).append(entry)

    wb.close()
    return fields


def assemble_words(fields: dict[int, list[dict]], variant: str) -> dict[int, int]:
    """Combine bit-field rows into whole 16-bit words for the chosen variant."""
    val_key = "v" if variant == "V" else "lm"
    words: dict[int, int] = {}

    for off, field_list in fields.items():
        word = 0
        for f in field_list:
            hi, lo = f["hi"], f["lo"]
            val = f[val_key]
            if hi is None or val is None:
                continue
            width = hi - lo + 1
            mask  = (1 << width) - 1
            word |= (val & mask) << lo
        words[off] = word & 0xFFFF

    return words


def apply_checksum(words: dict[int, int]) -> dict[int, int]:
    """Set word[0x3F] so that sum(word[0x00..0x3F]) + 0xBABA ≡ 0 (mod 0x10000)."""
    total = sum(words.get(i, 0) for i in range(CSUM_WORD))
    words[CSUM_WORD] = (-total + 0x10000 * 4 + CSUM_MAGIC) & 0xFFFF
    return words


def words_to_bytes(words: dict[int, int]) -> bytes:
    """Pack word dict into a little-endian binary blob."""
    max_off = max(words.keys(), default=CSUM_WORD)
    size    = max(max_off + 1, CSUM_WORD + 1)
    buf = bytearray(size * 2)
    for off, val in words.items():
        struct.pack_into("<H", buf, off * 2, val)
    return bytes(buf)


def words_to_txt(words: dict[int, int]) -> str:
    """Hex dump: 8 words per line, space-separated (matches Intel NVM format)."""
    max_off = max(words.keys(), default=CSUM_WORD)
    size    = max(max_off + 1, CSUM_WORD + 1)
    lines = []
    for base in range(0, size, 8):
        chunk = [f"{words.get(base + j, 0):04X}" for j in range(min(8, size - base))]
        lines.append(" ".join(chunk))
    return "\n".join(lines) + "\n"


def read_general_variable(xlsm_path: Path) -> dict[str, str]:
    """Read the 'general variable' sheet → {label: value}."""
    wb = load_workbook(str(xlsm_path), read_only=True, data_only=True, keep_vba=False)
    info = {}
    if GEN_SHEET in wb.sheetnames:
        ws = wb[GEN_SHEET]
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if row[0] and row[1] is not None:
                info[str(row[0]).strip()] = str(row[1]).strip()
    wb.close()
    return info


def apply_nvm_changes(words: dict[int, int], changes: list[dict],
                      variant: str) -> dict[int, int]:
    """
    Apply bit-level overrides to the assembled word dict.
    Each change: {offset, bit, new_value, variants}  where variants is ['V'], ['LM'], or ['V','LM']
    """
    for c in changes:
        if variant not in c["variants"]:
            continue
        off = c["offset"]
        bit = c["bit"]
        val = c["new_value"]
        word = words.get(off, 0)
        if val:
            word |=  (1 << bit)
        else:
            word &= ~(1 << bit)
        words[off] = word & 0xFFFF
    return words


# ── build pipeline ─────────────────────────────────────────────────────────

def build(platform_folder: Path,
          step: str,
          version: str,
          variants: list[str],
          output_dir: Path | None = None,
          nvm_changes: list[dict] | None = None) -> list[Path]:
    """
    Full build for one platform.
    Returns list of output .bin paths.
    """
    xlsm = find_xlsm(platform_folder)
    info = read_general_variable(xlsm)

    # Derive output name from general variable or folder name
    base_name = (info.get("image file name") or info.get("Project name")
                 or platform_folder.name)

    major, minor = (version.split(".", 1) + ["0"])[:2]
    ver_str = f"{major}.{minor}"

    if output_dir is None:
        output_dir = platform_folder.parent.parent.parent / "output" / platform_folder.name
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\nReading NVM map from: {xlsm.name}")
    fields = read_nvm_fields(xlsm)
    print(f"  Loaded {sum(len(v) for v in fields.values())} bit-field rows "
          f"across {len(fields)} unique offsets")

    outputs = []
    for variant in variants:
        label = VAR_LABEL[variant]
        suffix = VAR_SUFFIX[variant]
        fname = f"{base_name}_{step}_{ver_str}_Release_{suffix}"

        print(f"\nAssembling {label}...")
        words = assemble_words(fields, variant)
        if nvm_changes:
            words = apply_nvm_changes(words, nvm_changes, variant)
            applied = [c for c in nvm_changes if variant in c["variants"]]
            if applied:
                print(f"  Applying {len(applied)} NVM override(s):")
                for c in applied:
                    print(f"    0x{c['offset']:02X} bit {c['bit']} → {c['new_value']}")
        words = apply_checksum(words)

        bin_path = output_dir / f"{fname}.bin"
        txt_path = output_dir / f"{fname}.txt"

        bin_path.write_bytes(words_to_bytes(words))
        txt_path.write_text(words_to_txt(words), encoding="utf-8")

        csum = words.get(CSUM_WORD, 0)
        size = len(words_to_bytes(words))
        print(f"  ✓  {bin_path.name}  ({size} bytes, checksum=0x{csum:04X})")
        outputs.append(bin_path)

    print(f"\nOutput folder: {output_dir}")
    return outputs


# ── CLI entry point ────────────────────────────────────────────────────────

def _cli():
    parser = argparse.ArgumentParser(
        description="Build a GBE NVM binary image from an XLSM NVM map.")
    parser.add_argument("--platform", "-p",
                        help="Platform folder name (e.g. Nahum13_ptl_pcd_p_h)")
    parser.add_argument("--step",    "-s", default="A0",
                        help="Silicon step (default: A0)")
    parser.add_argument("--version", "-v", default="1.4",
                        help="Image version string (default: 1.4)")
    parser.add_argument("--variant", "-r", default="Both",
                        choices=["V", "LM", "Both"],
                        help="Variant: V, LM, or Both (default: Both)")
    parser.add_argument("--output",  "-o", default=None,
                        help="Output directory (default: ./output/<platform>)")
    args = parser.parse_args()

    root = _find_platforms_root()
    if root is None:
        print("ERROR: Cannot locate GBE_Image platforms folder.", file=sys.stderr)
        sys.exit(1)

    platforms = list_platforms(root)
    if not platforms:
        print(f"ERROR: No platforms found under {root}", file=sys.stderr)
        sys.exit(1)

    # Resolve platform
    if args.platform:
        matches = [p for p in platforms if p.name.lower() == args.platform.lower()]
        if not matches:
            print(f"ERROR: Platform '{args.platform}' not found. Available:")
            for p in platforms:
                print(f"  {p.name}")
            sys.exit(1)
        platform_folder = matches[0]
    else:
        print("Available platforms:")
        for i, p in enumerate(platforms):
            print(f"  [{i}] {p.name}")
        sys.exit(0)

    variants = ["V", "LM"] if args.variant == "Both" else [args.variant]
    out_dir  = Path(args.output) if args.output else None

    build(platform_folder, args.step, args.version, variants, out_dir)


if __name__ == "__main__":
    _cli()
