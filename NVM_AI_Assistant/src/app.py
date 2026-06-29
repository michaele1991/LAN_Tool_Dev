"""
Intel GBE NVM Tools
Tab 1 - AI Assistant: ask Claude about NVM config
Tab 2 - Image Builder: wizard to look up current values and generate a change report
"""

import json
import os
import subprocess
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from pathlib import Path
import threading
from datetime import datetime

REQUIRED_PACKAGES = ["anthropic", "openpyxl"]


def _ensure_packages():
    missing = []
    for pkg in REQUIRED_PACKAGES:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if not missing:
        return
    print(f"Installing: {', '.join(missing)} ...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet",
                           "--index-url", "https://pypi.org/simple/"] + missing)
    os.execv(sys.executable, [sys.executable] + sys.argv)


_ensure_packages()
import anthropic  # noqa

APP_TITLE   = "Intel GBE NVM Tools"
APP_VERSION = "v1.1"
MODEL       = "claude-opus-4-5"
MAX_TOKENS  = 4096
CONFIG_FILE = Path(__file__).parent / "config.json"
NVM_ROOT    = "Eng_GBE_Image"

# XLSM column indices (full nvm map sheet, data starts row index 6 after 0-based)
COL_OFFSET  = 0   # A - LAN Word Offset
COL_BITS    = 1   # B - bits
COL_RTL     = 2   # C - name in RTL
COL_CSPEC   = 3   # D - name in c-spec
COL_DESC    = 5   # F - description
COL_V       = 6   # G - V version
COL_LM      = 7   # H - LM version

SYSTEM_PROMPT = """You are an expert Intel GBE (Gigabit Ethernet) NVM (Non-Volatile Memory) configuration engineer.
You help engineers configure NVM settings for Intel Ethernet controllers
(Nahum7, Nahum8, Nahum9, Nahum10, Nahum11, Nahum13, MTL and similar platforms).
Your expertise: NVM word/bit field definitions, LAN SW (V) vs Non-LAN SW (LM) differences,
Device ID, WoL, PCIe, power management, EEE, checksum (word 0x3F = 0xBABA), C-Spec/RTL naming.
When answering: confirm offset+bits, give exact hex values for V/LM, note caveats, warn about risks.
Format changes as a table: | Offset | Bits | Field | V Value | LM Value | Reason |"""


def load_config():
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"api_key": "", "model": MODEL}


def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2), encoding="utf-8")


def find_nvm_root():
    here = Path(__file__).resolve().parent
    for candidate in [here, here.parent, here.parent.parent, here.parent.parent.parent]:
        p = candidate / NVM_ROOT
        if p.is_dir():
            return p
    return None


def norm_offset(s):
    """Normalize offset string to compare: '0x0a' -> '0ah', '0Ah' -> '0ah'"""
    s = s.strip().lower()
    if s.startswith("0x"):
        s = s[2:] + "h"
    elif not s.endswith("h"):
        s = s + "h"
    return s


def read_nvm_map(xlsm_path):
    """Read full nvm map sheet. Returns list of row dicts."""
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsm_path), read_only=True, data_only=True, keep_vba=False)
    rows = []
    if "full nvm map" in wb.sheetnames:
        ws = wb["full nvm map"]
        for row in ws.iter_rows(min_row=7, values_only=True):   # data starts row 7 (1-based)
            if row[COL_OFFSET] is not None:
                rows.append({
                    "offset":  str(row[COL_OFFSET]).strip() if row[COL_OFFSET] else "",
                    "bits":    str(row[COL_BITS]).strip()   if row[COL_BITS]   else "",
                    "rtl":     str(row[COL_RTL]).strip()    if row[COL_RTL]    else "",
                    "cspec":   str(row[COL_CSPEC]).strip()  if row[COL_CSPEC]  else "",
                    "desc":    str(row[COL_DESC]).strip()   if row[COL_DESC]   else "",
                    "v_val":   str(row[COL_V]).strip()      if row[COL_V]      else "N/A",
                    "lm_val":  str(row[COL_LM]).strip()     if row[COL_LM]     else "N/A",
                })
    wb.close()
    return rows


def lookup_register(nvm_rows, offset_raw, bits_raw=""):
    """Find matching rows by offset (and optionally bits)."""
    target_off = norm_offset(offset_raw)
    bits_raw = bits_raw.strip()
    matches = []
    for r in nvm_rows:
        row_off = norm_offset(r["offset"]) if r["offset"] else ""
        if row_off == target_off:
            if not bits_raw or r["bits"] == bits_raw:
                matches.append(r)
    return matches


class NvmApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE}  {APP_VERSION}")
        self.geometry("1200x840")
        self.minsize(900, 640)
        self.config_data = load_config()
        self.client = None
        self.conversation = []
        self.nvm_root = find_nvm_root()
        self.ai_excel_context = ""
        self.nvm_rows = []          # loaded rows for current Image Builder project
        self.change_list = []       # list of change dicts for Image Builder
        self._build_ui()
        self._load_projects()
        self._try_init_client()

    # ═══════════════════════════════════════════════════════════════════════
    # UI
    # ═══════════════════════════════════════════════════════════════════════

    def _build_ui(self):
        s = ttk.Style()
        try:
            s.theme_use("vista")
        except Exception:
            pass

        # Header
        hdr = tk.Frame(self, bg="#0071c5", height=46)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"  Intel GBE NVM Tools  {APP_VERSION}",
                 font=("Segoe UI", 13, "bold"), fg="white", bg="#0071c5").pack(side=tk.LEFT, padx=14, pady=8)
        tk.Label(hdr, text="AI: Claude (Anthropic)",
                 font=("Segoe UI", 9), fg="#a8d4f0", bg="#0071c5").pack(side=tk.RIGHT, padx=16)

        # Notebook
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        tab1 = ttk.Frame(self.nb)
        tab2 = ttk.Frame(self.nb)
        self.nb.add(tab1, text="  AI Assistant  ")
        self.nb.add(tab2, text="  Image Builder  ")

        self._build_ai_tab(tab1)
        self._build_builder_tab(tab2)

        # Status bar
        sb = tk.Frame(self, bg="#e8e8e8", height=22, relief=tk.SUNKEN, bd=1)
        sb.pack(side=tk.BOTTOM, fill=tk.X)
        sb.pack_propagate(False)
        self.api_status_lbl = tk.Label(sb, text="API: not connected", fg="#888",
                                       font=("Segoe UI", 8), bg="#e8e8e8")
        self.api_status_lbl.pack(side=tk.LEFT, padx=8)
        root_txt = f"NVM root: {self.nvm_root}" if self.nvm_root else f"'{NVM_ROOT}' not found"
        self.nvm_status_lbl = tk.Label(sb, text=root_txt,
                                       fg="#555" if self.nvm_root else "#cc0000",
                                       font=("Segoe UI", 8), bg="#e8e8e8")
        self.nvm_status_lbl.pack(side=tk.LEFT, padx=16)
        ttk.Button(sb, text="Settings", command=self._open_settings).pack(side=tk.RIGHT, padx=8)

    # ── Tab 1: AI Assistant ────────────────────────────────────────────────

    def _build_ai_tab(self, parent):
        main = tk.Frame(parent)
        main.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        form_frame = tk.Frame(main, width=300)
        form_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8))
        form_frame.pack_propagate(False)
        self._build_ai_form(form_frame)

        chat_frame = tk.Frame(main)
        chat_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._build_chat(chat_frame)

    def _card(self, parent, title):
        f = ttk.LabelFrame(parent, text=title, padding=(8, 4))
        f.pack(fill=tk.X, pady=(0, 8))
        return f

    def _build_ai_form(self, parent):
        tk.Label(parent, text="Ask AI about NVM config",
                 font=("Segoe UI", 10, "bold"), fg="#0071c5").pack(anchor=tk.W, pady=(4, 8))

        c = self._card(parent, "Project")
        c.columnconfigure(1, weight=1)
        tk.Label(c, text="Platform:", font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", pady=2)
        self.ai_project_var = tk.StringVar()
        self.ai_project_cb = ttk.Combobox(c, textvariable=self.ai_project_var,
                                          state="readonly", width=22)
        self.ai_project_cb.grid(row=0, column=1, sticky="ew", padx=(6, 0), pady=2)
        self.ai_project_cb.bind("<<ComboboxSelected>>", lambda _: self._ai_on_project_change())

        c2 = self._card(parent, "Register")
        c2.columnconfigure(1, weight=1)
        for i, (lbl, attr, hint) in enumerate([
            ("Offset (hex):", "ai_offset_var", ""),
            ("Bit(s):",       "ai_bits_var",   "e.g. 3  or  15:8"),
            ("Name:",         "ai_name_var",   "RTL/C-Spec (optional)"),
        ]):
            tk.Label(c2, text=lbl, font=("Segoe UI", 9)).grid(row=i*2, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            setattr(self, attr, var)
            ttk.Entry(c2, textvariable=var, width=18).grid(row=i*2, column=1, sticky="ew",
                                                           padx=(6, 0), pady=2)
            if hint:
                tk.Label(c2, text=hint, font=("Segoe UI", 8), fg="#999").grid(
                    row=i*2+1, column=0, columnspan=2, sticky="w")

        c3 = self._card(parent, "Mode")
        self.ai_mode_var = tk.StringVar(value="Both")
        for v, l in [("V", "V  (Consumer)"), ("LM", "LM  (Corporate)"), ("Both", "Both")]:
            ttk.Radiobutton(c3, text=l, value=v, variable=self.ai_mode_var).pack(anchor=tk.W, pady=1)

        c4 = self._card(parent, "Request")
        self.ai_value_var = tk.StringVar()
        ttk.Entry(c4, textvariable=self.ai_value_var, width=26).pack(fill=tk.X, pady=(0, 2))
        tk.Label(c4, text='"enable WoL"  "0x1"  "explain"',
                 font=("Segoe UI", 8), fg="#999").pack(anchor=tk.W)

        ttk.Button(parent, text="Ask AI", command=self._ai_ask).pack(fill=tk.X, pady=(10, 3))
        ttk.Button(parent, text="Clear Chat", command=self._clear_chat).pack(fill=tk.X, pady=2)
        ttk.Button(parent, text="Save Chat",  command=self._save_chat).pack(fill=tk.X, pady=2)

    def _build_chat(self, parent):
        paned = ttk.PanedWindow(parent, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True)

        hist = ttk.LabelFrame(paned, text="Conversation", padding=4)
        paned.add(hist, weight=4)
        self.chat_display = scrolledtext.ScrolledText(
            hist, wrap=tk.WORD, state=tk.DISABLED,
            font=("Segoe UI", 10), bg="#fafafa", relief=tk.FLAT)
        self.chat_display.pack(fill=tk.BOTH, expand=True)
        self.chat_display.tag_configure("user",      foreground="#0055aa", font=("Segoe UI", 10, "bold"))
        self.chat_display.tag_configure("assistant", foreground="#1a1a1a", font=("Segoe UI", 10))
        self.chat_display.tag_configure("system",    foreground="#888888", font=("Segoe UI", 9, "italic"))
        self.chat_display.tag_configure("error",     foreground="#cc0000", font=("Segoe UI", 10))

        fu = ttk.LabelFrame(paned, text="Follow-up  (Ctrl+Enter)", padding=4)
        paned.add(fu, weight=1)
        self.input_text = tk.Text(fu, height=4, wrap=tk.WORD,
                                  font=("Segoe UI", 10), relief=tk.SOLID, bd=1)
        self.input_text.pack(fill=tk.BOTH, expand=True)
        self.input_text.bind("<Control-Return>", lambda e: self._send_followup())
        br = ttk.Frame(fu)
        br.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(br, text="Send  [Ctrl+Enter]", command=self._send_followup).pack(side=tk.RIGHT, padx=4)
        self.status_var = tk.StringVar(value="Fill the form and click Ask AI")
        tk.Label(br, textvariable=self.status_var, font=("Segoe UI", 8), fg="#555").pack(side=tk.LEFT, padx=4)

    # ── Tab 2: Image Builder ───────────────────────────────────────────────

    def _build_builder_tab(self, parent):
        main = tk.Frame(parent)
        main.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        # LEFT: input form
        left = tk.Frame(main, width=320)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8))
        left.pack_propagate(False)
        self._build_builder_form(left)

        # RIGHT: results
        right = tk.Frame(main)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._build_builder_results(right)

    def _build_builder_form(self, parent):
        tk.Label(parent, text="GBE Image Change Builder",
                 font=("Segoe UI", 10, "bold"), fg="#0071c5").pack(anchor=tk.W, pady=(4, 8))

        # Project
        c = self._card(parent, "1. Project")
        c.columnconfigure(1, weight=1)
        tk.Label(c, text="Platform:", font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", pady=2)
        self.bld_project_var = tk.StringVar()
        self.bld_project_cb = ttk.Combobox(c, textvariable=self.bld_project_var,
                                           state="readonly", width=22)
        self.bld_project_cb.grid(row=0, column=1, sticky="ew", padx=(6, 0), pady=2)
        self.bld_project_cb.bind("<<ComboboxSelected>>", lambda _: self._bld_on_project_change())
        self.bld_xlsm_lbl = tk.Label(c, text="—", font=("Segoe UI", 8), fg="#0071c5",
                                     wraplength=220, justify=tk.LEFT)
        self.bld_xlsm_lbl.grid(row=1, column=0, columnspan=2, sticky="w")

        # Register
        c2 = self._card(parent, "2. Register to Change")
        c2.columnconfigure(1, weight=1)
        fields = [
            ("Offset (hex):", "bld_offset_var", "e.g.  0x15  or  15h"),
            ("Bit(s):",       "bld_bits_var",   "e.g.  3   or   15:8  (optional)"),
        ]
        for i, (lbl, attr, hint) in enumerate(fields):
            tk.Label(c2, text=lbl, font=("Segoe UI", 9)).grid(row=i*2, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            setattr(self, attr, var)
            ttk.Entry(c2, textvariable=var, width=16).grid(row=i*2, column=1, sticky="w",
                                                           padx=(6, 0), pady=2)
            tk.Label(c2, text=hint, font=("Segoe UI", 8), fg="#999").grid(
                row=i*2+1, column=0, columnspan=2, sticky="w")

        ttk.Button(c2, text="Lookup Current Values",
                   command=self._bld_lookup).grid(row=5, column=0, columnspan=2,
                                                   sticky="ew", pady=(10, 2))

        # Lookup result
        self.bld_lookup_lbl = tk.Label(c2, text="", font=("Segoe UI", 8),
                                       fg="#333", wraplength=280, justify=tk.LEFT)
        self.bld_lookup_lbl.grid(row=6, column=0, columnspan=2, sticky="w", pady=2)

        # New values
        c3 = self._card(parent, "3. New Values")
        c3.columnconfigure(1, weight=1)
        self.bld_mode_var = tk.StringVar(value="Both")
        for v, l in [("V", "V only"), ("LM", "LM only"), ("Both", "Both V and LM")]:
            ttk.Radiobutton(c3, text=l, value=v, variable=self.bld_mode_var,
                            command=self._bld_toggle_inputs).pack(anchor=tk.W, pady=1)

        self.bld_v_frame = ttk.Frame(c3)
        self.bld_v_frame.pack(fill=tk.X, pady=(4, 0))
        tk.Label(self.bld_v_frame, text="New V value:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.bld_new_v_var = tk.StringVar()
        ttk.Entry(self.bld_v_frame, textvariable=self.bld_new_v_var, width=12).pack(side=tk.LEFT, padx=(6, 0))

        self.bld_lm_frame = ttk.Frame(c3)
        self.bld_lm_frame.pack(fill=tk.X, pady=(4, 0))
        tk.Label(self.bld_lm_frame, text="New LM value:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.bld_new_lm_var = tk.StringVar()
        ttk.Entry(self.bld_lm_frame, textvariable=self.bld_new_lm_var, width=12).pack(side=tk.LEFT, padx=(6, 0))

        # Reason
        c4 = self._card(parent, "4. Reason / Comment (optional)")
        self.bld_reason_var = tk.StringVar()
        ttk.Entry(c4, textvariable=self.bld_reason_var, width=30).pack(fill=tk.X)

        # Action buttons
        ttk.Button(parent, text="+ Add to Change List",
                   command=self._bld_add_change).pack(fill=tk.X, pady=(10, 3))
        ttk.Button(parent, text="Generate Report + JSON",
                   command=self._bld_generate).pack(fill=tk.X, pady=3)
        ttk.Button(parent, text="Clear All",
                   command=self._bld_clear_all).pack(fill=tk.X, pady=3)

    def _build_builder_results(self, parent):
        # Change list (top)
        list_frame = ttk.LabelFrame(parent, text="Change List", padding=4)
        list_frame.pack(fill=tk.BOTH, expand=False, pady=(0, 6))

        cols = ("offset", "bits", "field", "old_v", "new_v", "old_lm", "new_lm", "reason")
        self.change_tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=7)
        headers = {"offset": "Offset", "bits": "Bits", "field": "Field Name",
                   "old_v": "Old V", "new_v": "New V",
                   "old_lm": "Old LM", "new_lm": "New LM", "reason": "Reason"}
        widths  = {"offset": 70, "bits": 60, "field": 180,
                   "old_v": 70, "new_v": 70, "old_lm": 70, "new_lm": 70, "reason": 200}
        for col in cols:
            self.change_tree.heading(col, text=headers[col])
            self.change_tree.column(col, width=widths[col], minwidth=50)
        sb = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL,
                           command=self.change_tree.xview)
        self.change_tree.configure(xscrollcommand=sb.set)
        self.change_tree.pack(fill=tk.BOTH, expand=True)
        sb.pack(fill=tk.X)

        ttk.Button(list_frame, text="Remove selected",
                   command=self._bld_remove_selected).pack(anchor=tk.E, pady=(4, 0))

        # Report (bottom)
        rep_frame = ttk.LabelFrame(parent, text="Generated Report", padding=4)
        rep_frame.pack(fill=tk.BOTH, expand=True)
        self.report_text = scrolledtext.ScrolledText(
            rep_frame, wrap=tk.WORD, font=("Consolas", 9), bg="#f8f8f8", relief=tk.FLAT)
        self.report_text.pack(fill=tk.BOTH, expand=True)
        self.report_text.tag_configure("header",  foreground="#0071c5", font=("Consolas", 9, "bold"))
        self.report_text.tag_configure("added",   foreground="#107c10")
        self.report_text.tag_configure("removed", foreground="#cc0000")
        self.report_text.tag_configure("info",    foreground="#555")

    # ═══════════════════════════════════════════════════════════════════════
    # Project loading (shared)
    # ═══════════════════════════════════════════════════════════════════════

    def _load_projects(self):
        if self.nvm_root and self.nvm_root.exists():
            projects = sorted(d.name for d in self.nvm_root.iterdir() if d.is_dir())
        else:
            projects = ["MTL_M_P", "Nahum11_mtl_m_p", "Nahum13_nvl_pch_s",
                        "Nahum10_adp_lp", "Nahum9_tgl_lp"]
        for cb in (self.ai_project_cb, self.bld_project_cb):
            cb["values"] = projects
        if projects:
            self.ai_project_cb.current(0)
            self.bld_project_cb.current(0)
            self._ai_on_project_change()
            self._bld_on_project_change()

    def _get_xlsm(self, project):
        if not self.nvm_root:
            return None
        p = self.nvm_root / project
        files = list(p.glob("*.xlsm")) if p.exists() else []
        return files[0] if files else None

    def _ai_on_project_change(self):
        xlsm = self._get_xlsm(self.ai_project_var.get())
        if xlsm:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(xlsm), read_only=True, data_only=True, keep_vba=False)
                lines = []
                for ws in wb.worksheets:
                    lines.append(f"\n== {ws.title} ==")
                    for row in ws.iter_rows(max_row=300, values_only=True):
                        if any(c is not None for c in row):
                            lines.append("\t".join("" if c is None else str(c) for c in row))
                wb.close()
                self.ai_excel_context = "\n".join(lines)
            except Exception:
                self.ai_excel_context = ""

    def _bld_on_project_change(self):
        proj = self.bld_project_var.get()
        xlsm = self._get_xlsm(proj)
        if xlsm:
            self.bld_xlsm_lbl.configure(text=xlsm.name, fg="#107c10")
            try:
                self.nvm_rows = read_nvm_map(xlsm)
                self.bld_xlsm_lbl.configure(
                    text=f"{xlsm.name}  ({len(self.nvm_rows)} rows)", fg="#107c10")
            except Exception as e:
                self.nvm_rows = []
                self.bld_xlsm_lbl.configure(text=f"Error: {e}", fg="#cc0000")
        else:
            self.nvm_rows = []
            self.bld_xlsm_lbl.configure(text="No XLSM found in project folder", fg="#cc0000")

    # ═══════════════════════════════════════════════════════════════════════
    # Image Builder logic
    # ═══════════════════════════════════════════════════════════════════════

    def _bld_toggle_inputs(self):
        mode = self.bld_mode_var.get()
        # Just enable/disable visual cue — entries always accessible
        self.bld_v_frame.configure()   # no-op, entries always visible

    def _bld_lookup(self):
        offset = self.bld_offset_var.get().strip()
        bits   = self.bld_bits_var.get().strip()
        if not offset:
            messagebox.showwarning("Lookup", "Enter an offset first.")
            return
        if not self.nvm_rows:
            messagebox.showwarning("Lookup", "Load a project with an XLSM first.")
            return
        matches = lookup_register(self.nvm_rows, offset, bits)
        if not matches:
            self.bld_lookup_lbl.configure(
                text=f"No match found for offset={offset}" + (f" bits={bits}" if bits else ""),
                fg="#cc0000")
            return
        lines = []
        for m in matches[:4]:
            lines.append(f"[{m['offset']}] [{m['bits']}]  {m['cspec'] or m['rtl']}")
            lines.append(f"  V={m['v_val']}   LM={m['lm_val']}")
        self.bld_lookup_lbl.configure(text="\n".join(lines), fg="#107c10")

    def _bld_add_change(self):
        offset = self.bld_offset_var.get().strip()
        bits   = self.bld_bits_var.get().strip()
        mode   = self.bld_mode_var.get()
        new_v  = self.bld_new_v_var.get().strip()
        new_lm = self.bld_new_lm_var.get().strip()
        reason = self.bld_reason_var.get().strip()

        if not offset:
            messagebox.showwarning("Add Change", "Offset is required.")
            return
        if mode in ("V", "Both") and not new_v:
            messagebox.showwarning("Add Change", "New V value is required.")
            return
        if mode in ("LM", "Both") and not new_lm:
            messagebox.showwarning("Add Change", "New LM value is required.")
            return

        # Lookup current values
        old_v = old_lm = field_name = bits_found = "—"
        if self.nvm_rows:
            matches = lookup_register(self.nvm_rows, offset, bits)
            if matches:
                m = matches[0]
                old_v    = m["v_val"]
                old_lm   = m["lm_val"]
                field_name = m["cspec"] or m["rtl"] or "—"
                bits_found = m["bits"]

        if not bits:
            bits = bits_found

        change = {
            "project":    self.bld_project_var.get(),
            "offset":     offset,
            "bits":       bits,
            "field":      field_name,
            "mode":       mode,
            "old_v":      old_v  if mode in ("V",  "Both") else "—",
            "new_v":      new_v  if mode in ("V",  "Both") else "—",
            "old_lm":     old_lm if mode in ("LM", "Both") else "—",
            "new_lm":     new_lm if mode in ("LM", "Both") else "—",
            "reason":     reason,
        }
        self.change_list.append(change)
        self.change_tree.insert("", tk.END, values=(
            offset, bits, field_name,
            change["old_v"], change["new_v"],
            change["old_lm"], change["new_lm"],
            reason,
        ))
        # Clear inputs
        for var in (self.bld_offset_var, self.bld_bits_var,
                    self.bld_new_v_var, self.bld_new_lm_var, self.bld_reason_var):
            var.set("")
        self.bld_lookup_lbl.configure(text="")

    def _bld_remove_selected(self):
        sel = self.change_tree.selection()
        for item in sel:
            idx = self.change_tree.index(item)
            self.change_tree.delete(item)
            if idx < len(self.change_list):
                self.change_list.pop(idx)

    def _bld_generate(self):
        if not self.change_list:
            messagebox.showwarning("Generate", "Add at least one change first.")
            return

        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        proj = self.change_list[0]["project"]
        lines = []
        lines.append(f"GBE NVM Change Report")
        lines.append(f"Generated : {ts}")
        lines.append(f"Project   : {proj}")
        lines.append(f"NVM Root  : {self.nvm_root or 'N/A'}")
        lines.append("=" * 70)
        lines.append("")

        for i, c in enumerate(self.change_list, 1):
            lines.append(f"Change #{i}")
            lines.append(f"  Offset    : {c['offset']}")
            lines.append(f"  Bits      : {c['bits']}")
            lines.append(f"  Field     : {c['field']}")
            lines.append(f"  Mode      : {c['mode']}")
            if c["mode"] in ("V", "Both"):
                lines.append(f"  V  before : {c['old_v']}")
                lines.append(f"  V  after  : {c['new_v']}  <-- CHANGE")
            if c["mode"] in ("LM", "Both"):
                lines.append(f"  LM before : {c['old_lm']}")
                lines.append(f"  LM after  : {c['new_lm']}  <-- CHANGE")
            if c["reason"]:
                lines.append(f"  Reason    : {c['reason']}")
            lines.append("")

        lines.append("=" * 70)
        lines.append("DIFF SUMMARY")
        lines.append("-" * 70)
        lines.append(f"{'Offset':<8} {'Bits':<8} {'Field':<30} {'Old V':<10} {'New V':<10} {'Old LM':<10} {'New LM':<10}")
        lines.append("-" * 70)
        for c in self.change_list:
            lines.append(f"{c['offset']:<8} {c['bits']:<8} {c['field'][:29]:<30} "
                         f"{c['old_v']:<10} {c['new_v']:<10} {c['old_lm']:<10} {c['new_lm']:<10}")

        report_txt = "\n".join(lines)

        # Display
        self.report_text.delete("1.0", tk.END)
        self.report_text.insert(tk.END, report_txt)

        # Save files
        out_dir = Path(filedialog.askdirectory(title="Save report + JSON to folder"))
        if not str(out_dir).strip():
            return

        ts_file = datetime.now().strftime("%Y%m%d_%H%M%S")
        txt_path = out_dir / f"NVM_Change_Report_{proj}_{ts_file}.txt"
        json_path = out_dir / f"change_request_{proj}_{ts_file}.json"

        txt_path.write_text(report_txt, encoding="utf-8")
        json_path.write_text(json.dumps({
            "generated": ts,
            "project": proj,
            "changes": self.change_list,
        }, indent=2), encoding="utf-8")

        messagebox.showinfo("Saved",
            f"Report saved:\n{txt_path.name}\n{json_path.name}\n\nIn: {out_dir}")

    def _bld_clear_all(self):
        if self.change_list and not messagebox.askyesno("Clear", "Clear all changes?"):
            return
        self.change_list.clear()
        for item in self.change_tree.get_children():
            self.change_tree.delete(item)
        self.report_text.delete("1.0", tk.END)
        self.bld_lookup_lbl.configure(text="")

    # ═══════════════════════════════════════════════════════════════════════
    # AI Assistant logic
    # ═══════════════════════════════════════════════════════════════════════

    def _try_init_client(self):
        key = self.config_data.get("api_key", "").strip()
        if key:
            self.client = anthropic.Anthropic(api_key=key)
            self.api_status_lbl.configure(text="API: connected", fg="#107c10")
        else:
            self.api_status_lbl.configure(text="API: no key — click Settings", fg="#cc0000")
            self._append_chat("system", "No API key. Click Settings in the status bar.")

    def _build_ai_prompt(self):
        project = self.ai_project_var.get()
        offset  = self.ai_offset_var.get().strip()
        bits    = self.ai_bits_var.get().strip()
        name    = self.ai_name_var.get().strip()
        mode    = self.ai_mode_var.get()
        value   = self.ai_value_var.get().strip()
        lines = [f"Project: {project}"]
        if offset:
            lines.append(f"Offset: 0x{offset.upper().lstrip('0X')}")
        if bits:
            lines.append(f"Bits: [{bits}]")
        if name:
            lines.append(f"Register: {name}")
        lines.append(f"Mode: {mode}")
        lines.append(f"Request: {value or 'Explain and recommend configuration.'}")
        return "\n".join(lines)

    def _ai_ask(self):
        if not self._check_client():
            return
        prompt = self._build_ai_prompt()
        self.conversation.clear()
        self.chat_display.configure(state=tk.NORMAL)
        self.chat_display.delete("1.0", tk.END)
        self.chat_display.configure(state=tk.DISABLED)
        self._submit(prompt)

    def _send_followup(self):
        if not self._check_client():
            return
        msg = self.input_text.get("1.0", tk.END).strip()
        if not msg:
            return
        self.input_text.delete("1.0", tk.END)
        self._submit(msg)

    def _submit(self, user_msg):
        self._append_chat("user", user_msg)
        self.conversation.append({"role": "user", "content": user_msg})
        self.status_var.set("Thinking...")
        self.update_idletasks()
        threading.Thread(target=self._call_api, daemon=True).start()

    def _call_api(self):
        try:
            system = SYSTEM_PROMPT
            if self.ai_excel_context:
                system += f"\n\n--- NVM MAP ({self.ai_project_var.get()}) ---\n{self.ai_excel_context[:10000]}\n---"
            resp = self.client.messages.create(
                model=self.config_data.get("model", MODEL),
                max_tokens=MAX_TOKENS,
                system=system,
                messages=self.conversation,
            )
            reply = resp.content[0].text
            self.conversation.append({"role": "assistant", "content": reply})
            self.after(0, lambda: self._append_chat("assistant", reply))
            self.after(0, lambda: self.status_var.set("Done — ask a follow-up or change the form"))
        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: self._append_chat("error", f"API Error: {err}"))
            self.after(0, lambda: self.status_var.set("Error — see chat"))

    def _check_client(self):
        if not self.client:
            messagebox.showwarning("No API Key", "Add your Anthropic API key in Settings.")
            self._open_settings()
            return False
        return True

    def _append_chat(self, role, text):
        self.chat_display.configure(state=tk.NORMAL)
        labels = {"user": "\nYou:\n", "assistant": "\nClaude:\n", "system": "\n", "error": "\nError: "}
        self.chat_display.insert(tk.END, labels.get(role, "\n"), role)
        self.chat_display.insert(tk.END, text + "\n", role)
        self.chat_display.insert(tk.END, "-" * 60 + "\n", "system")
        self.chat_display.configure(state=tk.DISABLED)
        self.chat_display.see(tk.END)

    def _clear_chat(self):
        if messagebox.askyesno("Clear Chat", "Clear conversation?"):
            self.conversation.clear()
            self.chat_display.configure(state=tk.NORMAL)
            self.chat_display.delete("1.0", tk.END)
            self.chat_display.configure(state=tk.DISABLED)

    def _save_chat(self):
        path = filedialog.asksaveasfilename(defaultextension=".txt",
                                            filetypes=[("Text", "*.txt")])
        if path:
            lines = [f"[{'You' if m['role']=='user' else 'Claude'}]\n{m['content']}\n"
                     for m in self.conversation]
            Path(path).write_text("\n".join(lines), encoding="utf-8")

    # ═══════════════════════════════════════════════════════════════════════
    # Settings
    # ═══════════════════════════════════════════════════════════════════════

    def _open_settings(self):
        dlg = tk.Toplevel(self)
        dlg.title("Settings")
        dlg.geometry("480x210")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.columnconfigure(1, weight=1)
        ttk.Label(dlg, text="Anthropic API Key:").grid(row=0, column=0, sticky="w", padx=16, pady=(18, 4))
        key_var = tk.StringVar(value=self.config_data.get("api_key", ""))
        key_entry = ttk.Entry(dlg, textvariable=key_var, width=42, show="*")
        key_entry.grid(row=0, column=1, padx=(4, 16), sticky="ew", pady=(18, 4))
        show_var = tk.BooleanVar()
        ttk.Checkbutton(dlg, text="Show", variable=show_var,
                        command=lambda: key_entry.configure(show="" if show_var.get() else "*")).grid(
            row=1, column=1, sticky="w", padx=4)
        ttk.Label(dlg, text="Model:").grid(row=2, column=0, sticky="w", padx=16, pady=(10, 4))
        model_var = tk.StringVar(value=self.config_data.get("model", MODEL))
        ttk.Combobox(dlg, textvariable=model_var, width=32,
                     values=["claude-opus-4-5", "claude-sonnet-4-5", "claude-haiku-4-5"]).grid(
            row=2, column=1, sticky="w", padx=(4, 16), pady=(10, 4))
        ttk.Label(dlg, text="console.anthropic.com",
                  foreground="#0071c5", font=("Segoe UI", 8)).grid(
            row=3, column=0, columnspan=2, sticky="w", padx=16)

        def _save():
            self.config_data["api_key"] = key_var.get().strip()
            self.config_data["model"]   = model_var.get()
            save_config(self.config_data)
            self._try_init_client()
            dlg.destroy()

        btns = ttk.Frame(dlg)
        btns.grid(row=4, column=0, columnspan=2, sticky="e", padx=16, pady=14)
        ttk.Button(btns, text="Cancel", command=dlg.destroy).pack(side=tk.RIGHT, padx=4)
        ttk.Button(btns, text="Save",   command=_save).pack(side=tk.RIGHT)


if __name__ == "__main__":
    NvmApp().mainloop()
