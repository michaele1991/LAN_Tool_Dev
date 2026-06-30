"""
GBE NVM Build Wizard — interactive 5-question CLI
==================================================
Run this directly: python wizard.py
The AI (GitHub Copilot) will invoke this automatically when asked
to "build a GBE image" or "create an NVM image".

Questions asked:
  Q1  Platform     — which silicon/NVM map
  Q2  Step         — silicon stepping (e.g. A0)
  Q3  Version      — image version string (e.g. 1.4)
  Q4  Variant      — V / LM / Both
  Q5  NVM changes  — bit-level overrides on top of the base map
  Q6  Confirm      — show full summary and build

Requirements: openpyxl  (pip install openpyxl)
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from build_nvm import (
    _find_platforms_root, list_platforms, find_xlsm, read_general_variable, build,
    VAR_SUFFIX, COL_OFFSET, COL_BITS, COL_RTL, COL_CSPEC, COL_V, COL_LM, DATA_ROW,
    NVM_SHEET, _parse_offset,
)
from openpyxl import load_workbook


# ── helpers ────────────────────────────────────────────────────────────────

def ask(prompt: str, default: str = "") -> str:
    full_prompt = f"{prompt} [{default}]: " if default else f"{prompt}: "
    try:
        answer = input(full_prompt).strip()
    except (EOFError, KeyboardInterrupt):
        print()
        sys.exit(0)
    return answer if answer else default


def choose(prompt: str, options: list[str], default_idx: int = 0) -> int:
    print(f"\n{prompt}")
    for i, opt in enumerate(options):
        marker = " *" if i == default_idx else ""
        print(f"  [{i}] {opt}{marker}")
    while True:
        raw = ask("  Enter number", str(default_idx))
        try:
            idx = int(raw)
            if 0 <= idx < len(options):
                return idx
        except ValueError:
            pass
        print(f"  Please enter a number between 0 and {len(options)-1}.")


def lookup_offset_bits(xlsm_path: Path, offset: int) -> list[dict]:
    """Return all bit-field rows at a given NVM offset."""
    wb = load_workbook(str(xlsm_path), read_only=True, data_only=True, keep_vba=False)
    rows = []
    if NVM_SHEET in wb.sheetnames:
        ws = wb[NVM_SHEET]
        for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
            if row[COL_OFFSET] is None:
                continue
            if _parse_offset(row[COL_OFFSET]) == offset:
                rows.append({
                    "bits": str(row[COL_BITS]  or "").strip(),
                    "name": str(row[COL_CSPEC] or row[COL_RTL] or "").strip(),
                    "v":    str(row[COL_V]     or "").strip(),
                    "lm":   str(row[COL_LM]    or "").strip(),
                })
    wb.close()
    return rows


def load_all_offsets(xlsm_path: Path) -> list[dict]:
    """
    Return an ordered list of all unique NVM offsets from the xlsm.
    Each entry: {offset_int, offset_hex, name, bits_list}
    where bits_list = [{bits, name, v, lm}, ...]
    """
    wb = load_workbook(str(xlsm_path), read_only=True, data_only=True, keep_vba=False)
    ordered: dict[int, dict] = {}
    if NVM_SHEET in wb.sheetnames:
        ws = wb[NVM_SHEET]
        for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
            if row[COL_OFFSET] is None:
                continue
            off = _parse_offset(row[COL_OFFSET])
            if off is None:
                continue
            if off not in ordered:
                name = str(row[COL_CSPEC] or row[COL_RTL] or "").strip()
                ordered[off] = {"offset": off, "name": name, "bits_list": []}
            ordered[off]["bits_list"].append({
                "bits": str(row[COL_BITS]  or "").strip(),
                "name": str(row[COL_CSPEC] or row[COL_RTL] or "").strip(),
                "v":    str(row[COL_V]     or "").strip(),
                "lm":   str(row[COL_LM]    or "").strip(),
            })
    wb.close()
    return [v for v in sorted(ordered.values(), key=lambda x: x["offset"])]


def find_register(all_offsets: list[dict], query: str) -> list[dict]:
    """
    Find registers by name (case-insensitive partial match), hex offset, or list index.
    Returns list of matching entries (usually 1, could be >1 for ambiguous names).
    """
    q = query.strip().lower()

    # Try list index
    try:
        idx = int(q)
        if 0 <= idx < len(all_offsets):
            return [all_offsets[idx]]
        return []
    except ValueError:
        pass

    # Try hex offset (0x58 or 58h or 58)
    try:
        off = int(q, 0) if q.startswith("0x") else int(q.rstrip("h"), 16)
        match = [r for r in all_offsets if r["offset"] == off]
        if match:
            return match
    except ValueError:
        pass

    # Partial name match (search across ALL bit-field names within each register)
    results = []
    for reg in all_offsets:
        # Check the register-level name or any individual bit-field name
        all_names = [reg["name"]] + [b["name"] for b in reg["bits_list"]]
        if any(q in n.lower() for n in all_names):
            results.append(reg)
    return results


def collect_nvm_changes(xlsm_path: Path) -> list[dict]:
    """
    Interactively collect bit-level NVM overrides.
    User inputs register name, hex offset, or list number — then bit + value.
    Returns list of: {offset, bit, new_value, variants}
    """
    changes = []
    all_offsets = load_all_offsets(xlsm_path)

    print("\n── Q5 / 5 ── NVM Modifications")
    print("  Specify changes by register name, hex offset, or list number.")
    print("  Examples:  'FEXTNVM12'  |  '0x58'  |  '87'  |  'Device ID'")
    print("  Type 'list' to show all registers, or press Enter to skip.\n")

    while True:
        raw = ask("  Register name / offset (or 'done')").strip()
        if not raw or raw.lower() in ("done", "none", "skip"):
            break

        if raw.lower() == "list":
            print(f"\n  {'#':>4}  {'Offset':6}  Name")
            print("  " + "-" * 55)
            for i, reg in enumerate(all_offsets):
                print(f"  {i:>4}  0x{reg['offset']:02X}     {reg['name'][:48]}")
            print()
            continue

        matches = find_register(all_offsets, raw)

        if not matches:
            print(f"  No register found matching '{raw}'. Try 'list' to browse all.")
            continue

        if len(matches) > 1:
            print(f"  Multiple matches for '{raw}':")
            for i, m in enumerate(matches):
                print(f"    [{i}]  0x{m['offset']:02X}  {m['name']}")
            sel = ask("  Which one?", "0")
            try:
                matches = [matches[int(sel)]]
            except (ValueError, IndexError):
                print("  Invalid selection.")
                continue

        reg_entry = matches[0]
        offset    = reg_entry["offset"]
        bit_rows  = reg_entry["bits_list"]

        print(f"\n  Register 0x{offset:02X} — {reg_entry['name']}:")
        print(f"  {'Bit(s)':8}  {'Current V':10}  {'Current LM':10}")
        print("  " + "-" * 32)
        for r in bit_rows:
            print(f"  {r['bits']:8}  {r['v']:10}  {r['lm']:10}")
        print()

        bit_raw = ask("  Which bit? (0–15)").strip()
        try:
            bit = int(bit_raw)
            if not (0 <= bit <= 15):
                raise ValueError
        except ValueError:
            print("  Must be a number 0–15.")
            continue

        cur_v  = next((r["v"]  for r in bit_rows if r["bits"].strip() == str(bit)), "?")
        cur_lm = next((r["lm"] for r in bit_rows if r["bits"].strip() == str(bit)), "?")

        new_raw = ask(f"  New value for bit {bit}? (0 or 1)").strip()
        try:
            new_val = int(new_raw)
            if new_val not in (0, 1):
                raise ValueError
        except ValueError:
            print("  Must be 0 or 1.")
            continue

        if cur_v == str(new_val) and cur_lm == str(new_val):
            print(f"  ⚠  Bit {bit} is already {new_val} in both V and LM — no change.")

        var_idx  = choose("  Apply to:",
                          ["Both  (V + LM)", "V only  (Consumer)", "LM only  (Corporate)"],
                          default_idx=0)
        variants = ["V", "LM"] if var_idx == 0 else (["V"] if var_idx == 1 else ["LM"])

        changes.append({"offset": offset, "bit": bit, "new_value": new_val,
                         "variants": variants})
        print(f"  ✓  0x{offset:02X} bit {bit} → {new_val}  ({'+'.join(variants)})")

        more = ask("\n  Any more changes? (y/n)", "n").lower()
        if more not in ("y", "yes"):
            break

    return changes


# ── wizard ─────────────────────────────────────────────────────────────────

def run_wizard():
    print("=" * 60)
    print("  GBE NVM Image Build Wizard")
    print("=" * 60)

    # Locate platforms root
    root = _find_platforms_root()
    if root is None:
        print("\nERROR: Cannot locate GBE_Image platforms folder.")
        print("Make sure you are running from inside the LAN_Tool_Dev repo.")
        sys.exit(1)

    platforms = list_platforms(root)
    if not platforms:
        print(f"\nERROR: No platform folders (with .xlsm) found under:\n  {root}")
        sys.exit(1)

    # ── Q1: Platform ───────────────────────────────────────────────────────
    print("\n── Q1 / 5 ── Platform")
    platform_names = [p.name for p in platforms]
    # Pre-select first platform alphabetically as default
    default_plat = next(
        (i for i, n in enumerate(platform_names) if "ptl_pcd_p_h" in n.lower()), 0
    )
    plat_idx = choose("Which platform / silicon project?",
                      platform_names, default_idx=default_plat)
    platform_folder = platforms[plat_idx]

    # Read defaults from xlsm general variable sheet
    try:
        xlsm = find_xlsm(platform_folder)
        info = read_general_variable(xlsm)
    except Exception:
        info = {}

    default_step = info.get("step", "A0")
    default_ver  = "1.4"

    # ── Q2: Step ───────────────────────────────────────────────────────────
    print("\n── Q2 / 5 ── Silicon Stepping")
    step = ask("Silicon step (e.g. A0, B0, C0)", default_step).upper().strip()
    if not step:
        step = default_step

    # ── Q3: Version ────────────────────────────────────────────────────────
    print("\n── Q3 / 5 ── Image Version")
    version = ask("Image version (e.g. 1.4, 2.0)", default_ver)
    # Validate format
    parts = version.split(".")
    if len(parts) != 2:
        print(f"  Using default version {default_ver!r}.")
        version = default_ver

    # ── Q4: Variant ────────────────────────────────────────────────────────
    print("\n── Q4 / 5 ── Build Variant")
    var_idx  = choose("Which variant(s) to build?", [
        "Both  — V (Consumer/LAN SW) + LM (Corporate/Non-LAN SW)",
        "V     — Consumer / LAN SW only",
        "LM    — Corporate / Non-LAN SW only",
    ], default_idx=0)
    variants = ["V", "LM"] if var_idx == 0 else (["V"] if var_idx == 1 else ["LM"])

    # ── Q5: NVM modifications ──────────────────────────────────────────────
    nvm_changes = collect_nvm_changes(xlsm)

    # ── Confirm ────────────────────────────────────────────────────────────
    output_dir = Path(__file__).resolve().parent.parent / "output" / platform_folder.name
    base_name  = (info.get("image file name") or info.get("Project name")
                  or platform_folder.name)
    major, minor = (version.split(".", 1) + ["0"])[:2]
    ver_str = f"{major}.{minor}"

    print("\n── Confirm Build ─────────────────────────────────────────")
    print(f"  Platform  : {platform_folder.name}")
    print(f"  XLSM      : {xlsm.name}")
    print(f"  Step      : {step}")
    print(f"  Version   : {ver_str}")
    print(f"  Variant(s): {' + '.join(variants)}")
    if nvm_changes:
        print(f"  NVM changes ({len(nvm_changes)}):")
        for c in nvm_changes:
            print(f"    0x{c['offset']:02X}  bit {c['bit']}  →  {c['new_value']}"
                  f"  ({'+'.join(c['variants'])})")
    else:
        print("  NVM changes: none (base map only)")
    print(f"  Output    : {output_dir}")
    print()
    for v in variants:
        fname = f"{base_name}_{step}_{ver_str}_Release_{VAR_SUFFIX[v]}"
        print(f"    {fname}.bin")
    print()
    confirm = ask("Build now? [Y/n]", "Y").upper()
    if confirm not in ("Y", "YES", ""):
        print("\nBuild cancelled.")
        sys.exit(0)

    # ── Run build ──────────────────────────────────────────────────────────
    print()
    try:
        outputs = build(platform_folder, step, version, variants, output_dir,
                        nvm_changes=nvm_changes)
    except Exception as exc:
        print(f"\nERROR during build: {exc}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    print(f"\n{'='*60}")
    print("  BUILD COMPLETE")
    print(f"  Output: {output_dir}")
    for p in outputs:
        print(f"    {p.name}  ({p.stat().st_size} bytes)")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    run_wizard()
