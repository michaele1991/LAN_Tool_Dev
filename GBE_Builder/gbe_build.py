"""
GBE Image Builder
-----------------
Interactive CLI: asks project / offset / bit / mode / value,
patches the NVM binary directly, recalculates checksum (word 0x3F = 0xBABA),
saves a change_request.json and a diff report.

Requirements: openpyxl   (pip install openpyxl)
NVM root folder name  : Eng_GBE_Image  (hardcoded, searched upward from this script)
"""

import json
import struct
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# ── auto-install openpyxl if missing ──────────────────────────────────────
try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet",
                           "--index-url", "https://pypi.org/simple/", "openpyxl"])
    from openpyxl import load_workbook

NVM_ROOT = "Eng_GBE_Image"         # primary name on target PCs
NVM_ROOT_FALLBACKS = ["GBE_Image"]  # fallback names for dev machines

# XLSM 'full nvm map' columns (0-based after iter_rows)
COL_OFFSET = 0   # A
COL_BITS   = 1   # B
COL_RTL    = 2   # C
COL_CSPEC  = 3   # D
COL_DESC   = 5   # F
COL_V      = 6   # G
COL_LM     = 7   # H


# ── helpers ────────────────────────────────────────────────────────────────

def find_nvm_root() -> Path | None:
    here = Path(__file__).resolve().parent
    search_names = [NVM_ROOT] + NVM_ROOT_FALLBACKS
    for p in [here, here.parent, here.parent.parent, here.parent.parent.parent]:
        for name in search_names:
            candidate = p / name
            if candidate.is_dir():
                return candidate
    return None


def norm_offset(s: str) -> str:
    s = s.strip().lower().lstrip("0x")
    return s.rstrip("h") + "h"


def read_nvm_rows(xlsm: Path) -> list[dict]:
    wb = load_workbook(str(xlsm), read_only=True, data_only=True, keep_vba=False)
    rows = []
    if "full nvm map" in wb.sheetnames:
        ws = wb["full nvm map"]
        for row in ws.iter_rows(min_row=7, values_only=True):
            if row[COL_OFFSET] is None:
                continue
            rows.append({
                "offset": str(row[COL_OFFSET]).strip(),
                "bits":   str(row[COL_BITS]).strip()  if row[COL_BITS]  else "",
                "rtl":    str(row[COL_RTL]).strip()   if row[COL_RTL]   else "",
                "cspec":  str(row[COL_CSPEC]).strip() if row[COL_CSPEC] else "",
                "desc":   str(row[COL_DESC]).strip()  if row[COL_DESC]  else "",
                "v_val":  str(row[COL_V]).strip()     if row[COL_V]     else "N/A",
                "lm_val": str(row[COL_LM]).strip()    if row[COL_LM]    else "N/A",
            })
    wb.close()
    return rows


def lookup(rows: list[dict], offset_raw: str, bits_raw: str = "") -> list[dict]:
    target = norm_offset(offset_raw)
    return [r for r in rows
            if norm_offset(r["offset"]) == target
            and (not bits_raw or str(r["bits"]).strip() == bits_raw.strip())]


def patch_bin(bin_path: Path, offset: int, bit: int, new_bit: int) -> tuple[int, int, int]:
    """Set/clear `bit` in NVM word at `offset`. Returns (old_word, new_word, checksum)."""
    data = bytearray(bin_path.read_bytes())
    byte_pos = offset * 2

    old_word = struct.unpack_from("<H", data, byte_pos)[0]
    if new_bit:
        new_word = old_word | (1 << bit)
    else:
        new_word = old_word & ~(1 << bit)
    struct.pack_into("<H", data, byte_pos, new_word)

    # Recalculate checksum at 0x3F
    csum_pos = 0x3F * 2
    struct.pack_into("<H", data, csum_pos, 0)
    s = sum(struct.unpack_from("<H", data, i)[0] for i in range(0, csum_pos, 2))
    csum = (0xBABA - s) & 0xFFFF
    struct.pack_into("<H", data, csum_pos, csum)

    bin_path.write_bytes(data)
    return old_word, new_word, csum


def patch_txt(txt_path: Path, offset: int, new_word: int, csum: int):
    """Update the hex word in the .txt image file."""
    lines = txt_path.read_text().splitlines()
    word_idx = 0
    out = []
    for line in lines:
        parts = line.split()
        new_parts = []
        for w in parts:
            if word_idx == offset:
                new_parts.append(f"{new_word:04X}")
            elif word_idx == 0x3F:
                new_parts.append(f"{csum:04X}")
            else:
                new_parts.append(w)
            word_idx += 1
        out.append(" ".join(new_parts))
    txt_path.write_text("\n".join(out))


def ask(prompt: str, options: list[str] | None = None) -> str:
    while True:
        if options:
            for i, o in enumerate(options, 1):
                print(f"  {i}. {o}")
        val = input(f"{prompt}: ").strip()
        if options:
            if val.isdigit() and 1 <= int(val) <= len(options):
                return options[int(val) - 1]
            # also accept text match
            matches = [o for o in options if o.lower().startswith(val.lower())]
            if len(matches) == 1:
                return matches[0]
            print("  -> enter a number or unambiguous prefix")
            continue
        if val:
            return val
        print("  -> required")


def hr(n=60):
    print("─" * n)


# ── main ───────────────────────────────────────────────────────────────────

def main():
    print()
    print("=" * 60)
    print("  Intel GBE Image Builder")
    print("=" * 60)

    nvm_root = find_nvm_root()
    if not nvm_root:
        print(f"\nERROR: '{NVM_ROOT}' folder not found near this script.")
        print("Place gbe_build.py inside or next to the Eng_GBE_Image folder.")
        sys.exit(1)

    print(f"\nNVM root : {nvm_root}")

    # ── 1. Project ────────────────────────────────────────────────────────
    projects = sorted(d.name for d in nvm_root.iterdir() if d.is_dir())
    print("\nQ1: Which project?")
    project_name = ask("   Enter number or name prefix", projects)
    project_dir  = nvm_root / project_name

    xlsm_files = list(project_dir.glob("*.xlsm"))
    if not xlsm_files:
        print(f"ERROR: no XLSM found in {project_dir}")
        sys.exit(1)
    xlsm = xlsm_files[0]
    print(f"   -> {project_name}  ({xlsm.name})")

    print("\nLoading NVM map...", end=" ", flush=True)
    nvm_rows = read_nvm_rows(xlsm)
    print(f"{len(nvm_rows)} rows loaded.")

    changes = []

    while True:
        hr()
        # ── 2. Offset ─────────────────────────────────────────────────────
        offset_str = ask("\nQ2: NVM word offset (hex, e.g. 0x59)")
        try:
            offset_int = int(offset_str.lower().lstrip("0x").rstrip("h"), 16)
        except ValueError:
            print("  -> invalid hex"); continue

        # Show all bits at this offset
        matches_all = lookup(nvm_rows, offset_str)
        if not matches_all:
            print(f"   Offset {offset_str} not found in XLSM (proceeding anyway)")
        else:
            print(f"\n   Offset {offset_str} — {matches_all[0]['cspec'] or matches_all[0]['rtl']}:")
            print(f"   {'Bit':<6} {'RTL name':<45} {'V':>8}  {'LM':>8}")
            print(f"   {'-'*6} {'-'*45} {'-'*8}  {'-'*8}")
            for r in matches_all:
                print(f"   {r['bits']:<6} {r['rtl'][:44]:<45} {r['v_val']:>8}  {r['lm_val']:>8}")

        # ── 3. Bit ────────────────────────────────────────────────────────
        bit_str  = ask("\nQ3: Which bit")
        try:
            bit_int = int(bit_str)
        except ValueError:
            print("  -> invalid"); continue

        match = next((r for r in matches_all if str(r["bits"]).strip() == bit_str.strip()), None)
        if match:
            print(f"\n   [{offset_str} bit {bit_int}]  {match['rtl']}")
            print(f"   {match['desc'][:120]}")
            print(f"   Current  V={match['v_val']}   LM={match['lm_val']}")
        else:
            print(f"   (bit {bit_int} not in XLSM map)")
            match = {"v_val": "?", "lm_val": "?", "rtl": "", "cspec": ""}

        # ── 4. Mode ───────────────────────────────────────────────────────
        mode = ask("\nQ4: V, LM, or Both", ["V", "LM", "Both"])

        # ── 5. New value ─────────────────────────────────────────────────
        new_val_str = ask("Q5: New value for this bit (0 or 1)")
        try:
            new_val = int(new_val_str)
            assert new_val in (0, 1)
        except Exception:
            print("  -> must be 0 or 1"); continue

        # Warn if same value
        current_v  = match["v_val"]
        current_lm = match["lm_val"]
        no_change_v  = (mode in ("V",  "Both") and current_v  not in ("?",) and current_v  == str(new_val))
        no_change_lm = (mode in ("LM", "Both") and current_lm not in ("?",) and current_lm == str(new_val))
        if no_change_v and mode == "V":
            print(f"   WARNING: V is already {current_v} — no change")
        if no_change_lm and mode == "LM":
            print(f"   WARNING: LM is already {current_lm} — no change")
        if no_change_v and no_change_lm:
            print("   WARNING: both values unchanged — skipping this entry")
            if ask("Add anyway? (y/n)", ["y", "n"]) == "n":
                pass
            else:
                pass

        changes.append({
            "offset":     offset_str,
            "offset_int": offset_int,
            "bit":        bit_int,
            "field":      match["rtl"] or match["cspec"],
            "mode":       mode,
            "old_v":      match["v_val"],
            "new_v":      str(new_val) if mode in ("V",  "Both") else "—",
            "old_lm":     match["lm_val"],
            "new_lm":     str(new_val) if mode in ("LM", "Both") else "—",
        })
        print(f"\n   Added: [{offset_str} bit {bit_int}]  {mode}  -> {new_val}")

        # ── Done? ─────────────────────────────────────────────────────────
        more = ask("\nAdd another change? (y/n)", ["y", "n"])
        if more == "n":
            break

    if not changes:
        print("\nNo changes — exiting.")
        sys.exit(0)

    # ── Summary before build ──────────────────────────────────────────────
    hr(60)
    print("\nCHANGE SUMMARY")
    hr(60)
    print(f"{'Offset':<8} {'Bit':<5} {'Field':<40} {'Old V':<8} {'New V':<8} {'Old LM':<8} {'New LM':<8}")
    print(f"{'-'*8} {'-'*5} {'-'*40} {'-'*8} {'-'*8} {'-'*8} {'-'*8}")
    for c in changes:
        print(f"{c['offset']:<8} {c['bit']:<5} {c['field'][:39]:<40} "
              f"{c['old_v']:<8} {c['new_v']:<8} {c['old_lm']:<8} {c['new_lm']:<8}")

    confirm = ask("\nBuild now? (y/n)", ["y", "n"])
    if confirm == "n":
        print("Aborted.")
        sys.exit(0)

    # ── Find output dirs ──────────────────────────────────────────────────
    out_dirs = [d for d in project_dir.iterdir() if d.is_dir()]
    cons_dirs = [d for d in out_dirs if any(x in d.name.lower() for x in ("cons", "v_", "_v_", "lan"))]
    corp_dirs = [d for d in out_dirs if any(x in d.name.lower() for x in ("corp", "lm", "non"))]
    # fallback: first two dirs
    if not cons_dirs:
        cons_dirs = out_dirs[:1]
    if not corp_dirs:
        corp_dirs = out_dirs[1:2] if len(out_dirs) > 1 else out_dirs[:1]

    print(f"\n  V  image dir : {cons_dirs[0].name if cons_dirs else 'not found'}")
    print(f"  LM image dir : {corp_dirs[0].name if corp_dirs else 'not found'}")

    results = []

    for change in changes:
        off  = change["offset_int"]
        bit  = change["bit"]
        mode = change["mode"]
        val  = int(change["new_v"]) if change["new_v"] != "—" else None

        # V image
        if mode in ("V", "Both") and cons_dirs:
            d = cons_dirs[0]
            bins = list(d.glob("*.bin"))
            txts = list(d.glob("*.txt"))
            if bins:
                shutil.copy2(bins[0], bins[0].with_suffix(".bin.bak"))
                old_w, new_w, csum = patch_bin(bins[0], off, bit, val)
                if txts:
                    patch_txt(txts[0], off, new_w, csum)
                results.append(("V", bins[0].name, off, bit, old_w, new_w, csum))
                print(f"\n  [V]  {bins[0].name}")
                print(f"       0x{off:02X}: 0x{old_w:04X} -> 0x{new_w:04X}  bit{bit}={val}  csum=0x{csum:04X}")

        # LM image
        lm_val = int(change["new_lm"]) if change["new_lm"] != "—" else None
        if mode in ("LM", "Both") and corp_dirs:
            d = corp_dirs[0]
            bins = list(d.glob("*.bin"))
            txts = list(d.glob("*.txt"))
            if bins:
                shutil.copy2(bins[0], bins[0].with_suffix(".bin.bak"))
                old_w, new_w, csum = patch_bin(bins[0], off, bit, lm_val)
                if txts:
                    patch_txt(txts[0], off, new_w, csum)
                results.append(("LM", bins[0].name, off, bit, old_w, new_w, csum))
                print(f"\n  [LM] {bins[0].name}")
                print(f"       0x{off:02X}: 0x{old_w:04X} -> 0x{new_w:04X}  bit{bit}={lm_val}  csum=0x{csum:04X}")

    # ── Save change_request.json ──────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = project_dir / f"change_request_{ts}.json"
    json_path.write_text(json.dumps({
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "project":   project_name,
        "changes":   changes,
    }, indent=2))

    # ── Save diff report ──────────────────────────────────────────────────
    report_lines = [
        "GBE NVM Change Report",
        f"Generated : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Project   : {project_name}",
        "=" * 60,
    ]
    for i, c in enumerate(changes, 1):
        report_lines += [
            f"\nChange #{i}",
            f"  Offset : {c['offset']}",
            f"  Bit    : {c['bit']}",
            f"  Field  : {c['field']}",
            f"  Mode   : {c['mode']}",
        ]
        if c["mode"] in ("V", "Both"):
            report_lines.append(f"  V   : {c['old_v']} -> {c['new_v']}  <-- CHANGE")
        if c["mode"] in ("LM", "Both"):
            report_lines.append(f"  LM  : {c['old_lm']} -> {c['new_lm']}  <-- CHANGE")

    report_lines += [
        "\n" + "=" * 60,
        "BINARY PATCH RESULTS",
        "-" * 60,
    ]
    for img, name, off, bit, old_w, new_w, csum in results:
        report_lines.append(
            f"[{img}] {name}  0x{off:02X}[{bit}]: 0x{old_w:04X}->0x{new_w:04X}  csum=0x{csum:04X}"
        )

    rpt_path = project_dir / f"change_report_{ts}.txt"
    rpt_path.write_text("\n".join(report_lines))

    hr(60)
    print(f"\nDone.")
    print(f"  Report : {rpt_path.name}")
    print(f"  JSON   : {json_path.name}")
    print(f"  Location: {project_dir}")
    hr(60)


if __name__ == "__main__":
    main()
