import json
import os
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

CHECKSUM_WORD = 0x3F
CHECKSUM_VALUE = 0xBABA
CHECKSUM_LINE = 7


@dataclass
class RegisterDef:
    name: str
    word_index: int
    description: str = ""
    rtl_name: str = ""


class GbeImageEditor(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Intel® GBE NVM Configuration Studio  v2.1")
        self.geometry("1280x820")
        self.minsize(1100, 650)
        self.status_var = tk.StringVar(value="Ready")
        
        # Modern color scheme
        self.colors = {
            'bg': '#f0f0f0',
            'header_bg': '#0071c5',  # Intel blue
            'header_fg': 'white',
            'param_bg': '#ffffff',
            'accent': '#00c7fd',  # Light blue accent
            'border': '#d0d0d0'
        }
        self.configure(bg=self.colors['bg'])

        self.current_folder: Path | None = None
        self.bin_files: list[Path] = []
        self.current_bin: Path | None = None
        self.current_txt: Path | None = None
        self.current_xlsm: Path | None = None
        self.words: list[int] = [0xFFFF] * 128
        self.words_v: list[int] = []
        self.words_lm: list[int] = []
        self.dirty_bits_v: dict = {}   # {word_idx: (set_mask, clear_mask)}
        self.dirty_bits_lm: dict = {}
        self.registers: list[RegisterDef] = []
        self.gbe_folders: list[Path] = []

        self._build_ui()
        self._load_gbe_folders()
        self.update_notes_with_parameters()

    def _build_ui(self) -> None:
        # Create custom style
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Header.TLabel', font=('Segoe UI', 11, 'bold'), foreground='#0071c5')
        style.configure('Title.TLabel', font=('Segoe UI', 14, 'bold'), foreground='#0071c5')
        style.configure('Param.TLabel', font=('Segoe UI', 9), foreground='#333333')
        style.configure('Accent.TButton', font=('Segoe UI', 9, 'bold'))
        style.configure('Build.TButton', font=('Segoe UI', 9, 'bold'))
        style.map('Build.TButton',
                  foreground=[('!active', 'white'), ('active', 'white'), ('pressed', 'white')],
                  background=[('!active', '#107c10'), ('active', '#0e6b0e'), ('pressed', '#0a5a0a')])

        # ── Menu bar ──────────────────────────────────────────────────────────
        menubar = tk.Menu(self, font=('Segoe UI', 9))
        self.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0, font=('Segoe UI', 9))
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Load Excel", accelerator="Ctrl+L", command=self.load_from_excel)
        file_menu.add_command(label="Build NVM", accelerator="Ctrl+B", command=self.run_build_flow)
        file_menu.add_separator()
        file_menu.add_command(label="Open Output Folder", command=self._open_output_folder)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit)

        build_menu = tk.Menu(menubar, tearoff=0, font=('Segoe UI', 9))
        menubar.add_cascade(label="Build", menu=build_menu)
        build_menu.add_command(label="Sync Edits to Build", command=self.sync_all_excel_changes)
        build_menu.add_separator()
        build_menu.add_command(label="Build NVM  (Python)", command=self.run_build_flow)
        build_menu.add_command(label="VBA Build  (Legacy Excel Macro)", command=self.generate_nvm)
        build_menu.add_separator()
        build_menu.add_command(label="Read NVM \u2192 Excel", command=self.read_existing_nvm)

        help_menu = tk.Menu(menubar, tearoff=0, font=('Segoe UI', 9))
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self._show_about)

        self.bind_all("<Control-l>", lambda e: self.load_from_excel())
        self.bind_all("<Control-b>", lambda e: self.run_build_flow())

        # ── Header bar ────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=self.colors['header_bg'], height=56)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(header, text="Intel\u00ae GBE NVM Configuration Studio",
                font=('Segoe UI', 14, 'bold'),
                fg=self.colors['header_fg'],
                bg=self.colors['header_bg']).pack(side=tk.LEFT, padx=20, pady=12)
        tk.Label(header, text="v2.1  |  GBE NVM Engineering Tool",
                font=('Segoe UI', 9),
                fg='#a8d4f0',
                bg=self.colors['header_bg']).pack(side=tk.RIGHT, padx=20, pady=12)
        
        # Toolbar
        top = ttk.Frame(self, padding=12)
        top.pack(fill=tk.X)

        self.folder_var = tk.StringVar()
        self.bin_var = tk.StringVar()
        self.txt_var = tk.StringVar()

        ttk.Label(top, text="Project Folder:", style='Param.TLabel').pack(side=tk.LEFT, padx=(0, 6))
        self.folder_combo = ttk.Combobox(top, textvariable=self.folder_var, width=50, state="readonly", font=('Segoe UI', 9))
        self.folder_combo.pack(side=tk.LEFT, padx=(0, 12))
        self.folder_combo.bind("<<ComboboxSelected>>", lambda _e: self.open_folder())
        ttk.Button(top, text="Refresh", command=self._load_gbe_folders).pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="New Project", command=self._open_new_project_dialog).pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="Load Excel  [Ctrl+L]", command=self.load_from_excel, style='Accent.TButton').pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="Build NVM  [Ctrl+B]", command=self.run_build_flow, style='Build.TButton').pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="Open Output Folder", command=self._open_output_folder).pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="Version Diff", command=self.show_version_diff).pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="Clone NVM Images", command=self._clone_nvm_images).pack(side=tk.LEFT, padx=4)

        # Parameters section - modern card design
        param_outer = tk.Frame(self, bg=self.colors['bg'])
        param_outer.pack(fill=tk.X, padx=16, pady=(8, 12))
        
        param_frame = tk.Frame(param_outer, bg=self.colors['param_bg'], relief=tk.FLAT, bd=1)
        param_frame.pack(fill=tk.X, padx=1, pady=1)
        
        # Header
        param_header = tk.Frame(param_frame, bg=self.colors['accent'], height=35)
        param_header.pack(fill=tk.X)
        param_header.pack_propagate(False)
        tk.Label(param_header, text="NVM Configuration Parameters", 
                font=('Segoe UI', 10, 'bold'), fg='white', bg=self.colors['accent']).pack(side=tk.LEFT, padx=12, pady=6)
        
        # Parameter grid
        param_content = tk.Frame(param_frame, bg=self.colors['param_bg'], padx=16, pady=12)
        param_content.pack(fill=tk.X)
        
        # Define all StringVars first (moved from later in code)
        self.project_var = tk.StringVar(value="MTL_M_P")
        self.silicon_var = tk.StringVar(value="Nahum11")
        self.step_var = tk.StringVar(value="C0")
        self.major_var = tk.StringVar(value="1")
        self.minor_var = tk.StringVar(value="4")
        self.device_id_var = tk.StringVar(value="0x0DC9")
        self.sku_device_id_var = tk.StringVar(value="0x0DCA")
        self.nvm_output_var = tk.StringVar(value="MTL_M_P")
        self.lan_sw_var = tk.StringVar(value="lan")
        self.lm_v_var = tk.StringVar(value="Both")
        
        # Add trace to update when changed
        for var in [self.project_var, self.silicon_var, self.step_var, self.major_var, 
                   self.minor_var, self.device_id_var, self.sku_device_id_var, 
                   self.nvm_output_var, self.lan_sw_var, self.lm_v_var]:
            var.trace('w', lambda *args: self.on_param_change())
        
        # Create 4-column grid layout
        params = [
            ("Project Name:", self.project_var, 0, 0),
            ("Silicon:", self.silicon_var, 0, 2),
            ("Step:", self.step_var, 0, 4),
            ("NVM Output:", self.nvm_output_var, 0, 6),
            ("Major Version:", self.major_var, 1, 0),
            ("Minor Version:", self.minor_var, 1, 2),
            ("Device ID:", self.device_id_var, 1, 4),
            ("SKU Device ID:", self.sku_device_id_var, 1, 6),
        ]
        
        for label_text, var, row, col in params:
            tk.Label(param_content, text=label_text, font=('Segoe UI', 9, 'bold'), 
                    bg=self.colors['param_bg'], fg='#333333', anchor=tk.W).grid(row=row, column=col, sticky=tk.W, padx=(0, 4), pady=6)
            entry = tk.Entry(param_content, textvariable=var, font=('Segoe UI', 9), 
                           width=18, relief=tk.SOLID, bd=1)
            entry.grid(row=row, column=col+1, sticky=tk.W, padx=(0, 20), pady=6)
        
        # LAN/V selection row
        select_frame = tk.Frame(param_content, bg=self.colors['param_bg'])
        select_frame.grid(row=2, column=0, columnspan=8, sticky=tk.W, pady=(8, 0))
        
        tk.Label(select_frame, text="LAN/Non-LAN SW:", font=('Segoe UI', 9, 'bold'), 
                bg=self.colors['param_bg'], fg='#333333').pack(side=tk.LEFT, padx=(0, 8))
        tk.Radiobutton(select_frame, text="LAN SW", value="lan", variable=self.lan_sw_var,
                      bg=self.colors['param_bg'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=4)
        tk.Radiobutton(select_frame, text="Non-LAN SW", value="non_lan", variable=self.lan_sw_var,
                      bg=self.colors['param_bg'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=4)
        
        tk.Label(select_frame, text="      LM/V Select:", font=('Segoe UI', 9, 'bold'), 
                bg=self.colors['param_bg'], fg='#333333').pack(side=tk.LEFT, padx=(20, 8))
        tk.Radiobutton(select_frame, text="V", value="V", variable=self.lm_v_var,
                      bg=self.colors['param_bg'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=4)
        tk.Radiobutton(select_frame, text="LM", value="LM", variable=self.lm_v_var,
                      bg=self.colors['param_bg'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=4)
        tk.Radiobutton(select_frame, text="Both", value="Both", variable=self.lm_v_var,
                      bg=self.colors['param_bg'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=4)
        
        # Notes section (below parameters)
        notes_frame = tk.Frame(param_frame, bg=self.colors['param_bg'])
        notes_frame.pack(fill=tk.X, padx=16, pady=12)
        
        tk.Label(notes_frame, text="Project Notes:", font=('Segoe UI', 9, 'bold'), 
                bg=self.colors['param_bg'], fg='#333333', anchor=tk.W).pack(anchor=tk.W, pady=(0, 4))
        
        self.notes_text = tk.Text(notes_frame, height=3, wrap=tk.WORD, font=("Segoe UI", 9),
                                 relief=tk.SOLID, bd=1)
        notes_scroll = ttk.Scrollbar(notes_frame, orient=tk.VERTICAL, command=self.notes_text.yview)
        notes_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.notes_text.configure(yscrollcommand=notes_scroll.set)
        self.notes_text.pack(fill=tk.X)
        self.notes_text.insert("1.0", "Enter project notes and documentation here...")
        self.notes_text.bind('<FocusIn>', lambda e: self.notes_text.delete("1.0", tk.END) if self.notes_text.get("1.0", "end-1c") == "Enter project notes and documentation here..." else None)

        # ── Status bar (packed before main so it sticks to the bottom) ──────
        status_bar = tk.Frame(self, bg='#e8e8e8', height=24, relief=tk.SUNKEN, bd=1)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        status_bar.pack_propagate(False)
        tk.Label(status_bar, textvariable=self.status_var, anchor=tk.W,
                 font=('Segoe UI', 8), bg='#e8e8e8', fg='#333333').pack(side=tk.LEFT, padx=10)
        tk.Label(status_bar, text="Intel\u00ae GBE NVM Configuration Studio  v2.1",
                 anchor=tk.E, font=('Segoe UI', 8), bg='#e8e8e8', fg='#888888').pack(side=tk.RIGHT, padx=10)

        main = tk.Frame(self, bg=self.colors['bg'])
        main.pack(fill=tk.BOTH, expand=True, padx=12)

        # Create notebook for tabs with modern styling
        style = ttk.Style()
        style.configure('TNotebook', background=self.colors['bg'])
        style.configure('TNotebook.Tab', font=('Segoe UI', 10), padding=[20, 8])
        
        main_paned = ttk.PanedWindow(main, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True)
        nb_frame = ttk.Frame(main_paned)
        main_paned.add(nb_frame, weight=3)
        self.notebook = ttk.Notebook(nb_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=(0, 4))

        # Initialize variables needed for tree creation
        self.show_less_var = tk.BooleanVar(value=False)  # Start with full view
        self.show_less_var.trace('w', lambda *args: self._on_show_less_changed())

        # Tab 1: Excel Data View
        excel_tab = ttk.Frame(self.notebook)
        self.notebook.add(excel_tab, text="Excel Data")

        # Use PanedWindow for resizable panels
        excel_paned = ttk.PanedWindow(excel_tab, orient=tk.HORIZONTAL)
        excel_paned.pack(fill=tk.BOTH, expand=True)

        # Excel data tree (left side)
        excel_left = ttk.Frame(excel_paned)
        excel_paned.add(excel_left, weight=3)
        self.excel_tree_parent = excel_left  # Store for tree recreation

        # Define all possible columns
        self.excel_columns_full = [
            ("offset",         "Offset",       70),
            ("bits",           "Bits",         80),
            ("rtl_name",       "RTL Name",    180),
            ("name",           "C-Spec Name", 220),
            ("bits_owner",     "Owner",       100),
            ("v",              "V",            80),
            ("lm",             "LM",           80),
            ("values_read",    "Values Read",  110),
            ("final_value",    "Final Value",  90),
            ("comments",       "Comments",    200),
            ("changed_version","Changed Ver",  110),
        ]

        # Simple view: hide the mostly-N/A columns
        # (RTL Name, Owner, Values Read, Final Value, Changed Ver are typically N/A)
        self.excel_columns_simple = [
            ("offset",   "Offset",      70),
            ("bits",     "Bits",        80),
            ("name",     "C-Spec Name", 240),
            ("v",        "V",           90),
            ("lm",       "LM",          90),
            ("comments", "Comments",   260),
        ]
        
        # Start with simple view
        self._create_excel_tree(excel_left)
        
        # Description panel (right side, resizable)
        excel_right = ttk.Frame(excel_paned, padding=(12, 0, 0, 0))
        excel_paned.add(excel_right, weight=1)

        ttk.Label(excel_right, text="Bit Description", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)
        
        self.bit_desc_text = tk.Text(excel_right, wrap=tk.WORD, font=("Segoe UI", 9), height=10)
        self.bit_desc_text.pack(fill=tk.BOTH, expand=True, pady=(8, 0))
        desc_scrollbar = ttk.Scrollbar(excel_right, orient=tk.VERTICAL, command=self.bit_desc_text.yview)
        desc_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.bit_desc_text.configure(yscrollcommand=desc_scrollbar.set)

        # Edit values section
        ttk.Separator(excel_right, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(12, 8))
        ttk.Label(excel_right, text="Edit Values", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)
        
        edit_frame = ttk.Frame(excel_right)
        edit_frame.pack(fill=tk.X, pady=(8, 0))
        
        # V value
        v_frame = ttk.Frame(edit_frame)
        v_frame.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=4)
        self.update_v_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(v_frame, text="V Value:", variable=self.update_v_var).pack(side=tk.LEFT)
        self.v_value_var = tk.StringVar()
        self.v_entry = ttk.Entry(v_frame, textvariable=self.v_value_var, width=15)
        self.v_entry.pack(side=tk.LEFT, padx=(8, 0))
        
        # LM value
        lm_frame = ttk.Frame(edit_frame)
        lm_frame.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=4)
        self.update_lm_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(lm_frame, text="LM Value:", variable=self.update_lm_var).pack(side=tk.LEFT)
        self.lm_value_var = tk.StringVar()
        self.lm_entry = ttk.Entry(lm_frame, textvariable=self.lm_value_var, width=15)
        self.lm_entry.pack(side=tk.LEFT, padx=(8, 0))
        
        # Save button
        button_frame = ttk.Frame(excel_right)
        button_frame.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(button_frame, text="Save Changes", command=self.save_bit_changes).pack(side=tk.LEFT)
        
        # Store selected row index
        self.selected_row_idx = None

        # Tab 2: LCD Extension (Non LAN SW)
        lcd_non_lan_tab = ttk.Frame(self.notebook)
        self.notebook.add(lcd_non_lan_tab, text="LCD Non-LAN")
        
        self.lcd_non_lan_tree = ttk.Treeview(lcd_non_lan_tab, show="headings", height=20)
        self.lcd_non_lan_tree.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.lcd_non_lan_tree.bind("<Button-3>", lambda e: self.show_tab_context_menu(e, "lcd_non_lan"))
        lcd_non_lan_scroll = ttk.Scrollbar(lcd_non_lan_tab, orient=tk.VERTICAL, command=self.lcd_non_lan_tree.yview)
        lcd_non_lan_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.lcd_non_lan_tree.configure(yscrollcommand=lcd_non_lan_scroll.set)
        
        # Tab 3: LCD Extension (LAN SW)
        lcd_lan_tab = ttk.Frame(self.notebook)
        self.notebook.add(lcd_lan_tab, text="LCD LAN")
        
        self.lcd_lan_tree = ttk.Treeview(lcd_lan_tab, show="headings", height=20)
        self.lcd_lan_tree.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.lcd_lan_tree.bind("<Button-3>", lambda e: self.show_tab_context_menu(e, "lcd_lan"))
        lcd_lan_scroll = ttk.Scrollbar(lcd_lan_tab, orient=tk.VERTICAL, command=self.lcd_lan_tree.yview)
        lcd_lan_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.lcd_lan_tree.configure(yscrollcommand=lcd_lan_scroll.set)
        
        # Tab 4: ISCSI Module
        iscsi_tab = ttk.Frame(self.notebook)
        self.notebook.add(iscsi_tab, text="ISCSI Module")
        
        self.iscsi_tree = ttk.Treeview(iscsi_tab, show="headings", height=20)
        self.iscsi_tree.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.iscsi_tree.bind("<Button-3>", lambda e: self.show_tab_context_menu(e, "iscsi"))
        iscsi_scroll = ttk.Scrollbar(iscsi_tab, orient=tk.VERTICAL, command=self.iscsi_tree.yview)
        iscsi_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.iscsi_tree.configure(yscrollcommand=iscsi_scroll.set)
        
        # Tab 5: Basic NVM Map
        basic_nvm_tab = ttk.Frame(self.notebook)
        self.notebook.add(basic_nvm_tab, text="Basic NVM")
        
        self.basic_nvm_tree = ttk.Treeview(basic_nvm_tab, show="headings", height=20)
        self.basic_nvm_tree.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.basic_nvm_tree.bind("<Button-3>", lambda e: self.show_tab_context_menu(e, "basic_nvm"))
        basic_nvm_scroll = ttk.Scrollbar(basic_nvm_tab, orient=tk.VERTICAL, command=self.basic_nvm_tree.yview)
        basic_nvm_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.basic_nvm_tree.configure(yscrollcommand=basic_nvm_scroll.set)

        # Tab 6: Word list
        word_tab = ttk.Frame(self.notebook)
        self.notebook.add(word_tab, text="Word Values")

        word_left = ttk.Frame(word_tab)
        word_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(word_left, columns=("index", "value"), show="headings", height=20)
        self.tree.heading("index", text="Word Index")
        self.tree.heading("value", text="Value (hex)")
        self.tree.column("index", width=150, minwidth=80, stretch=True, anchor=tk.E)
        self.tree.column("value", width=150, minwidth=80, stretch=True, anchor=tk.W)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree.bind("<<TreeviewSelect>>", self.on_word_select)

        scrollbar = ttk.Scrollbar(word_left, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Right panel — part of main PanedWindow for native drag-resize
        right_outer = ttk.Frame(main_paned)
        main_paned.add(right_outer, weight=0)

        # Scroll arrow buttons at the very top
        right_arrows = ttk.Frame(right_outer)
        right_arrows.pack(side=tk.TOP, fill=tk.X)

        right_vscroll = ttk.Scrollbar(right_outer, orient=tk.VERTICAL)
        right_vscroll.pack(side=tk.RIGHT, fill=tk.Y)

        right_canvas = tk.Canvas(right_outer, width=240, highlightthickness=0, bd=0,
                                 bg=self.colors['bg'])
        right_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right_canvas.configure(yscrollcommand=right_vscroll.set)
        right_vscroll.configure(command=right_canvas.yview)

        right = ttk.Frame(right_canvas, padding=(12, 0, 4, 0))
        _rwin = right_canvas.create_window((0, 0), window=right, anchor='nw')

        def _on_right_inner_cfg(event):
            right_canvas.configure(scrollregion=right_canvas.bbox("all"))

        def _on_right_canvas_cfg(event):
            right_canvas.itemconfig(_rwin, width=event.width)

        right.bind('<Configure>', _on_right_inner_cfg)
        right_canvas.bind('<Configure>', _on_right_canvas_cfg)

        def _mw_right(event):
            right_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_mw(e=None):
            right_canvas.bind_all('<MouseWheel>', _mw_right)

        def _unbind_mw(e=None):
            right_canvas.unbind_all('<MouseWheel>')

        right_canvas.bind('<Enter>', _bind_mw)
        right_canvas.bind('<Leave>', _unbind_mw)
        right.bind('<Enter>', _bind_mw)
        right.bind('<Leave>', _unbind_mw)

        ttk.Button(right_arrows, text="▲", width=4,
                   command=lambda: right_canvas.yview_scroll(-3, "units")).pack(side=tk.LEFT, padx=(2, 0), pady=1)
        ttk.Button(right_arrows, text="▼", width=4,
                   command=lambda: right_canvas.yview_scroll(3, "units")).pack(side=tk.LEFT, padx=2, pady=1)
        ttk.Button(right_arrows, text="⏫", width=4,
                   command=lambda: right_canvas.yview_moveto(0)).pack(side=tk.LEFT, padx=(0, 2), pady=1)

        # Define remaining StringVars
        self.index_var = tk.StringVar()
        self.value_var = tk.StringVar()
        self.bit_var = tk.StringVar()
        self.register_name_var = tk.StringVar()
        self.register_desc_var = tk.StringVar()
        self.excel_col_var = tk.StringVar(value="V")
        self.hide_values_var = tk.BooleanVar(value=False)
        self.create_with_read_var = tk.BooleanVar(value=False)
        
        # Initialize excel_data
        self.excel_data = []
        self.lcd_non_lan_data = []
        self.lcd_lan_data = []
        self.iscsi_data = []
        self.basic_nvm_data = []

        # Quick Actions Section
        ttk.Label(right, text="Quick Actions", font=('Segoe UI', 11, 'bold'), foreground='#0071c5').pack(anchor=tk.W, pady=(0, 12))
        
        ttk.Checkbutton(right, text="Hide N/A columns", variable=self.show_less_var).pack(anchor=tk.W, pady=(6, 0))
        ttk.Checkbutton(right, text="Hide values read", variable=self.hide_values_var).pack(anchor=tk.W, pady=(4, 0))
        ttk.Checkbutton(right, text="Create w/ read values", variable=self.create_with_read_var).pack(anchor=tk.W, pady=(4, 0))
        
        ttk.Button(right, text="Sync Edits to Build", command=self.sync_all_excel_changes, width=28).pack(anchor=tk.W, pady=(8, 0))

        ttk.Separator(right, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=12)

        # Generate buttons
        ttk.Button(right, text="VBA Build (Legacy)", command=self.generate_nvm, width=28).pack(anchor=tk.W, pady=4)
        ttk.Button(right, text="Version Diff", command=self.show_version_diff, width=28).pack(anchor=tk.W, pady=4)
        ttk.Button(right, text="Read NVM \u2192 Excel", command=self.read_existing_nvm, width=28).pack(anchor=tk.W, pady=4)
        ttk.Button(right, text="Open Output Folder", command=self._open_output_folder, width=28).pack(anchor=tk.W, pady=4)
        ttk.Button(right, text="Exit", command=self.quit, width=28).pack(anchor=tk.W, pady=(12, 0))

        ttk.Separator(right, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=12)

        # Register Editor Section
        ttk.Label(right, text="Register Editor", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(4, 6))

        ttk.Label(right, text="Excel Column:").pack(anchor=tk.W, pady=(4, 2))
        excel_col_frame = ttk.Frame(right)
        excel_col_frame.pack(anchor=tk.W)
        ttk.Radiobutton(excel_col_frame, text="V", value="V", variable=self.excel_col_var).pack(side=tk.LEFT)
        ttk.Radiobutton(excel_col_frame, text="LM", value="LM", variable=self.excel_col_var).pack(side=tk.LEFT)
        ttk.Radiobutton(excel_col_frame, text="Both", value="Both", variable=self.excel_col_var).pack(side=tk.LEFT)

        ttk.Label(right, text="Register Name:").pack(anchor=tk.W, pady=(6, 2))
        ttk.Entry(right, textvariable=self.register_name_var, width=20).pack(anchor=tk.W)

        ttk.Label(right, text="Description:").pack(anchor=tk.W, pady=(4, 2))
        desc_label = ttk.Label(right, textvariable=self.register_desc_var, width=28, foreground="#0066CC", wraplength=180, justify=tk.LEFT)
        desc_label.pack(anchor=tk.W)

        ttk.Label(right, text="Word Index:").pack(anchor=tk.W, pady=(4, 2))
        ttk.Entry(right, textvariable=self.index_var, width=20, state="readonly").pack(anchor=tk.W)

        ttk.Label(right, text="Current Value:").pack(anchor=tk.W, pady=(4, 2))
        ttk.Entry(right, textvariable=self.value_var, width=20, state="readonly").pack(anchor=tk.W)

        ttk.Label(right, text="Bit Number (0-15):").pack(anchor=tk.W, pady=(6, 2))
        ttk.Entry(right, textvariable=self.bit_var, width=10).pack(anchor=tk.W)

        bit_btns = ttk.Frame(right)
        bit_btns.pack(anchor=tk.W, pady=(4, 0))
        ttk.Button(bit_btns, text="Set Bit", command=lambda: self.set_clear_bit(True)).pack(side=tk.LEFT)
        ttk.Button(bit_btns, text="Clear Bit", command=lambda: self.set_clear_bit(False)).pack(side=tk.LEFT, padx=4)

        ttk.Separator(right, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        ttk.Label(right, text="Registers", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W)

        # Search filter row
        reg_search_frame = ttk.Frame(right)
        reg_search_frame.pack(fill=tk.X, pady=(4, 2))
        ttk.Label(reg_search_frame, text="Search:", font=('Segoe UI', 8)).pack(side=tk.LEFT, padx=(0, 4))
        self.reg_search_var = tk.StringVar()
        self.reg_search_var.trace_add("write", lambda *_: self._filter_registers())
        reg_search_entry = ttk.Entry(reg_search_frame, textvariable=self.reg_search_var, width=18)
        reg_search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(reg_search_frame, text="✕", width=3,
                   command=self._clear_reg_search).pack(side=tk.LEFT, padx=(4, 0))

        self.reg_tree = ttk.Treeview(right, columns=("name", "index"), show="headings", height=8)
        self.reg_tree.heading("name", text="Name")
        self.reg_tree.heading("index", text="Word")
        self.reg_tree.column("name", width=160)
        self.reg_tree.column("index", width=55, anchor=tk.E)
        self.reg_tree.pack(fill=tk.X, pady=(4, 0))
        self.reg_tree.bind("<<TreeviewSelect>>", self.on_register_select)

        # Row 1: edit operations
        reg_btns = ttk.Frame(right)
        reg_btns.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(reg_btns, text="Add",        command=self.add_register,                   width=8).pack(side=tk.LEFT)
        ttk.Button(reg_btns, text="Edit",       command=self.edit_register,                  width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(reg_btns, text="Remove",     command=self.remove_register,                width=8).pack(side=tk.LEFT)
        ttk.Button(reg_btns, text="From Excel", command=self._populate_registers_from_excel, width=11).pack(side=tk.LEFT, padx=2)
        # Row 2: persistence
        reg_btns2 = ttk.Frame(right)
        reg_btns2.pack(fill=tk.X, pady=(2, 0))
        ttk.Button(reg_btns2, text="Save Map", command=self.save_registers, width=11).pack(side=tk.LEFT)
        ttk.Button(reg_btns2, text="Load Map", command=self.load_registers, width=11).pack(side=tk.LEFT, padx=2)
        
    def _create_excel_tree(self, parent):
        """Create or recreate the Excel tree with current column configuration."""
        # Destroy existing tree if it exists
        if hasattr(self, 'excel_tree'):
            self.excel_tree.destroy()
            if hasattr(self, 'excel_scrollbar'):
                self.excel_scrollbar.destroy()
            if hasattr(self, 'excel_xscrollbar'):
                self.excel_xscrollbar.destroy()
        
        # Choose columns based on show_less_var
        columns_to_use = self.excel_columns_simple if self.show_less_var.get() else self.excel_columns_full
        col_ids = [col[0] for col in columns_to_use]
        
        self.excel_tree = ttk.Treeview(
            parent,
            columns=col_ids,
            show="headings",
            height=25
        )
        
        # Configure columns
        for col_id, col_name, col_width in columns_to_use:
            self.excel_tree.heading(col_id, text=col_name)
            self.excel_tree.column(col_id, width=col_width, minwidth=50, stretch=True, anchor=tk.W)
        
        # Style alternating rows
        self.excel_tree.tag_configure('oddrow',  background='#f9f9f9')
        self.excel_tree.tag_configure('evenrow', background='#ffffff')

        # Pack scrollbar FIRST so tree doesn't steal all space
        self.excel_scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.excel_tree.yview)
        self.excel_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.excel_xscrollbar = ttk.Scrollbar(parent, orient=tk.HORIZONTAL, command=self.excel_tree.xview)
        self.excel_xscrollbar.pack(side=tk.BOTTOM, fill=tk.X)

        self.excel_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.excel_tree.configure(yscrollcommand=self.excel_scrollbar.set,
                                  xscrollcommand=self.excel_xscrollbar.set)
        
        self.excel_tree.bind("<<TreeviewSelect>>", self.on_excel_row_select)
        self.excel_tree.bind("<Double-1>", self.on_excel_tree_double_click)
        self.excel_tree.bind("<Button-3>", self.show_excel_context_menu)

        # Refresh data if available
        if hasattr(self, 'excel_data') and self.excel_data:
            self.refresh_excel_tree()

    def on_param_change(self) -> None:
        """Called when any parameter is changed."""
        # Could be used for auto-save or validation in the future
        pass

    def _load_general_variables(self, ws) -> None:
        """Load parameters from 'general variable' sheet and populate UI fields."""
        try:
            # Read key-value pairs from the sheet (Column A = key, Column B = value)
            for row_num in range(2, min(15, ws.max_row + 1)):  # Rows 2-14 contain variables
                key = ws.cell(row_num, 1).value
                value = ws.cell(row_num, 2).value
                
                if not key:
                    continue
                
                key_str = str(key).strip().lower()
                value_str = str(value).strip() if value else ""
                
                # Map Excel keys to UI variables
                if "silicon" in key_str:
                    self.silicon_var.set(value_str)
                elif "step" in key_str:
                    self.step_var.set(value_str)
                elif "project" in key_str or "nahum" in key_str.lower():
                    self.project_var.set(value_str)
                elif "device id" in key_str and "sku" not in key_str:
                    self.device_id_var.set(value_str)
                elif "sku" in key_str:
                    self.sku_device_id_var.set(value_str)
                elif "image file" in key_str:
                    # Extract project name from image file name
                    if value_str:
                        self.nvm_output_var.set(value_str.replace("GBE_", "").replace("_ALL", ""))
                elif "version" in key_str and "image" in key_str:
                    # Parse version number (handles "1.04", "104", or "1.4")
                    if value_str:
                        # Remove any leading zeros and handle decimal format
                        if '.' in value_str:
                            parts = value_str.split('.')
                            self.major_var.set(parts[0])
                            self.minor_var.set(parts[1] if len(parts) > 1 else '0')
                        elif value_str.isdigit() and len(value_str) >= 2:
                            # Format like "104" -> "1.04"
                            self.major_var.set(value_str[0])
                            self.minor_var.set(value_str[1:])
                        else:
                            self.major_var.set(value_str)
                            self.minor_var.set('0')
        except Exception as e:
            print(f"Warning: Could not fully load general variables: {e}")
    
    def _on_show_less_changed(self) -> None:
        """Rebuild Excel tree when show_less checkbox changes."""
        if hasattr(self, 'excel_tree_parent'):
            self._create_excel_tree(self.excel_tree_parent)
    
    def _load_additional_sheets(self, wb) -> None:
        """Load data from LCD, ISCSI, and Basic NVM Map sheets."""
        # Initialize data structures
        self.lcd_non_lan_data = []
        self.lcd_lan_data = []
        self.iscsi_data = []
        self.basic_nvm_data = []
        
        # Load LCD Extension (Non LAN SW)
        if "LCD extention non LAN SW" in wb.sheetnames:
            ws = wb["LCD extention non LAN SW"]
            for row_num in range(2, ws.max_row + 1):
                row_data = []
                has_data = False
                for col_num in range(1, 6):  # 5 columns
                    val = ws.cell(row_num, col_num).value
                    if val is not None:
                        has_data = True
                    row_data.append(str(val) if val is not None else '')
                if has_data:
                    self.lcd_non_lan_data.append(row_data)
        
        # Load LCD Extension (LAN SW)
        if "LCD extention LAN SW" in wb.sheetnames:
            ws = wb["LCD extention LAN SW"]
            for row_num in range(2, ws.max_row + 1):
                row_data = []
                has_data = False
                for col_num in range(1, 6):  # 5 columns
                    val = ws.cell(row_num, col_num).value
                    if val is not None:
                        has_data = True
                    row_data.append(str(val) if val is not None else '')
                if has_data:
                    self.lcd_lan_data.append(row_data)
        
        # Load ISCSI_MODULE
        if "ISCSI_MODULE" in wb.sheetnames:
            ws = wb["ISCSI_MODULE"]
            for row_num in range(2, ws.max_row + 1):
                row_data = []
                has_data = False
                for col_num in range(1, 6):  # 5 columns
                    val = ws.cell(row_num, col_num).value
                    if val is not None:
                        has_data = True
                    row_data.append(str(val) if val is not None else '')
                if has_data:
                    self.iscsi_data.append(row_data)
        
        # Load basic nvm map
        if "basic nvm map" in wb.sheetnames:
            ws = wb["basic nvm map"]
            for row_num in range(2, ws.max_row + 1):
                row_data = []
                has_data = False
                for col_num in range(1, 6):  # 5 columns
                    val = ws.cell(row_num, col_num).value
                    if val is not None:
                        has_data = True
                    row_data.append(str(val) if val is not None else '')
                if has_data:
                    self.basic_nvm_data.append(row_data)
    
    def refresh_additional_trees(self) -> None:
        """Populate additional Excel sheet trees with loaded data."""
        # Common column headers for all additional sheets (5 columns)
        headers = ["Offset", "Bits", "Name", "Value", "Description"]
        
        # Configure and populate LCD Non-LAN tree
        if hasattr(self, 'lcd_non_lan_tree') and hasattr(self, 'lcd_non_lan_data'):
            self.lcd_non_lan_tree['columns'] = tuple(range(len(headers)))
            for i, header in enumerate(headers):
                self.lcd_non_lan_tree.heading(i, text=header)
                self.lcd_non_lan_tree.column(i, width=150, minwidth=50, stretch=True)
            self.lcd_non_lan_tree.delete(*self.lcd_non_lan_tree.get_children())
            for row_data in self.lcd_non_lan_data:
                self.lcd_non_lan_tree.insert("", tk.END, values=row_data)
            self.lcd_non_lan_tree.bind("<Double-1>", lambda e: self.on_tree_double_click(e, self.lcd_non_lan_tree, self.lcd_non_lan_data))
        
        # Configure and populate LCD LAN tree
        if hasattr(self, 'lcd_lan_tree') and hasattr(self, 'lcd_lan_data'):
            self.lcd_lan_tree['columns'] = tuple(range(len(headers)))
            for i, header in enumerate(headers):
                self.lcd_lan_tree.heading(i, text=header)
                self.lcd_lan_tree.column(i, width=150, minwidth=50, stretch=True)
            self.lcd_lan_tree.delete(*self.lcd_lan_tree.get_children())
            for row_data in self.lcd_lan_data:
                self.lcd_lan_tree.insert("", tk.END, values=row_data)
            self.lcd_lan_tree.bind("<Double-1>", lambda e: self.on_tree_double_click(e, self.lcd_lan_tree, self.lcd_lan_data))
        
        # Configure and populate ISCSI tree
        if hasattr(self, 'iscsi_tree') and hasattr(self, 'iscsi_data'):
            self.iscsi_tree['columns'] = tuple(range(len(headers)))
            for i, header in enumerate(headers):
                self.iscsi_tree.heading(i, text=header)
                self.iscsi_tree.column(i, width=150, minwidth=50, stretch=True)
            self.iscsi_tree.delete(*self.iscsi_tree.get_children())
            for row_data in self.iscsi_data:
                self.iscsi_tree.insert("", tk.END, values=row_data)
            self.iscsi_tree.bind("<Double-1>", lambda e: self.on_tree_double_click(e, self.iscsi_tree, self.iscsi_data))
        
        # Configure and populate Basic NVM tree
        if hasattr(self, 'basic_nvm_tree') and hasattr(self, 'basic_nvm_data'):
            self.basic_nvm_tree['columns'] = tuple(range(len(headers)))
            for i, header in enumerate(headers):
                self.basic_nvm_tree.heading(i, text=header)
                self.basic_nvm_tree.column(i, width=150, minwidth=50, stretch=True)
            self.basic_nvm_tree.delete(*self.basic_nvm_tree.get_children())
            for row_data in self.basic_nvm_data:
                self.basic_nvm_tree.insert("", tk.END, values=row_data)
            self.basic_nvm_tree.bind("<Double-1>", lambda e: self.on_tree_double_click(e, self.basic_nvm_tree, self.basic_nvm_data))

    def update_notes_with_parameters(self) -> None:
        """No longer needed - parameters are now directly editable in the UI."""
        pass

    def _set_status(self, msg: str) -> None:
        """Update the status bar message."""
        if hasattr(self, 'status_var'):
            self.status_var.set(msg)
            self.update_idletasks()

    def _open_output_folder(self) -> None:
        """Open the current project folder in Windows Explorer."""
        if self.current_folder and self.current_folder.exists():
            os.startfile(str(self.current_folder))
        else:
            messagebox.showinfo(
                "Open Output Folder",
                "No project folder loaded yet.\nSelect a project from the dropdown first."
            )

    def _show_about(self) -> None:
        """Show the About dialog."""
        win = tk.Toplevel(self)
        win.title("About  \u2014  Intel\u00ae GBE NVM Configuration Studio")
        win.geometry("440x300")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        hdr = tk.Frame(win, bg=self.colors['header_bg'], height=60)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Intel\u00ae GBE NVM Configuration Studio",
                 font=('Segoe UI', 12, 'bold'), fg='white',
                 bg=self.colors['header_bg']).pack(expand=True, pady=16)

        body = tk.Frame(win, bg='white', padx=28, pady=16)
        body.pack(fill=tk.BOTH, expand=True)

        info_rows = [
            ("Version:",   "2.1"),
            ("Purpose:",   "GBE NVM image generation and editing"),
            ("Supports:",  "MTL-M/P  (Nahum11) and compatible platforms"),
            ("Runtime:",   "Python 3.13  +  openpyxl  +  pywin32"),
            ("Build mode:", "Python (native) and VBA Legacy"),
        ]
        for label, value in info_rows:
            row = tk.Frame(body, bg='white')
            row.pack(fill=tk.X, pady=4)
            tk.Label(row, text=label, width=14, anchor=tk.W,
                     font=('Segoe UI', 9, 'bold'), bg='white', fg='#222').pack(side=tk.LEFT)
            tk.Label(row, text=value, anchor=tk.W,
                     font=('Segoe UI', 9), bg='white', fg='#555').pack(side=tk.LEFT)

        ttk.Button(win, text="Close", command=win.destroy).pack(pady=14)

    def sync_all_excel_changes(self) -> None:
        """Sync all edited values from excel_data to word arrays."""
        if not self.excel_data:
            messagebox.showwarning("Sync Changes", "No Excel data loaded.")
            return
        
        try:
            synced_count = 0
            
            for row_data in self.excel_data:
                word_idx = row_data.get('word_idx')
                if word_idx is None:
                    continue
                
                # Parse bit range
                bits_str = str(row_data.get('bits', ''))
                if ':' not in bits_str:
                    continue
                
                parts = bits_str.split(':')
                high_bit = int(parts[0].strip())
                low_bit = int(parts[1].strip())
                num_bits = high_bit - low_bit + 1
                mask = (1 << num_bits) - 1
                
                # Sync V value
                v_val = row_data.get('v', '')
                if v_val and str(v_val).upper() not in ['N/A', 'NONE', '']:
                    try:
                        v_str = str(v_val).replace('0x', '').replace('0X', '').strip()
                        v_int = int(v_str, 16)
                        
                        while len(self.words_v) <= word_idx:
                            self.words_v.append(0xFFFF)
                        
                        current_word = self.words_v[word_idx]
                        current_word &= ~(mask << low_bit)
                        current_word |= (v_int & mask) << low_bit
                        self.words_v[word_idx] = current_word
                        # Record dirty delta
                        fms = mask << low_bit
                        s, c = self.dirty_bits_v.get(word_idx, (0, 0))
                        c |= fms; s = (s & ~fms) | ((v_int & mask) << low_bit)
                        self.dirty_bits_v[word_idx] = (s, c)
                        synced_count += 1
                    except:
                        pass
                
                # Sync LM value
                lm_val = row_data.get('lm', '')
                if lm_val and str(lm_val).upper() not in ['N/A', 'NONE', '']:
                    try:
                        lm_str = str(lm_val).replace('0x', '').replace('0X', '').strip()
                        lm_int = int(lm_str, 16)
                        
                        while len(self.words_lm) <= word_idx:
                            self.words_lm.append(0xFFFF)
                        
                        current_word = self.words_lm[word_idx]
                        current_word &= ~(mask << low_bit)
                        current_word |= (lm_int & mask) << low_bit
                        self.words_lm[word_idx] = current_word
                        # Record dirty delta
                        fms = mask << low_bit
                        s, c = self.dirty_bits_lm.get(word_idx, (0, 0))
                        c |= fms; s = (s & ~fms) | ((lm_int & mask) << low_bit)
                        self.dirty_bits_lm[word_idx] = (s, c)
                        synced_count += 1
                    except:
                        pass
            
            # Update display array based on current selection
            if self.excel_col_var.get() == "V":
                self.words = self.words_v.copy()
            else:
                self.words = self.words_lm.copy()
            
            self.refresh_word_list()
            messagebox.showinfo("Sync Complete", f"Successfully synced {synced_count} values from Excel data to build arrays.\n\nYour edits are ready for the next build!")
            
        except Exception as e:
            messagebox.showerror("Sync Error", f"Error syncing changes: {e}")

    def _clone_nvm_images(self) -> None:
        """Download nd_pae_sw-ccd_lan_nvm_images into GBE_Image/ — works on any PC."""
        import threading
        import shutil
        import zipfile
        import urllib.request

        REPO_URL  = "https://github.com/michaele1991/nd_pae_sw-ccd_lan_nvm_images"
        ZIP_URL   = f"{REPO_URL}/archive/refs/heads/master.zip"
        workspace_root = Path(__file__).parent.parent
        gbe_image_root = workspace_root / "GBE_Image"

        # Confirm if GBE_Image already has content
        existing = [f for f in gbe_image_root.iterdir() if f.is_dir()] if gbe_image_root.exists() else []
        if existing:
            if not messagebox.askyesno(
                "Clone NVM Images",
                f"GBE_Image already contains {len(existing)} project folder(s).\n"
                "Download will add/update from the repository.\nContinue?"
            ):
                return

        # Progress dialog
        dlg = tk.Toplevel(self)
        dlg.title("Downloading NVM Images")
        dlg.geometry("480x160")
        dlg.resizable(False, False)
        dlg.grab_set()
        tk.Label(dlg, text="Downloading NVM image repository…", font=('Segoe UI', 10)).pack(pady=(18, 6))
        prog = ttk.Progressbar(dlg, mode='indeterminate', length=400)
        prog.pack(pady=4)
        status_var = tk.StringVar(value="Connecting…")
        tk.Label(dlg, textvariable=status_var, font=('Segoe UI', 9), fg='#555').pack(pady=4)
        prog.start(12)

        def do_download():
            try:
                tmp_zip  = workspace_root / "_nvm_download.zip"
                tmp_dir  = workspace_root / "_nvm_extract_tmp"

                # ── Try git clone first (faster, incremental) ──────────────
                git_ok = False
                self.after(0, lambda: status_var.set("Trying git clone…"))
                git_result = subprocess.run(
                    ["git", "clone", "--depth", "1", "-b", "master", REPO_URL, str(tmp_dir)],
                    capture_output=True, text=True
                )
                if git_result.returncode == 0:
                    git_ok = True
                else:
                    # ── Fallback: download ZIP via urllib (no git needed) ──
                    self.after(0, lambda: status_var.set("Downloading ZIP archive…"))
                    if tmp_dir.exists():
                        shutil.rmtree(tmp_dir)
                    urllib.request.urlretrieve(ZIP_URL, str(tmp_zip))

                    self.after(0, lambda: status_var.set("Extracting…"))
                    tmp_dir.mkdir(parents=True, exist_ok=True)
                    with zipfile.ZipFile(str(tmp_zip), 'r') as zf:
                        zf.extractall(str(tmp_dir))
                    tmp_zip.unlink(missing_ok=True)

                    # ZIP extracts into a single subfolder e.g. "nd_pae_sw-...-master/"
                    subdirs = [d for d in tmp_dir.iterdir() if d.is_dir()]
                    if len(subdirs) == 1:
                        # Flatten: move contents of that subfolder up
                        inner = subdirs[0]
                        for item in list(inner.iterdir()):
                            item.rename(tmp_dir / item.name)
                        inner.rmdir()

                # ── Move project folders into GBE_Image/ ───────────────────
                self.after(0, lambda: status_var.set("Installing project folders…"))
                gbe_image_root.mkdir(parents=True, exist_ok=True)
                moved = 0
                for item in tmp_dir.iterdir():
                    if item.name.startswith('.') or item.name in ('.git', '__MACOSX'):
                        continue
                    dest = gbe_image_root / item.name
                    if dest.exists():
                        shutil.rmtree(dest) if dest.is_dir() else dest.unlink()
                    shutil.move(str(item), str(dest))
                    moved += 1

                shutil.rmtree(tmp_dir, ignore_errors=True)
                method = "git clone" if git_ok else "ZIP download"
                self.after(0, lambda: _finish(True, f"{moved} item(s) installed via {method}"))

            except Exception as exc:
                for p in [workspace_root / "_nvm_download.zip", workspace_root / "_nvm_extract_tmp"]:
                    if p.exists():
                        shutil.rmtree(p, ignore_errors=True) if p.is_dir() else p.unlink(missing_ok=True)
                self.after(0, lambda: _finish(False, str(exc)))

        def _finish(ok: bool, msg: str):
            prog.stop()
            dlg.destroy()
            if ok:
                messagebox.showinfo("Download Complete", f"NVM images installed successfully.\n{msg}")
                self._load_gbe_folders()
            else:
                messagebox.showerror("Download Failed", f"Failed to download NVM images:\n{msg}")

        threading.Thread(target=do_download, daemon=True).start()

    def _open_new_project_dialog(self) -> None:
        """Open a dialog to create a new GBE project folder from an existing template."""
        workspace_root = Path(__file__).parent.parent
        gbe_image_root = workspace_root / "GBE_Image"
        existing = sorted([f.name for f in gbe_image_root.iterdir() if f.is_dir()]) if gbe_image_root.exists() else []

        dlg = tk.Toplevel(self)
        dlg.title("New Project")
        dlg.geometry("520x580")
        dlg.resizable(False, False)
        dlg.grab_set()

        # Title bar
        hdr = tk.Frame(dlg, bg='#0071c5', height=40)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Create New GBE Project", bg='#0071c5', fg='white',
                 font=('Segoe UI', 12, 'bold')).pack(side=tk.LEFT, padx=16, pady=8)

        body = ttk.Frame(dlg, padding=18)
        body.pack(fill=tk.BOTH, expand=True)

        def _row(parent, label, var, row, placeholder=""):
            ttk.Label(parent, text=label, font=('Segoe UI', 9)).grid(row=row, column=0, sticky='w', pady=4, padx=(0, 12))
            e = ttk.Entry(parent, textvariable=var, width=28)
            e.grid(row=row, column=1, sticky='ew', pady=4)
            return e

        body.columnconfigure(1, weight=1)

        folder_var   = tk.StringVar()
        silicon_var  = tk.StringVar(value=self.silicon_var.get())
        step_var     = tk.StringVar(value=self.step_var.get())
        major_var    = tk.StringVar(value=self.major_var.get())
        minor_var    = tk.StringVar(value=self.minor_var.get())
        device_var   = tk.StringVar(value=self.device_id_var.get())
        sku_var      = tk.StringVar(value=self.sku_device_id_var.get())
        output_var   = tk.StringVar()
        lan_var      = tk.StringVar(value=self.lan_sw_var.get())
        lmv_var      = tk.StringVar(value=self.lm_v_var.get())
        template_var = tk.StringVar(value=existing[0] if existing else "")

        # Auto-fill NVM output from folder name
        def _sync_output(*_):
            if not output_var.get() or output_var.get() == folder_var.get():
                output_var.set(folder_var.get())
        folder_var.trace_add("write", _sync_output)

        r = 0
        _row(body, "Project Folder Name:", folder_var, r); r += 1
        _row(body, "NVM Output Name:",     output_var,  r); r += 1
        _row(body, "Silicon:",             silicon_var, r); r += 1
        _row(body, "Step:",                step_var,    r); r += 1
        _row(body, "Major Version:",       major_var,   r); r += 1
        _row(body, "Minor Version:",       minor_var,   r); r += 1
        _row(body, "Device ID:",           device_var,  r); r += 1
        _row(body, "SKU Device ID:",       sku_var,     r); r += 1

        # LAN SW radio
        ttk.Label(body, text="LAN Mode:", font=('Segoe UI', 9)).grid(row=r, column=0, sticky='w', pady=4)
        lan_frame = ttk.Frame(body)
        lan_frame.grid(row=r, column=1, sticky='w'); r += 1
        ttk.Radiobutton(lan_frame, text="LAN SW",     value="lan",     variable=lan_var).pack(side=tk.LEFT)
        ttk.Radiobutton(lan_frame, text="Non-LAN SW", value="non_lan", variable=lan_var).pack(side=tk.LEFT, padx=8)

        # LM/V radio
        ttk.Label(body, text="LM/V Mode:", font=('Segoe UI', 9)).grid(row=r, column=0, sticky='w', pady=4)
        lmv_frame = ttk.Frame(body)
        lmv_frame.grid(row=r, column=1, sticky='w'); r += 1
        ttk.Radiobutton(lmv_frame, text="V",    value="V",    variable=lmv_var).pack(side=tk.LEFT)
        ttk.Radiobutton(lmv_frame, text="LM",   value="LM",   variable=lmv_var).pack(side=tk.LEFT, padx=8)
        ttk.Radiobutton(lmv_frame, text="Both", value="Both", variable=lmv_var).pack(side=tk.LEFT)

        # Template selector
        ttk.Label(body, text="Copy XLSM from:", font=('Segoe UI', 9)).grid(row=r, column=0, sticky='w', pady=4)
        tmpl_cb = ttk.Combobox(body, textvariable=template_var, values=existing, state="readonly", width=28)
        tmpl_cb.grid(row=r, column=1, sticky='ew', pady=4); r += 1

        err_var = tk.StringVar()
        err_lbl = ttk.Label(body, textvariable=err_var, foreground='red', font=('Segoe UI', 8), wraplength=420)
        err_lbl.grid(row=r, column=0, columnspan=2, sticky='w', pady=(4, 0)); r += 1

        # Buttons
        btn_frame = ttk.Frame(dlg, padding=(18, 0, 18, 14))
        btn_frame.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Button(btn_frame, text="Cancel", command=dlg.destroy).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(btn_frame, text="Create Project", style='Accent.TButton',
                   command=lambda: _do_create()).pack(side=tk.RIGHT)

        def _do_create():
            folder_name = folder_var.get().strip()
            if not folder_name:
                err_var.set("Project folder name is required."); return

            dest = gbe_image_root / folder_name
            if dest.exists():
                err_var.set(f"Folder already exists: {dest}"); return

            tmpl_name = template_var.get()
            if not tmpl_name:
                err_var.set("Select a template project to copy XLSM from."); return

            tmpl_dir = gbe_image_root / tmpl_name
            xlsm_files = list(tmpl_dir.glob("*.xlsm"))
            if not xlsm_files:
                err_var.set(f"No .xlsm found in template: {tmpl_name}"); return

            import shutil
            try:
                dest.mkdir(parents=True)

                # Copy Build.bat and calc_csum.py verbatim
                for fname in ("Build.bat", "calc_csum.py"):
                    src_f = tmpl_dir / fname
                    if src_f.exists():
                        shutil.copy2(str(src_f), str(dest / fname))

                # Copy XLSM, rename to <folder_name_lower>_nvm_map.xlsm
                new_xlsm_name = folder_name.lower() + "_nvm_map.xlsm"
                shutil.copy2(str(xlsm_files[0]), str(dest / new_xlsm_name))

                # Generate run_excel_macro.vbs with updated filenames
                vbs_content = (
                    "Option Explicit\n\n"
                    "    LaunchMacro\n\n"
                    "    Sub LaunchMacro() \n"
                    "      Dim xl\n"
                    "      Dim xlBook      \n"
                    "      Dim sCurPath\n\n"
                    "        sCurPath = CreateObject(\"Scripting.FileSystemObject\").GetAbsolutePathName(\".\")\n"
                    f"        Set xl = CreateObject(\"Excel.application\")\n"
                    f"        Set xlBook = xl.Workbooks.Open(sCurPath & \"\\{new_xlsm_name}\", 0, True)   \n"
                    "        xl.Application.Visible = False\n"
                    f"        xl.Application.run \"{new_xlsm_name}!Module1.genNvmCMDline\"\n"
                    "        xl.DisplayAlerts = False        \n"
                    "        xlBook.Save = True\n"
                    "        xl.activewindow.close\n"
                    "        xl.Quit\n"
                    "        Set xlBook = Nothing\n"
                    "        Set xl = Nothing\n"
                    "        End Sub \n"
                )
                (dest / "run_excel_macro.vbs").write_text(vbs_content, encoding="utf-8")

                # Write parameters.json with the entered values
                import json as _json
                params = {
                    "projectName":   folder_name,
                    "nvmOutput":     output_var.get().strip() or folder_name,
                    "silicon":       silicon_var.get().strip(),
                    "step":          step_var.get().strip(),
                    "version":       {"major": major_var.get().strip(), "minor": minor_var.get().strip()},
                    "deviceId":      device_var.get().strip(),
                    "skuDeviceId":   sku_var.get().strip(),
                    "lanMode":       lan_var.get(),
                    "lmvMode":       lmv_var.get(),
                }
                (dest / "parameters.json").write_text(_json.dumps(params, indent=2), encoding="utf-8")

                dlg.destroy()
                self._load_gbe_folders()
                # Select the new project
                self.folder_var.set(folder_name)
                self.open_folder()
                messagebox.showinfo("New Project", f"Project '{folder_name}' created successfully.")

            except Exception as ex:
                err_var.set(f"Error: {ex}")

    def _load_gbe_folders(self) -> None:
        """Scan GBE_Image folder for project subfolders."""
        workspace_root = Path(__file__).parent.parent
        gbe_image_root = workspace_root / "GBE_Image"
        
        if not gbe_image_root.exists():
            gbe_image_root.mkdir(parents=True, exist_ok=True)
        
        self.gbe_folders = [f for f in gbe_image_root.iterdir() if f.is_dir() and not f.name.startswith(".")]
        self.folder_combo["values"] = [f.name for f in self.gbe_folders]
        
        if self.gbe_folders and not self.folder_var.get():
            self.folder_combo.current(0)
            self.open_folder()

    def open_folder(self) -> None:
        """Load the selected project folder."""
        if not self.folder_var.get():
            return
        
        # Find the folder by name
        folder_name = self.folder_var.get()
        folder_path = None
        for f in self.gbe_folders:
            if f.name == folder_name:
                folder_path = f
                break
        
        if not folder_path or not folder_path.exists():
            messagebox.showerror("Error", f"Folder not found: {folder_name}")
            return
        
        self.current_folder = folder_path
        self._load_xlsm()
        self.load_registers(auto=True)
        self._set_status(f"Project: {folder_path.name}")

    def _load_xlsm(self) -> None:
        """Find and load the Excel workbook in the current folder."""
        if not self.current_folder:
            return
        xlsm_files = list(self.current_folder.glob("*.xlsm"))
        if xlsm_files:
            self.current_xlsm = xlsm_files[0]
        else:
            self.current_xlsm = None

    def scan_bins(self, silent: bool = False) -> None:
        """Deprecated - no longer used without bin UI."""
        pass

    def load_selected_bin(self) -> None:
        if not self.bin_var.get():
            return
        self.current_bin = Path(self.bin_var.get())
        if not self.current_bin.exists():
            messagebox.showerror("Missing file", "Selected .bin file does not exist.")
            return
        self.words = self.read_bin(self.current_bin)
        self.refresh_word_list()
        self.detect_txt()

    def detect_txt(self) -> None:
        if not self.current_bin:
            return
        candidate = self.current_bin.with_suffix(".txt")
        if candidate.exists():
            self.current_txt = candidate
            self.txt_var.set(str(candidate))
        else:
            self.current_txt = None
            self.txt_var.set("")

    def pick_txt_file(self) -> None:
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if not file_path:
            return
        self.current_txt = Path(file_path)
        self.txt_var.set(file_path)

    def read_bin(self, path: Path) -> list[int]:
        data = path.read_bytes()
        words = []
        for i in range(0, len(data), 2):
            if i + 2 <= len(data):
                words.append(int.from_bytes(data[i:i + 2], "little"))
        return words

    def write_bin(self, path: Path) -> None:
        data = b"".join(w.to_bytes(2, "little") for w in self.words)
        path.write_bytes(data)

    def refresh_word_list(self) -> None:
        self.tree.delete(*self.tree.get_children())
        for idx, word in enumerate(self.words):
            self.tree.insert("", tk.END, values=(idx, f"0x{word:04X}"))

    def on_word_select(self, _event: tk.Event) -> None:
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        if not values:
            return
        index = int(values[0])
        # Ensure words list is large enough
        while len(self.words) <= index:
            self.words.append(0xFFFF)
        self.index_var.set(str(index))
        self.value_var.set(f"0x{self.words[index]:04X}")

    def parse_index(self) -> int:
        text = self.index_var.get().strip().lower()
        if text.startswith("0x"):
            return int(text, 16)
        return int(text)

    def parse_value(self) -> int:
        text = self.value_var.get().strip().lower()
        if not text.startswith("0x"):
            raise ValueError("Value must be in hex format (0x1234).")
        val = int(text, 16)
        if not 0 <= val <= 0xFFFF:
            raise ValueError("Value must be a 16-bit number.")
        return val

    def apply_value(self) -> None:
        try:
            index = self.parse_index()
            value = self.parse_value()
        except Exception as exc:
            messagebox.showerror("Invalid input", str(exc))
            return
        if not 0 <= index < len(self.words):
            messagebox.showerror("Invalid index", "Word index out of range.")
            return
        self.words[index] = value
        self.refresh_word_list()

    def _update_excel_word(self, word_index: int, value: int) -> None:
        """Update the Excel workbook with the new word value."""
        if not self.current_xlsm or not self.current_xlsm.exists():
            return
        
        wb = load_workbook(str(self.current_xlsm), keep_vba=True)
        if "full nvm map" not in wb.sheetnames:
            wb.close()
            return
        
        ws = wb["full nvm map"]
        excel_mode = self.excel_col_var.get()
        col_letters = []
        if excel_mode == "V":
            col_letters = ["G"]
        elif excel_mode == "LM":
            col_letters = ["H"]
        else:  # Both
            col_letters = ["G", "H"]
        
        # Find the row(s) for this word index
        row = 6
        while ws.cell(row, 1).value is not None:
            cell_word = ws.cell(row, 1).value
            # Check if this matches our word index (could be hex like "00h" or decimal)
            try:
                if isinstance(cell_word, str):
                    cell_word = cell_word.replace("h", "").strip()
                    cell_word_num = int(cell_word, 16)
                else:
                    cell_word_num = int(cell_word)
                
                if cell_word_num == word_index:
                    # Check if this is the last range for this word (15:0 typically)
                    bit_range = ws.cell(row, 2).value
                    if bit_range and ":" in str(bit_range):
                        parts = str(bit_range).split(":")
                        if parts[0].strip() == "15" and parts[1].strip() == "0":
                            # This is the full word, update it in selected columns
                            for col_letter in col_letters:
                                ws.cell(row, ord(col_letter) - ord('A') + 1).value = f"0x{value:04X}"
                            break
            except:
                pass
            row += 1
        
        wb.save(str(self.current_xlsm))
        wb.close()

    def set_clear_bit(self, set_bit: bool) -> None:
        try:
            # Get index from register name or direct input
            if self.register_name_var.get().strip():
                reg_name = self.register_name_var.get().strip()
                reg = next((r for r in self.registers if r.name == reg_name), None)
                if not reg:
                    messagebox.showerror("Invalid register", f"Register '{reg_name}' not found.")
                    return
                index = reg.word_index
            elif self.index_var.get().strip():
                index = self.parse_index()
            else:
                messagebox.showerror("Invalid input", "Enter a register name or select a register from the list.")
                return

            bit_str = self.bit_var.get().strip()
            if not bit_str:
                messagebox.showerror("Invalid input", "Enter a bit number (0-15) in the 'Bit Number' field.")
                return
            bit = int(bit_str)
        except ValueError:
            messagebox.showerror("Invalid input", "Bit number must be an integer between 0 and 15.")
            return
        except Exception as exc:
            messagebox.showerror("Invalid input", f"Enter a register name and bit number (0-15). Error: {exc}")
            return
        
        if not 0 <= bit <= 15:
            messagebox.showerror("Invalid bit", "Bit must be between 0 and 15.")
            return
        
        # Ensure words list is large enough
        while len(self.words) <= index:
            self.words.append(0xFFFF)
        
        mask = 1 << bit
        if set_bit:
            self.words[index] |= mask
        else:
            self.words[index] &= ~mask
        new_val = self.words[index]
        self.value_var.set(f"0x{new_val:04X}")

        # Mirror change into words_v and words_lm so build picks it up
        while len(self.words_v) <= index:
            self.words_v.append(0xFFFF)
        while len(self.words_lm) <= index:
            self.words_lm.append(0xFFFF)
        col = self.excel_col_var.get()
        if col in ("V", "Both"):
            self.words_v[index] = new_val
            s, c = self.dirty_bits_v.get(index, (0, 0))
            if set_bit:
                s |= (1 << bit); c &= ~(1 << bit)
            else:
                c |= (1 << bit); s &= ~(1 << bit)
            self.dirty_bits_v[index] = (s, c)
        if col in ("LM", "Both"):
            self.words_lm[index] = new_val
            s, c = self.dirty_bits_lm.get(index, (0, 0))
            if set_bit:
                s |= (1 << bit); c &= ~(1 << bit)
            else:
                c |= (1 << bit); s &= ~(1 << bit)
            self.dirty_bits_lm[index] = (s, c)

        self.refresh_word_list()

    def update_checksum(self) -> None:
        if len(self.words) <= CHECKSUM_WORD:
            messagebox.showerror("Checksum", "Binary file is too small to update checksum.")
            return
        total = sum(self.words[0:CHECKSUM_WORD])
        checksum = (-total + (2 ** 32) + CHECKSUM_VALUE) & 0xFFFF
        self.words[CHECKSUM_WORD] = checksum
        self.refresh_word_list()
        messagebox.showinfo("Checksum", f"Checksum updated to 0x{checksum:04X}")

    def update_txt_checksum(self) -> None:
        if not self.current_txt or not self.current_txt.exists():
            messagebox.showerror("TXT", "No TXT file selected.")
            return
        if len(self.words) <= CHECKSUM_WORD:
            messagebox.showerror("TXT", "Binary file is too small to update checksum.")
            return
        checksum = self.words[CHECKSUM_WORD]
        lines = self.current_txt.read_text(encoding="utf-8").splitlines()
        if len(lines) <= CHECKSUM_LINE:
            messagebox.showerror("TXT", "TXT file does not have the expected checksum line.")
            return
        parts = lines[CHECKSUM_LINE].split()
        if not parts:
            messagebox.showerror("TXT", "Checksum line is empty.")
            return
        parts[-1] = f"{checksum:04X}"
        lines[CHECKSUM_LINE] = " ".join(parts)
        self.current_txt.write_text("\n".join(lines) + "\n", encoding="utf-8")
        messagebox.showinfo("TXT", "TXT checksum updated.")

    def save_bin(self) -> None:
        if not self.current_bin:
            messagebox.showerror("Save", "No bin file loaded.")
            return
        self.write_bin(self.current_bin)
        messagebox.showinfo("Save", "Bin file saved.")

    def save_bin_as(self) -> None:
        file_path = filedialog.asksaveasfilename(defaultextension=".bin", filetypes=[("Bin files", "*.bin")])
        if not file_path:
            return
        self.write_bin(Path(file_path))
        messagebox.showinfo("Save As", "Bin file saved.")

    def load_from_excel(self) -> None:
        """Load word values from Excel workbook into memory."""
        if not self.current_xlsm or not self.current_xlsm.exists():
            messagebox.showerror("Load from Excel", "No Excel workbook found. Please select a project folder first.")
            return

        if load_workbook is None:
            messagebox.showerror("Load from Excel", "openpyxl is not installed.")
            return

        self._set_status(f"Loading Excel data from {self.current_xlsm.name}\u2026")
        try:
            wb = load_workbook(str(self.current_xlsm), keep_vba=False, data_only=True)
            
            # Load parameters from 'general variable' sheet if it exists
            if "general variable" in wb.sheetnames:
                self._load_general_variables(wb["general variable"])
            
            if "full nvm map" not in wb.sheetnames:
                messagebox.showerror("Load from Excel", "'full nvm map' sheet not found in workbook.")
                wb.close()
                return
            
            ws = wb["full nvm map"]
            
            # Read both V and LM words from Excel
            self.words_v = []
            self.words_lm = []
            self.words = []  # Default words array
            self.excel_data = []
            # Bit-delta dirty tracking: {word_idx: (set_mask, clear_mask)}
            # Only user-changed bits are recorded; applied as overlay on macro output at build time
            self.dirty_bits_v: dict = {}
            self.dirty_bits_lm: dict = {}
            row = 6  # Start at row 6 (after headers)
            
            while ws.cell(row, 1).value is not None:
                cell_word = ws.cell(row, 1).value
                
                # Parse word index
                try:
                    if isinstance(cell_word, str):
                        cell_word = cell_word.replace("h", "").strip()
                        word_idx = int(cell_word, 16)
                    else:
                        word_idx = int(cell_word)
                except:
                    row += 1
                    continue
                
                # Read Excel row data for display (all 13 columns)
                offset_val = ws.cell(row, 1).value  # Column A: Offset
                bits_val = ws.cell(row, 2).value    # Column B: bits
                rtl_name_val = ws.cell(row, 3).value  # Column C: name in RTL
                name_val = ws.cell(row, 4).value    # Column D: name in c-spec
                bits_owner_val = ws.cell(row, 5).value  # Column E: bits owner
                description_val = ws.cell(row, 6).value  # Column F: description
                v_val = ws.cell(row, 7).value       # Column G: V version
                lm_val = ws.cell(row, 8).value      # Column H: LM version
                values_read_val = ws.cell(row, 9).value  # Column I: Values read
                criteria_val = ws.cell(row, 10).value  # Column J: criteria for change
                final_val = ws.cell(row, 11).value  # Column K: final value
                comments_val = ws.cell(row, 12).value  # Column L: comments
                changed_ver_val = ws.cell(row, 13).value  # Column M: changed in version
                
                # Only store rows with meaningful names (skip N/A and empty)
                if name_val and str(name_val).strip() not in ['N/A', 'None', 'name in RTL', 'name in c-spec']:
                    self.excel_data.append({
                        'offset': str(offset_val) if offset_val else '',
                        'bits': str(bits_val) if bits_val else '',
                        'rtl_name': str(rtl_name_val) if rtl_name_val else '',
                        'name': str(name_val) if name_val else '',
                        'bits_owner': str(bits_owner_val) if bits_owner_val else '',
                        'description': str(description_val) if description_val else '',
                        'v': str(v_val) if v_val else '',
                        'lm': str(lm_val) if lm_val else '',
                        'values_read': str(values_read_val) if values_read_val else '',
                        'criteria': str(criteria_val) if criteria_val else '',
                        'final_value': str(final_val) if final_val else '',
                        'comments': str(comments_val) if comments_val else '',
                        'changed_version': str(changed_ver_val) if changed_ver_val else '',
                        'word_idx': word_idx
                    })
                
                # Read bit field values for ALL rows (not just 15:0)
                bit_range = ws.cell(row, 2).value
                if bit_range and ":" in str(bit_range):
                    try:
                        parts = str(bit_range).split(":")
                        high_bit = int(parts[0].strip())
                        low_bit = int(parts[1].strip())
                        num_bits = high_bit - low_bit + 1
                        mask = (1 << num_bits) - 1
                        
                        # Ensure word arrays are large enough
                        while len(self.words_v) <= word_idx:
                            self.words_v.append(0xFFFF)
                        while len(self.words_lm) <= word_idx:
                            self.words_lm.append(0xFFFF)
                        
                        # Read V value (column G = 7)
                        v_cell = ws.cell(row, 7).value
                        if v_cell and str(v_cell).upper() not in ['N/A', 'NONE', '']:
                            v_str = str(v_cell).replace("0x", "").replace("0X", "").strip()
                            try:
                                v_value = int(v_str, 16)
                                # Update word with bit-level precision
                                current_word = self.words_v[word_idx]
                                current_word &= ~(mask << low_bit)  # Clear target bits
                                current_word |= (v_value & mask) << low_bit  # Set new bits
                                self.words_v[word_idx] = current_word
                            except:
                                pass
                        
                        # Read LM value (column H = 8)
                        lm_cell = ws.cell(row, 8).value
                        if lm_cell and str(lm_cell).upper() not in ['N/A', 'NONE', '']:
                            lm_str = str(lm_cell).replace("0x", "").replace("0X", "").strip()
                            try:
                                lm_value = int(lm_str, 16)
                                # Update word with bit-level precision
                                current_word = self.words_lm[word_idx]
                                current_word &= ~(mask << low_bit)  # Clear target bits
                                current_word |= (lm_value & mask) << low_bit  # Set new bits
                                self.words_lm[word_idx] = current_word
                            except:
                                pass
                    except:
                        pass
                
                row += 1
            
            # Load other sheets
            self._load_additional_sheets(wb)
            
            wb.close()

            # Override words_v / words_lm from Official .bin files if present
            # This ensures the baseline is the official reference, not stale Excel cache
            if self.current_folder:
                official_dir = self.current_folder / "Official_images"
                if official_dir.exists():
                    for subdir in official_dir.iterdir():
                        if not subdir.is_dir():
                            continue
                        bins = list(subdir.glob("*.bin"))
                        if not bins:
                            continue
                        raw = bins[0].read_bytes()
                        loaded = [int.from_bytes(raw[i:i+2], "little") for i in range(0, len(raw)-1, 2)]
                        name_lower = subdir.name.lower()
                        if "cons" in name_lower:
                            self.words_v = loaded
                        elif "corp" in name_lower:
                            self.words_lm = loaded

            # Set default words to V version
            self.words = self.words_v.copy() if self.words_v else []
            
            self.refresh_word_list()
            self.refresh_excel_tree()
            self.refresh_additional_trees()
            self.update_notes_with_parameters()
            self._populate_registers_from_excel(silent=True)
            self._set_status(
                f"Loaded: {self.current_xlsm.name}  \u2014  "
                f"{len(self.excel_data)} registers  |  "
                f"V: {len(self.words_v)} words  |  LM: {len(self.words_lm)} words"
            )

            load_msg = (
                f"Excel data loaded successfully!\n\n"
                f"  Registers loaded:      {len(self.excel_data)}\n"
                f"  V (Corporate) words:   {len(self.words_v)}\n"
                f"  LM (Consumer) words:   {len(self.words_lm)}\n\n"
                f"All bit-field values are now in memory."
            )
            messagebox.showinfo("Load from Excel", load_msg)

        except Exception as exc:
            self._set_status(f"Error loading Excel: {exc}")
            messagebox.showerror("Load from Excel", f"Error: {exc}")
    
    def refresh_excel_tree(self) -> None:
        """Populate Excel data tree with loaded data."""
        self.excel_tree.delete(*self.excel_tree.get_children())
        
        # Determine which columns to display based on show_less checkbox
        columns_to_use = self.excel_columns_simple if self.show_less_var.get() else self.excel_columns_full
        
        for idx, row_data in enumerate(self.excel_data):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            
            # Build values tuple dynamically based on current columns
            # Each col is a tuple: (key, display_name, width)
            values = []
            for col_key, col_name, col_width in columns_to_use:
                val = row_data.get(col_key, 'N/A')
                values.append(val if val not in [None, ''] else 'N/A')
            
            self.excel_tree.insert("", tk.END, values=tuple(values), tags=(tag,))
    
    def on_excel_row_select(self, _event: tk.Event) -> None:
        """Show bit description when selecting a row in Excel data."""
        sel = self.excel_tree.selection()
        if not sel:
            return
        
        # Get selected item index
        item = sel[0]
        idx = self.excel_tree.index(item)
        
        if idx < len(self.excel_data):
            row_data = self.excel_data[idx]
            self.selected_row_idx = idx
            
            # Show only the description text
            if row_data.get('description') and str(row_data['description']) not in ['N/A', 'None', '']:
                desc_text = row_data['description']
            else:
                desc_text = "(No description available)"
            
            # Update description text widget
            self.bit_desc_text.delete(1.0, tk.END)
            self.bit_desc_text.insert(1.0, desc_text)
            
            # Update edit fields with current values
            self.v_value_var.set(row_data.get('v', ''))
            self.lm_value_var.set(row_data.get('lm', ''))

    def save_bit_changes(self) -> None:
        """Save the edited V and/or LM values to the word data."""
        if self.selected_row_idx is None or self.selected_row_idx >= len(self.excel_data):
            messagebox.showwarning("Save Changes", "Please select a bit to edit.")
            return
        
        row_data = self.excel_data[self.selected_row_idx]
        
        # Check which values to update
        update_v = self.update_v_var.get()
        update_lm = self.update_lm_var.get()
        
        if not update_v and not update_lm:
            messagebox.showwarning("Save Changes", "Please check V and/or LM to update.")
            return
        
        # Get word index and bit range
        word_idx = row_data['word_idx']
        bits_str = row_data['bits']
        
        try:
            # Parse bit range
            if ":" in bits_str:
                parts = bits_str.split(":")
                high_bit = int(parts[0].strip())
                low_bit = int(parts[1].strip())
            else:
                # Single bit
                high_bit = low_bit = int(bits_str.strip())
            
            # Create bit mask
            num_bits = high_bit - low_bit + 1
            mask = (1 << num_bits) - 1
            
            # Ensure we have enough words in both arrays
            while len(self.words_v) <= word_idx:
                self.words_v.append(0xFFFF)
            while len(self.words_lm) <= word_idx:
                self.words_lm.append(0xFFFF)
            
            updated = []
            
            # Update V value if checked
            if update_v:
                new_v = self.v_value_var.get().strip()
                if not new_v:
                    messagebox.showwarning("Save Changes", "Please enter a V value.")
                    return
                
                new_v_str = new_v.replace("0x", "").replace("0X", "").strip()
                new_v_int = int(new_v_str, 16)
                
                # Update V word array
                current_word = self.words_v[word_idx]
                current_word &= ~(mask << low_bit)
                current_word |= (new_v_int & mask) << low_bit
                self.words_v[word_idx] = current_word
                # Record dirty bit delta
                fms = mask << low_bit
                s, c = self.dirty_bits_v.get(word_idx, (0, 0))
                c |= fms
                s = (s & ~fms) | ((new_v_int & mask) << low_bit)
                self.dirty_bits_v[word_idx] = (s, c)
                # Update excel_data
                self.excel_data[self.selected_row_idx]['v'] = new_v
                updated.append('V')
            
            # Update LM value if checked
            if update_lm:
                new_lm = self.lm_value_var.get().strip()
                if not new_lm:
                    messagebox.showwarning("Save Changes", "Please enter an LM value.")
                    return
                
                new_lm_str = new_lm.replace("0x", "").replace("0X", "").strip()
                new_lm_int = int(new_lm_str, 16)
                
                # Update LM word array
                current_word = self.words_lm[word_idx]
                current_word &= ~(mask << low_bit)
                current_word |= (new_lm_int & mask) << low_bit
                self.words_lm[word_idx] = current_word
                # Record dirty bit delta
                fms = mask << low_bit
                s, c = self.dirty_bits_lm.get(word_idx, (0, 0))
                c |= fms
                s = (s & ~fms) | ((new_lm_int & mask) << low_bit)
                self.dirty_bits_lm[word_idx] = (s, c)
                # Update excel_data
                self.excel_data[self.selected_row_idx]['lm'] = new_lm
                updated.append('LM')
            
            # Update default words array (use V version)
            self.words = self.words_v.copy()
            
            # Refresh the tree display
            self.refresh_excel_tree()
            
            updated_str = " and ".join(updated)
            messagebox.showinfo("Save Changes", f"Successfully updated {updated_str} value(s) for {row_data['name']}")
            
        except ValueError as e:
            messagebox.showerror("Save Changes", f"Invalid hex value: {e}")
        except Exception as e:
            messagebox.showerror("Save Changes", f"Error updating value: {e}")

    def _write_words_to_excel(self, variant: str, nvm_name: str, major: str, minor: str) -> None:
        """Write ONLY build parameters (B2-B12) to Excel. Never touches nvm_map G/H cells."""
        wb = load_workbook(str(self.current_xlsm), keep_vba=True)
        ws_var = wb["general variable"]
        ws_var["B2"] = nvm_name
        ws_var["B3"] = self.silicon_var.get().strip()
        ws_var["B4"] = self.step_var.get().strip()
        ws_var["B5"] = nvm_name
        ws_var["B7"] = int(major + minor.zfill(2))
        ws_var["B8"] = self.device_id_var.get().strip()
        ws_var["B9"] = self.sku_device_id_var.get().strip()
        ws_var["B10"] = (variant == "LM")
        ws_var["B11"] = (variant == "V")
        ws_var["B12"] = (self.lan_sw_var.get() == "lan")
        wb.save(str(self.current_xlsm))
        wb.close()

    def _apply_dirty_bits_to_output(self, folder: Path, variant: str) -> None:
        """
        After the macro writes output files, patch only user-changed bits into them.
        Uses bit-delta (set_mask, clear_mask) — macro baseline is preserved for all other bits.
        """
        dirty = self.dirty_bits_v if variant == "V" else self.dirty_bits_lm
        if not dirty:
            return
        bin_files = list(folder.glob("*.bin"))
        txt_files = list(folder.glob("*.txt"))
        if not bin_files:
            return
        bin_path = bin_files[0]

        raw = bytearray(bin_path.read_bytes())
        for word_idx, (set_mask, clear_mask) in dirty.items():
            byte_off = word_idx * 2
            if byte_off + 2 > len(raw):
                continue
            word = int.from_bytes(raw[byte_off:byte_off+2], "little")
            word = (word | set_mask) & ~clear_mask
            raw[byte_off:byte_off+2] = word.to_bytes(2, "little")
        bin_path.write_bytes(bytes(raw))

        # Rebuild txt from the patched binary
        if txt_files:
            words = [int.from_bytes(raw[i:i+2], "little") for i in range(0, len(raw)-1, 2)]
            lines = []
            for i in range(0, len(words), 8):
                lines.append(" ".join(f"{words[i+j]:04X}" for j in range(min(8, len(words)-i))))
            txt_files[0].write_text("\n".join(lines) + "\n", encoding="utf-8")

    def run_build_flow(self) -> None:
        """Build NVM: write params to Excel, run VBA macro per variant, update checksums."""
        if not self.current_folder or not self.current_xlsm:
            messagebox.showerror("Build NVM", "No project folder / Excel workbook loaded.\nSelect a project from the dropdown first.")
            return
        if load_workbook is None:
            messagebox.showerror("Build NVM", "openpyxl is not installed.")
            return

        version_selection = self.lm_v_var.get()
        build_v  = version_selection in ["V",  "Both"]
        build_lm = version_selection in ["LM", "Both"]

        if not build_v and not build_lm:
            messagebox.showerror("Build NVM", "Please select V, LM, or Both.")
            return

        nvm_name = self.nvm_output_var.get().strip() or self.project_var.get().strip() or "GBE_Output"
        major    = self.major_var.get().strip() or "1"
        minor    = self.minor_var.get().strip() or "0"
        version  = f"{major}.{minor}"

        lines_preview = []
        if build_v:  lines_preview.append(f"  V  (Consumer)  ->  {nvm_name}_{version}_Release_Cons_Prod_NA")
        if build_lm: lines_preview.append(f"  LM (Corporate) ->  {nvm_name}_{version}_Release_Corp_Prod_NA")

        if messagebox.askquestion(
            "Build NVM",
            "Generate NVM image files via Excel macro:\n\n"
            + "\n".join(lines_preview)
            + "\n\nContinue?",
            icon='question'
        ) != 'yes':
            return

        try:
            import win32com.client
            import pythoncom
            import time

            variants = []
            if build_v:  variants.append("V")
            if build_lm: variants.append("LM")

            for variant in variants:
                label = "Consumer (V)" if variant == "V" else "Corporate (LM)"
                self._set_status(f"Writing parameters + word values for {label}...")

                # Step 1: write parameters AND all modified word values via openpyxl
                self._write_words_to_excel(variant, nvm_name, major, minor)

                # Step 2: run genNvmCMDline macro via Excel COM
                self._set_status(f"Running NVM macro for {label}...")
                pythoncom.CoInitialize()
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                wb_com = excel.Workbooks.Open(
                    str(self.current_xlsm.absolute()), UpdateLinks=0, ReadOnly=True
                )
                macro = f"{self.current_xlsm.name}!Module1.genNvmCMDline"
                excel.Application.Run(macro)
                wb_com.Close(SaveChanges=False)
                excel.Quit()
                del excel, wb_com
                pythoncom.CoUninitialize()
                time.sleep(0.5)

                # Step 2b: apply user bit-delta patches to the macro's output files
                sku_suffix = "Cons" if variant == "V" else "Corp"
                for subfolder in self.current_folder.iterdir():
                    if subfolder.is_dir() and sku_suffix in subfolder.name:
                        self._apply_dirty_bits_to_output(subfolder, variant)

            # Step 3: update checksums in each built output folder (skip Official_images)
            self._set_status("Updating checksums...")
            for subfolder in self.current_folder.iterdir():
                if subfolder.is_dir() and subfolder.name != "Official_images":
                    self._apply_checksum_to_folder(subfolder)

            self._set_status(f"Build complete — {nvm_name}_{version}")
            self.dirty_bits_v.clear()
            self.dirty_bits_lm.clear()
            messagebox.showinfo("Build NVM", f"Build complete!\n\nOutput in:\n  {self.current_folder}")
            self.scan_bins(silent=True)

        except Exception as exc:
            self._set_status(f"Build error: {exc}")
            messagebox.showerror("Build NVM", f"Error during build:\n{exc}")
            try:
                excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


    def _generate_full_variant(self, variant: str) -> list[int]:
        """
        Write parameters to Excel, force COM recalculation, then read all
        sheets to produce a complete 4096-word NVM image.
        variant 'V'  -> column G (7) = Consumer (Cons_Prod_NA)
        variant 'LM' -> column H (8) = Corporate (Corp_Prod_NA)
        """
        # Step 1: write params and recalculate formulas via Excel COM
        self._prepare_excel_for_build(variant)

        # Step 2: read the now-correct cached values
        col = 7 if variant == "V" else 8
        lan_sw   = self.lan_sw_var.get()
        lcd_name = "LCD extention LAN SW" if lan_sw == "lan" else "LCD extention non LAN SW"

        wb     = load_workbook(str(self.current_xlsm), data_only=True)
        ws_map = wb["full nvm map"]
        words: list[int] = []

        def _places(br) -> int:
            if br and ":" in str(br):
                try:
                    p = str(br).split(":")
                    return int(p[0].strip()) - int(p[1].strip()) + 1
                except (ValueError, IndexError):
                    pass
            return 16

        def _hexval(raw):
            if raw is None:
                return None
            s = str(raw).replace("0x", "").replace("0X", "").strip()
            try:
                return int(s, 16)
            except ValueError:
                return None

        def _flush(buf: str) -> int:
            if len(buf) > 16:
                buf = buf[-16:]
            return int(buf, 2) if buf else 0

        def _process_sheet(ws, col_idx: int) -> None:
            """Append words from a sheet to the words list using bit-field assembly."""
            buf = ""
            row = 6 if ws == ws_map else 2
            while ws.cell(row, 1).value is not None:
                rdata  = ws.cell(row, col_idx).value
                brange = ws.cell(row, 2).value
                if rdata is not None and brange is not None:
                    v = _hexval(rdata)
                    if v is not None:
                        buf += format(v, f"0{_places(brange)}b")
                if ws.cell(row, 1).value != ws.cell(row + 1, 1).value:
                    words.append(_flush(buf))
                    buf = ""
                row += 1

        # 1. Full NVM map
        _process_sheet(ws_map, col)

        # 2. LCD Extension (col 3, same for both variants)
        last_lcd_addr: int | None = None
        if lcd_name in wb.sheetnames:
            ws_lcd = wb[lcd_name]
            buf = ""
            row = 2
            while ws_lcd.cell(row, 1).value is not None:
                waddr  = ws_lcd.cell(row, 1).value
                rdata  = ws_lcd.cell(row, 3).value
                brange = ws_lcd.cell(row, 2).value
                if waddr is not None and rdata is not None and brange is not None:
                    a = str(waddr).replace("h", "").replace("0x", "").strip()
                    try:
                        last_lcd_addr = int(a, 16)
                    except ValueError:
                        row += 1
                        continue
                    v = _hexval(rdata)
                    if v is not None:
                        buf += format(v, f"0{_places(brange)}b")
                if ws_lcd.cell(row, 1).value != ws_lcd.cell(row + 1, 1).value:
                    words.append(_flush(buf))
                    buf = ""
                row += 1

        # 3. FFFF fill from LCD end to ISCSI start (read from G308)
        iscsi_start = 0x200
        raw_g308 = ws_map.cell(308, 7).value
        if raw_g308 is not None:
            v = _hexval(raw_g308)
            if v is not None:
                iscsi_start = v
        while len(words) < iscsi_start:
            words.append(0xFFFF)

        # 4. ISCSI Module (col 3, same for both variants)
        last_iscsi_addr: int | None = None
        if "ISCSI_MODULE" in wb.sheetnames:
            ws_iscsi = wb["ISCSI_MODULE"]
            buf = ""
            row = 2
            while ws_iscsi.cell(row, 1).value is not None:
                waddr  = ws_iscsi.cell(row, 1).value
                rdata  = ws_iscsi.cell(row, 3).value
                brange = ws_iscsi.cell(row, 2).value
                if waddr is not None and rdata is not None and brange is not None:
                    a = str(waddr).replace("h", "").replace("0x", "").strip()
                    try:
                        last_iscsi_addr = int(a, 16)
                    except ValueError:
                        row += 1
                        continue
                    v = _hexval(rdata)
                    if v is not None:
                        buf += format(v, f"0{_places(brange)}b")
                if ws_iscsi.cell(row, 1).value != ws_iscsi.cell(row + 1, 1).value:
                    words.append(_flush(buf))
                    buf = ""
                row += 1

        wb.close()

        # 5. FFFF fill from ISCSI end to 4088 (matches VBA)
        while len(words) < 4088:
            words.append(0xFFFF)

        # 6. Pad to exactly 4096
        while len(words) < 4096:
            words.append(0xFFFF)

        return words[:4096]

    def _prepare_excel_for_build(self, variant: str) -> None:
        """
        Use Excel COM to write parameters and recalculate all formulas.
        This ensures openpyxl data_only reads see the correct computed values.
        Falls back silently if COM is unavailable.
        """
        try:
            import win32com.client
            import pythoncom
            import time
            pythoncom.CoInitialize()

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(str(self.current_xlsm.absolute()),
                                      UpdateLinks=0, ReadOnly=False)
            ws_var = wb.Worksheets("general variable")
            ws_map = wb.Worksheets("full nvm map")

            # Write NVM parameters
            ws_var.Range("B2").Value = self.project_var.get().strip()
            ws_var.Range("B3").Value = self.silicon_var.get().strip()
            ws_var.Range("B4").Value = self.step_var.get().strip()
            ws_var.Range("B5").Value = self.nvm_output_var.get().strip()

            major = self.major_var.get().strip() or "1"
            minor = (self.minor_var.get().strip() or "0").zfill(2)
            nvm_ver = int(major + minor)
            ws_var.Range("B7").Value = nvm_ver
            ws_var.Range("B8").Value = self.device_id_var.get().strip()
            ws_var.Range("B9").Value = self.sku_device_id_var.get().strip()

            lan_sw = (self.lan_sw_var.get() == "lan")
            ws_var.Range("B12").Value = lan_sw

            # Set LM/V flags
            ws_var.Range("B10").Value = (variant == "LM")   # lm_nvm
            ws_var.Range("B11").Value = (variant == "V")    # v_nvm

            # Calculate LCD extension pointer
            ext_name = "LCD extention LAN SW" if lan_sw else "LCD extention non LAN SW"
            ws_ext = wb.Worksheets(ext_name)
            lcd_rows = ws_ext.UsedRange.Rows.Count
            lcd_ptr = (lcd_rows - 1) // 2
            ws_map.Range("G84").Value = f"0x{lcd_ptr:02X}"
            ws_map.Range("H84").Value = f"0x{lcd_ptr:02X}"

            # Write version in hex format used by Excel
            ver_high = nvm_ver // 100
            ver_low  = nvm_ver % 10
            ver_hex  = f"0x{ver_high:X}0{ver_low:X}4"
            ws_map.Range("G11").Value = ver_hex
            ws_map.Range("H11").Value = ver_hex

            excel.CalculateFull()
            time.sleep(2)

            wb.Save()
            wb.Close(SaveChanges=True)
            excel.Quit()
            del wb, excel
            pythoncom.CoUninitialize()
            self._set_status(f"Excel recalculated for {variant} variant.")

        except Exception as exc:
            self._set_status(f"COM recalc skipped ({exc}); using cached values.")
            try:
                excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _generate_txt_file(self, txt_path: Path) -> None:
        """Generate .txt file from words in memory."""
        lines = []
        words_per_line = 8
        
        for i in range(0, len(self.words), words_per_line):
            line_words = []
            for j in range(words_per_line):
                if i + j < len(self.words):
                    line_words.append(f"{self.words[i + j]:04X}")
                else:
                    line_words.append("FFFF")
            lines.append(" ".join(line_words))
        
        txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    
    def _apply_checksum_to_folder(self, folder: Path) -> None:
        """Read .bin in folder, compute checksum, write back to .bin and .txt."""
        bin_files = list(folder.glob("*.bin"))
        txt_files = list(folder.glob("*.txt"))
        if not bin_files or not txt_files:
            return
        bin_path = bin_files[0]
        txt_path = txt_files[0]

        raw = bin_path.read_bytes()
        words = [int.from_bytes(raw[i:i+2], "little") for i in range(0, len(raw), 2)]

        sum_val = sum(words[:CHECKSUM_WORD])
        checksum = (-sum_val + (2 ** 32) + CHECKSUM_VALUE) & 0xFFFF

        # Patch bin
        with bin_path.open("r+b") as f:
            f.seek(CHECKSUM_WORD * 2)
            f.write(checksum.to_bytes(2, "little"))

        # Patch txt (word at CHECKSUM_LINE * 8 words/line, last word on that line)
        lines = txt_path.read_text(encoding="utf-8").splitlines()
        if len(lines) > CHECKSUM_LINE:
            parts = lines[CHECKSUM_LINE].split()
            if len(parts) >= 8:
                parts[-1] = f"{checksum:04X}"
                lines[CHECKSUM_LINE] = " ".join(parts)
                txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def _update_checksum_in_files(self, txt_path: Path, bin_path: Path) -> None:
        """Update checksums in both txt and bin files."""
        # Calculate checksum from first 0x3F words
        sum_val = sum(self.words[0:CHECKSUM_WORD])
        checksum = (-sum_val + (2 ** 32) + CHECKSUM_VALUE) & 0xFFFF
        
        # Update in memory
        if len(self.words) > CHECKSUM_WORD:
            self.words[CHECKSUM_WORD] = checksum
        
        # Update bin file
        self.write_bin(bin_path)
        
        # Update txt file
        lines = txt_path.read_text(encoding="utf-8").splitlines()
        if len(lines) > CHECKSUM_LINE:
            parts = lines[CHECKSUM_LINE].split()
            if len(parts) >= 8:
                parts[-1] = f"{checksum:04X}"
                lines[CHECKSUM_LINE] = " ".join(parts)
                txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def _run_command(self, args: list[str], cwd: Path) -> None:
        result = subprocess.run(
            args,
            cwd=str(cwd),
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            stdout = result.stdout.strip()
            stderr = result.stderr.strip()
            details = "\n".join([part for part in [stdout, stderr] if part])
            raise RuntimeError(f"Command failed: {' '.join(args)}\n{details}")

    def load_registers(self, auto: bool = False) -> None:
        if auto and not self.current_folder:
            return
        reg_path = None
        if self.current_folder:
            candidate = self.current_folder / "registers.json"
            if candidate.exists():
                reg_path = candidate
        if not auto:
            file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
            if file_path:
                reg_path = Path(file_path)
        if not reg_path:
            self.registers = []
            self.refresh_registers()
            return
        data = json.loads(reg_path.read_text(encoding="utf-8"))
        loaded = []
        for item in data:
            item.setdefault('rtl_name', '')
            try:
                loaded.append(RegisterDef(**item))
            except TypeError:
                loaded.append(RegisterDef(name=item.get('name',''), word_index=item.get('word_index',0),
                                          description=item.get('description',''), rtl_name=item.get('rtl_name','')))
        self.registers = loaded
        self.refresh_registers()

    def save_registers(self) -> None:
        if self.current_folder:
            reg_path = self.current_folder / "registers.json"
        else:
            file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
            if not file_path:
                return
            reg_path = Path(file_path)
        data = [reg.__dict__ for reg in self.registers]
        reg_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
        messagebox.showinfo("Registers", f"Saved register map to {reg_path}")

    def refresh_registers(self) -> None:
        self.reg_tree.delete(*self.reg_tree.get_children())
        flt = self.reg_search_var.get().lower() if hasattr(self, "reg_search_var") else ""
        for reg in self.registers:
            rtl = getattr(reg, 'rtl_name', '') or ''
            display_name = rtl if rtl else reg.name
            if flt:
                if flt not in reg.name.lower() and flt not in rtl.lower():
                    continue
            self.reg_tree.insert("", tk.END, values=(display_name, reg.word_index))

    def _filter_registers(self) -> None:
        self.refresh_registers()

    def _clear_reg_search(self) -> None:
        self.reg_search_var.set("")
        self.refresh_registers()

    def _populate_registers_from_excel(self, silent: bool = False) -> None:
        """Populate register list from loaded Excel data (C-Spec names + word offsets)."""
        if not self.excel_data:
            if not silent:
                messagebox.showinfo("Registers", "Load an Excel file first.")
            return
        added = 0
        existing_names = {r.name for r in self.registers}
        for row in self.excel_data:
            name = row.get("name") or row.get("rtl_name") or row.get("description") or ""
            name = str(name).strip()
            if not name or name in ("N/A", "None", ""):
                continue
            word_idx = row.get("word_idx") or row.get("offset") or 0
            try:
                if isinstance(word_idx, str):
                    word_idx = int(word_idx, 16) if word_idx.startswith("0x") else int(word_idx, 0)
                word_idx = int(word_idx)
            except (ValueError, TypeError):
                word_idx = 0
            if name not in existing_names:
                rtl = str(row.get('rtl_name') or '').strip()
                self.registers.append(RegisterDef(name=name, word_index=word_idx, rtl_name=rtl))
                existing_names.add(name)
                added += 1
        self.refresh_registers()
        if not silent:
            messagebox.showinfo("Registers", f"Added {added} registers from Excel data.")

    def on_register_select(self, _event: tk.Event) -> None:
        sel = self.reg_tree.selection()
        if not sel:
            return
        values = self.reg_tree.item(sel[0], "values")
        if not values:
            return
        display_name = str(values[0])
        self.register_name_var.set(display_name)
        self.index_var.set(str(values[1]))

        # Find register by rtl_name or c-spec name
        def _match(r):
            rtl = getattr(r, 'rtl_name', '') or ''
            return rtl == display_name or r.name == display_name
        reg = next((r for r in self.registers if _match(r)), None)
        if reg and reg.description:
            self.register_desc_var.set(reg.description)
        else:
            self.register_desc_var.set("")
        
        try:
            idx = int(values[1])
            # Ensure words list is large enough
            while len(self.words) <= idx:
                self.words.append(0xFFFF)
            self.value_var.set(f"0x{self.words[idx]:04X}")
        except Exception:
            pass

    def add_register(self) -> None:
        self._open_register_editor()

    def edit_register(self) -> None:
        sel = self.reg_tree.selection()
        if not sel:
            messagebox.showwarning("Registers", "Select a register to edit.")
            return
        values = self.reg_tree.item(sel[0], "values")
        name = values[0]
        reg = next((r for r in self.registers if r.name == name), None)
        if reg:
            self._open_register_editor(reg)

    def remove_register(self) -> None:
        sel = self.reg_tree.selection()
        if not sel:
            return
        values = self.reg_tree.item(sel[0], "values")
        name = values[0]
        self.registers = [r for r in self.registers if r.name != name]
        self.refresh_registers()

    def _open_register_editor(self, reg: RegisterDef | None = None) -> None:
        win = tk.Toplevel(self)
        win.title("Register")
        win.geometry("360x200")

        name_var = tk.StringVar(value=reg.name if reg else "")
        index_var = tk.StringVar(value=str(reg.word_index) if reg else "")
        desc_var = tk.StringVar(value=reg.description if reg else "")

        ttk.Label(win, text="Name:").pack(anchor=tk.W, padx=12, pady=(12, 2))
        ttk.Entry(win, textvariable=name_var, width=40).pack(anchor=tk.W, padx=12)

        ttk.Label(win, text="Word Index:").pack(anchor=tk.W, padx=12, pady=(8, 2))
        ttk.Entry(win, textvariable=index_var, width=20).pack(anchor=tk.W, padx=12)

        ttk.Label(win, text="Description:").pack(anchor=tk.W, padx=12, pady=(8, 2))
        ttk.Entry(win, textvariable=desc_var, width=40).pack(anchor=tk.W, padx=12)

        def save() -> None:
            try:
                idx = int(index_var.get().strip())
            except Exception:
                messagebox.showerror("Invalid", "Word index must be an integer.")
                return
            new_reg = RegisterDef(name=name_var.get().strip(), word_index=idx, description=desc_var.get().strip())
            if not new_reg.name:
                messagebox.showerror("Invalid", "Name is required.")
                return
            self.registers = [r for r in self.registers if r.name != new_reg.name]
            self.registers.append(new_reg)
            self.refresh_registers()
            win.destroy()

        ttk.Button(win, text="Save", command=save).pack(pady=12)

    def on_excel_tree_double_click(self, event) -> None:
        """Handle double-click on Excel data tree to edit cell."""
        region = self.excel_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        # Get clicked row and column
        item = self.excel_tree.identify_row(event.y)
        column = self.excel_tree.identify_column(event.x)
        
        if not item:
            return
        
        # Get row index
        row_idx = self.excel_tree.index(item)
        if row_idx >= len(self.excel_data):
            return
        
        # Get column index (column is like '#1', '#2', etc.)
        col_idx = int(column.replace('#', '')) - 1
        
        # Get current column configuration
        columns_to_use = self.excel_columns_simple if self.show_less_var.get() else self.excel_columns_full
        if col_idx >= len(columns_to_use):
            return
        
        col_key, col_name, col_width = columns_to_use[col_idx]
        current_value = self.excel_data[row_idx].get(col_key, '')
        
        # Create popup dialog for editing
        edit_win = tk.Toplevel(self)
        edit_win.title(f"Edit {col_name}")
        edit_win.geometry("400x150")
        edit_win.transient(self)
        edit_win.grab_set()
        
        ttk.Label(edit_win, text=f"Edit {col_name}:", font=('Segoe UI', 10, 'bold')).pack(pady=10, padx=10)
        
        edit_var = tk.StringVar(value=str(current_value))
        entry = ttk.Entry(edit_win, textvariable=edit_var, width=50)
        entry.pack(pady=10, padx=10)
        entry.focus()
        entry.select_range(0, tk.END)
        
        def save_edit():
            new_value = edit_var.get()
            self.excel_data[row_idx][col_key] = new_value
            
            # If editing V or LM columns, update the word arrays
            if col_key in ['v', 'lm']:
                try:
                    row_data = self.excel_data[row_idx]
                    word_idx = row_data.get('word_idx')
                    
                    if word_idx is not None:
                        # Parse bit range to determine which bits to update
                        bits_str = str(row_data.get('bits', ''))
                        
                        # Check if this is a full word update (15:0) or partial
                        if ':' in bits_str:
                            parts = bits_str.split(':')
                            high_bit = int(parts[0].strip())
                            low_bit = int(parts[1].strip())
                            
                            # Calculate mask
                            num_bits = high_bit - low_bit + 1
                            mask = (1 << num_bits) - 1
                            
                            # Parse new value
                            new_val_str = new_value.replace('0x', '').replace('0X', '').strip()
                            if new_val_str and new_val_str.upper() != 'N/A':
                                new_val_int = int(new_val_str, 16)
                                
                                # Update the appropriate word array
                                if col_key == 'v':
                                    # Ensure array is large enough
                                    while len(self.words_v) <= word_idx:
                                        self.words_v.append(0xFFFF)
                                    
                                    # Update V word
                                    current_word = self.words_v[word_idx]
                                    current_word &= ~(mask << low_bit)  # Clear bits
                                    current_word |= (new_val_int & mask) << low_bit  # Set new bits
                                    self.words_v[word_idx] = current_word
                                    fms = mask << low_bit
                                    s, c = self.dirty_bits_v.get(word_idx, (0, 0))
                                    c |= fms; s = (s & ~fms) | ((new_val_int & mask) << low_bit)
                                    self.dirty_bits_v[word_idx] = (s, c)
                                    
                                elif col_key == 'lm':
                                    # Ensure array is large enough
                                    while len(self.words_lm) <= word_idx:
                                        self.words_lm.append(0xFFFF)
                                    
                                    # Update LM word
                                    current_word = self.words_lm[word_idx]
                                    current_word &= ~(mask << low_bit)  # Clear bits
                                    current_word |= (new_val_int & mask) << low_bit  # Set new bits
                                    self.words_lm[word_idx] = current_word
                                    fms = mask << low_bit
                                    s, c = self.dirty_bits_lm.get(word_idx, (0, 0))
                                    c |= fms; s = (s & ~fms) | ((new_val_int & mask) << low_bit)
                                    self.dirty_bits_lm[word_idx] = (s, c)
                                
                                # Update the default words array based on excel_col_var
                                if self.excel_col_var.get() == "V":
                                    self.words = self.words_v.copy()
                                else:
                                    self.words = self.words_lm.copy()
                                
                                # Refresh word list to show changes
                                self.refresh_word_list()
                except Exception as e:
                    print(f"Warning: Could not update word array: {e}")
            
            self.refresh_excel_tree()
            
            # Show confirmation message
            if col_key in ['v', 'lm']:
                messagebox.showinfo("Value Updated", f"{col_name} updated successfully!\nThe change will be included in the next build.")
            
            edit_win.destroy()
        
        def cancel_edit():
            edit_win.destroy()
        
        btn_frame = ttk.Frame(edit_win)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Save", command=save_edit).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=cancel_edit).pack(side=tk.LEFT, padx=5)
        
        entry.bind('<Return>', lambda e: save_edit())
        entry.bind('<Escape>', lambda e: cancel_edit())
    
    def on_tree_double_click(self, event, tree, data_list) -> None:
        """Handle double-click on any tree to edit cell."""
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        # Get clicked row and column
        item = tree.identify_row(event.y)
        column = tree.identify_column(event.x)
        
        if not item:
            return
        
        # Get row index and column index
        row_idx = tree.index(item)
        if row_idx >= len(data_list):
            return
        
        col_idx = int(column.replace('#', '')) - 1
        if col_idx >= len(data_list[row_idx]):
            return
        
        current_value = data_list[row_idx][col_idx]
        
        # Create popup dialog for editing
        edit_win = tk.Toplevel(self)
        edit_win.title(f"Edit Value")
        edit_win.geometry("400x150")
        edit_win.transient(self)
        edit_win.grab_set()
        
        ttk.Label(edit_win, text=f"Edit value:", font=('Segoe UI', 10, 'bold')).pack(pady=10, padx=10)
        
        edit_var = tk.StringVar(value=str(current_value))
        entry = ttk.Entry(edit_win, textvariable=edit_var, width=50)
        entry.pack(pady=10, padx=10)
        entry.focus()
        entry.select_range(0, tk.END)
        
        def save_edit():
            new_value = edit_var.get()
            data_list[row_idx][col_idx] = new_value
            self.refresh_additional_trees()
            edit_win.destroy()
        
        def cancel_edit():
            edit_win.destroy()
        
        btn_frame = ttk.Frame(edit_win)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Save", command=save_edit).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=cancel_edit).pack(side=tk.LEFT, padx=5)
        
        entry.bind('<Return>', lambda e: save_edit())
        entry.bind('<Escape>', lambda e: cancel_edit())
    
    def show_excel_context_menu(self, event) -> None:
        """Show context menu for Excel Data tab."""
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Add Column to Full View", command=self.add_excel_column_dialog)
        menu.add_command(label="Toggle View (Simple/Full)", command=lambda: self.show_less_var.set(not self.show_less_var.get()))
        menu.tk_popup(event.x_root, event.y_root)
    
    def add_excel_column_dialog(self) -> None:
        """Show dialog to add a new column to Excel Data."""
        dialog = tk.Toplevel(self)
        dialog.title("Add Column to Excel Data")
        dialog.geometry("400x200")
        dialog.transient(self)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Column Key (internal name):", font=('Segoe UI', 10)).pack(pady=5, padx=10, anchor=tk.W)
        key_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=key_var, width=50).pack(pady=2, padx=10)
        
        ttk.Label(dialog, text="Column Display Name:", font=('Segoe UI', 10)).pack(pady=5, padx=10, anchor=tk.W)
        name_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=name_var, width=50).pack(pady=2, padx=10)
        
        ttk.Label(dialog, text="Column Width:", font=('Segoe UI', 10)).pack(pady=5, padx=10, anchor=tk.W)
        width_var = tk.StringVar(value="150")
        ttk.Entry(dialog, textvariable=width_var, width=50).pack(pady=2, padx=10)
        
        def save_column():
            key = key_var.get().strip()
            name = name_var.get().strip()
            try:
                width = int(width_var.get().strip())
            except:
                width = 150
            
            if not key or not name:
                messagebox.showwarning("Add Column", "Please enter both key and display name.")
                return
            
            # Add to full columns list
            self.excel_columns_full.append((key, name, width))
            
            # Add empty value to all existing data rows
            for row in self.excel_data:
                if key not in row:
                    row[key] = ''
            
            # Recreate tree if in full view
            if not self.show_less_var.get():
                self._create_excel_tree(self.excel_tree_parent)
                self.refresh_excel_tree()
            
            dialog.destroy()
            messagebox.showinfo("Add Column", f"Column '{name}' added to full view!")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Add", command=save_column).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def show_tab_context_menu(self, event, tab_name) -> None:
        """Show context menu for adding columns/rows."""
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Add New Column", command=lambda: self.add_column_dialog(tab_name))
        menu.add_command(label="Add New Row", command=lambda: self.add_row_to_tab(tab_name))
        menu.tk_popup(event.x_root, event.y_root)
    
    def add_column_dialog(self, tab_name) -> None:
        """Show dialog to add a new column to a tab."""
        dialog = tk.Toplevel(self)
        dialog.title("Add New Column")
        dialog.geometry("350x120")
        dialog.transient(self)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Column Name:", font=('Segoe UI', 10)).pack(pady=10, padx=10)
        
        col_name_var = tk.StringVar()
        entry = ttk.Entry(dialog, textvariable=col_name_var, width=40)
        entry.pack(pady=5, padx=10)
        entry.focus()
        
        def save_column():
            col_name = col_name_var.get().strip()
            if not col_name:
                messagebox.showwarning("Add Column", "Please enter a column name.")
                return
            
            # Add column to the appropriate data structure
            if tab_name == "lcd_non_lan":
                for row in self.lcd_non_lan_data:
                    row.append('')  # Add empty value for new column
            elif tab_name == "lcd_lan":
                for row in self.lcd_lan_data:
                    row.append('')
            elif tab_name == "iscsi":
                for row in self.iscsi_data:
                    row.append('')
            elif tab_name == "basic_nvm":
                for row in self.basic_nvm_data:
                    row.append('')
            
            self.refresh_additional_trees()
            dialog.destroy()
            messagebox.showinfo("Add Column", f"Column '{col_name}' added successfully!")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Add", command=save_column).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        entry.bind('<Return>', lambda e: save_column())
        entry.bind('<Escape>', lambda e: dialog.destroy())
    
    def add_row_to_tab(self, tab_name) -> None:
        """Add a new empty row to a tab."""
        if tab_name == "lcd_non_lan":
            num_cols = len(self.lcd_non_lan_data[0]) if self.lcd_non_lan_data else 5
            self.lcd_non_lan_data.append([''] * num_cols)
        elif tab_name == "lcd_lan":
            num_cols = len(self.lcd_lan_data[0]) if self.lcd_lan_data else 5
            self.lcd_lan_data.append([''] * num_cols)
        elif tab_name == "iscsi":
            num_cols = len(self.iscsi_data[0]) if self.iscsi_data else 5
            self.iscsi_data.append([''] * num_cols)
        elif tab_name == "basic_nvm":
            num_cols = len(self.basic_nvm_data[0]) if self.basic_nvm_data else 5
            self.basic_nvm_data.append([''] * num_cols)
        
        self.refresh_additional_trees()
        messagebox.showinfo("Add Row", "New row added successfully!")

    def generate_nvm(self) -> None:
        """VBA Legacy Build: write params via COM, run Excel macro genNvmCMDline, then calc_csum."""
        build_root = self.current_folder
        if not build_root:
            folder = filedialog.askdirectory(title="Select build folder with Excel workbook")
            if not folder:
                return
            build_root = Path(folder)

        xlsm_files = list(build_root.glob("*.xlsm"))
        if not xlsm_files:
            messagebox.showerror("VBA Build", "No .xlsm file found in the selected folder.")
            return
        xlsm_path = xlsm_files[0]

        py_path = build_root / "calc_csum.py"
        if not py_path.exists():
            messagebox.showerror("VBA Build", "calc_csum.py not found in the project folder.")
            return

        try:
            import win32com.client
            import pythoncom
            import time

            self._set_status("VBA Build: writing parameters + word values to Excel...")
            nvm_name_vba = self.nvm_output_var.get().strip() or self.project_var.get().strip()
            major = self.major_var.get().strip() or "1"
            minor = (self.minor_var.get().strip() or "0")
            variant_vba = self.lm_v_var.get()
            # Write for each selected variant
            variants_vba = []
            if variant_vba in ("V", "Both"):   variants_vba.append("V")
            if variant_vba in ("LM", "Both"):  variants_vba.append("LM")
            if not variants_vba:               variants_vba = ["V"]
            for v in variants_vba:
                self._write_words_to_excel(v, nvm_name_vba, major, minor)

            self._set_status("VBA Build: running macro genNvmCMDline...")
            pythoncom.CoInitialize()
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(xlsm_path.absolute()), UpdateLinks=0, ReadOnly=True)
            macro_name = f"{xlsm_path.name}!Module1.genNvmCMDline"
            excel.Application.Run(macro_name)
            time.sleep(1)

            wb.Close(SaveChanges=False)
            excel.Quit()
            del wb, excel
            pythoncom.CoUninitialize()

            # Apply user bit-delta patches to output files
            for v in variants_vba:
                sku_suffix = "Cons" if v == "V" else "Corp"
                for subfolder in build_root.iterdir():
                    if subfolder.is_dir() and sku_suffix in subfolder.name:
                        self._apply_dirty_bits_to_output(subfolder, v)

            # Update checksums in each output subfolder (skip Official_images)
            self._set_status("VBA Build: updating checksums...")
            for subfolder in build_root.iterdir():
                if subfolder.is_dir() and subfolder.name != "Official_images":
                    self._apply_checksum_to_folder(subfolder)

            self.dirty_bits_v.clear()
            self.dirty_bits_lm.clear()
            self._set_status("VBA Build complete.")
            messagebox.showinfo("VBA Build", "Excel macro ran successfully.\nChecksum updated.\nOutput files are in the project folder.")

        except Exception as exc:
            self._set_status(f"VBA Build error: {exc}")
            messagebox.showerror("VBA Build", f"Error:\n{exc}\n\nMake sure Microsoft Excel is installed.")
            try:
                excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _generate_json_params(self, build_root: Path) -> None:
        """Generate JSON parameter files in output folders."""
        params = {
            "projectName": self.project_var.get().strip(),
            "silicon": self.silicon_var.get().strip(),
            "step": self.step_var.get().strip(),
            "version": {
                "major": self.major_var.get().strip(),
                "minor": self.minor_var.get().strip()
            },
            "deviceId": self.device_id_var.get().strip(),
            "skuDeviceId": self.sku_device_id_var.get().strip(),
            "nvmOutput": self.nvm_output_var.get().strip() or self.project_var.get().strip(),
            "lanMode": self.lan_sw_var.get(),
            "lmvMode": self.lm_v_var.get(),
            "hideValuesRead": self.hide_values_var.get(),
            "createWithReadValues": self.create_with_read_var.get()
        }
        
        # Find output folders (they follow the pattern: projectname_version_SKU)
        for subfolder in build_root.iterdir():
            if subfolder.is_dir() and not subfolder.name.startswith("."):
                # Check if it contains bin/txt files
                bin_files = list(subfolder.glob("*.bin"))
                if bin_files:
                    # Generate JSON file
                    json_path = subfolder / "parameters.json"
                    json_path.write_text(json.dumps(params, indent=2), encoding="utf-8")

    def read_existing_nvm(self) -> None:
        if load_workbook is None:
            messagebox.showerror("Read NVM", "openpyxl is not installed. Install it to use this feature.")
            return

        build_root = self.current_folder
        if not build_root:
            folder = filedialog.askdirectory(title="Select build folder with Excel workbook")
            if not folder:
                return
            build_root = Path(folder)

        # Find the .xlsm file
        xlsm_files = list(build_root.glob("*.xlsm"))
        if not xlsm_files:
            messagebox.showerror("Read NVM", "No .xlsm file found in the selected folder.")
            return
        xlsm_path = xlsm_files[0]

        # Select bin file to read
        bin_file = filedialog.askopenfilename(
            title="Select NVM bin file to read",
            filetypes=[("Binary files", "*.bin"), ("All files", "*.*")],
        )
        if not bin_file:
            return

        try:
            # Read the bin file and populate column I in "full nvm map" sheet
            wb = load_workbook(str(xlsm_path), keep_vba=True)
            if "full nvm map" not in wb.sheetnames:
                messagebox.showerror("Read NVM", "'full nvm map' sheet not found in workbook.")
                wb.close()
                return

            ws = wb["full nvm map"]
            bin_data = Path(bin_file).read_bytes()
            words = [int.from_bytes(bin_data[i:i + 2], "little") for i in range(0, len(bin_data), 2)]

            # Populate column I with hex values starting from row 6
            row = 6
            word_idx = 0
            while ws.cell(row, 1).value and word_idx < len(words):
                word_num = ws.cell(row, 1).value
                # Check if this is the last range for this word
                next_word_num = ws.cell(row + 1, 1).value if ws.cell(row + 1, 1).value else None
                if word_num != next_word_num:
                    # Write the word value to column I
                    bit_range = ws.cell(row, 2).value  # e.g., "15:0"
                    if bit_range and ":" in str(bit_range):
                        parts = str(bit_range).split(":")
                        msb = int(parts[0])
                        lsb = int(parts[1])
                        mask = (1 << (msb - lsb + 1)) - 1
                        value = (words[word_idx] >> lsb) & mask
                        ws.cell(row, 9).value = f"0x{value:X}"
                    word_idx += 1
                row += 1

            # Show column I
            ws.column_dimensions["I"].hidden = False

            wb.save(str(xlsm_path))
            wb.close()

            messagebox.showinfo("Read NVM", f"Successfully read {bin_file} into Excel workbook.")

        except Exception as exc:
            messagebox.showerror("Read NVM", f"Error: {exc}")


    def show_version_diff(self) -> None:
        """Open a window listing all built output versions for the current project
        and showing word-level diffs between any two selected versions."""
        if not self.current_folder or not self.current_folder.exists():
            messagebox.showwarning("Version Diff", "No project folder loaded.\nSelect a project first.")
            return

        # ── Collect all subfolders that contain a .bin file ──────────────────
        import struct as _struct
        import datetime as _dt

        entries = []  # list of (display_name, bin_path)
        for sub in sorted(self.current_folder.iterdir()):
            if not sub.is_dir():
                continue
            bins = list(sub.glob("*.bin"))
            if not bins:
                continue
            bin_path = bins[0]
            entries.append((sub.name, bin_path))

        if not entries:
            messagebox.showinfo("Version Diff", "No built output versions found in this project.\nRun a build first.")
            return

        # ── Build window ─────────────────────────────────────────────────────
        win = tk.Toplevel(self)
        win.title(f"Version Diff  —  {self.current_folder.name}")
        win.geometry("1200x700")
        win.transient(self)

        # Header
        hdr = tk.Frame(win, bg=self.colors['header_bg'], height=44)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"Version Diff  ·  {self.current_folder.name}",
                 font=('Segoe UI', 11, 'bold'), fg='white',
                 bg=self.colors['header_bg']).pack(side=tk.LEFT, padx=16, pady=8)
        tk.Label(hdr, text="Select two rows, then click Compare",
                 font=('Segoe UI', 9), fg='#a8d4f0',
                 bg=self.colors['header_bg']).pack(side=tk.RIGHT, padx=16, pady=8)

        # ── Top pane: version list ────────────────────────────────────────────
        top_pane = ttk.Frame(win, padding=(8, 8, 8, 4))
        top_pane.pack(fill=tk.X)

        cols_v = ("name", "variant", "words", "size_kb", "modified")
        ver_tree = ttk.Treeview(top_pane, columns=cols_v, show="headings", height=7, selectmode="extended")
        for col, hd, w in [
            ("name",     "Version / Folder",  380),
            ("variant",  "Variant",            90),
            ("words",    "Words",              70),
            ("size_kb",  "Size (KB)",          80),
            ("modified", "Modified",           160),
        ]:
            ver_tree.heading(col, text=hd)
            ver_tree.column(col, width=w, minwidth=50, stretch=(col == "name"))

        ver_tree.tag_configure('official', foreground='#0071c5', font=('Segoe UI', 9, 'bold'))
        ver_tree.tag_configure('even', background='#f9f9f9')
        ver_tree.tag_configure('odd',  background='#ffffff')

        vsb = ttk.Scrollbar(top_pane, orient=tk.VERTICAL, command=ver_tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        ver_tree.configure(yscrollcommand=vsb.set)
        ver_tree.pack(fill=tk.X)

        # Populate version list
        bin_data: dict[str, bytes] = {}  # name -> raw bytes
        for i, (name, bin_path) in enumerate(entries):
            raw = bin_path.read_bytes()
            bin_data[name] = raw
            variant = "Corp" if "corp" in name.lower() else ("Cons" if "cons" in name.lower() else "—")
            words   = len(raw) // 2
            size_kb = round(len(raw) / 1024, 1)
            mtime   = _dt.datetime.fromtimestamp(bin_path.stat().st_mtime).strftime("%Y-%m-%d  %H:%M")
            tag     = ('official', 'even' if i % 2 == 0 else 'odd') if 'official' in name.lower() else ('even' if i % 2 == 0 else 'odd',)
            ver_tree.insert("", tk.END, iid=name, values=(name, variant, words, size_kb, mtime), tags=tag)

        # ── Controls ─────────────────────────────────────────────────────────
        ctrl = ttk.Frame(win, padding=(8, 2, 8, 4))
        ctrl.pack(fill=tk.X)

        lbl_a = tk.StringVar(value="A: (none)")
        lbl_b = tk.StringVar(value="B: (none)")
        sel_a: list = [None]
        sel_b: list = [None]

        def _pick(slot_var, slot_list, lbl_var):
            sel = ver_tree.selection()
            if not sel:
                messagebox.showwarning("Version Diff", "Select a row first.", parent=win)
                return
            slot_list[0] = sel[0]
            lbl_var.set(f"{['A','B'][slot_list is sel_b]}: {sel[0]}")

        ttk.Button(ctrl, text="Set as A",  command=lambda: _pick(lbl_a, sel_a, lbl_a), width=12).pack(side=tk.LEFT, padx=4)
        tk.Label(ctrl, textvariable=lbl_a, font=('Segoe UI', 9, 'bold'), fg='#0071c5', width=45, anchor=tk.W).pack(side=tk.LEFT)
        ttk.Button(ctrl, text="Set as B",  command=lambda: _pick(lbl_b, sel_b, lbl_b), width=12).pack(side=tk.LEFT, padx=(20, 4))
        tk.Label(ctrl, textvariable=lbl_b, font=('Segoe UI', 9, 'bold'), fg='#107c10', width=45, anchor=tk.W).pack(side=tk.LEFT)

        # ── Diff table ───────────────────────────────────────────────────────
        diff_frame = ttk.LabelFrame(win, text="Word-level Differences", padding=(8, 4))
        diff_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=(4, 8))

        summary_var = tk.StringVar(value="Select two versions and click Compare.")
        tk.Label(diff_frame, textvariable=summary_var,
                 font=('Segoe UI', 9, 'bold'), fg='#333', anchor=tk.W).pack(anchor=tk.W, pady=(0, 4))

        cols_d = ("word", "offset", "name", "val_a", "val_b", "changed_bits")
        diff_tree = ttk.Treeview(diff_frame, columns=cols_d, show="headings", height=18)
        for col, hd, w in [
            ("word",         "Word (dec)",     80),
            ("offset",       "Offset (hex)",   90),
            ("name",         "C-Spec Name",   280),
            ("val_a",        "Value  A",        90),
            ("val_b",        "Value  B",        90),
            ("changed_bits", "Changed Bits",   160),
        ]:
            diff_tree.heading(col, text=hd)
            diff_tree.column(col, width=w, minwidth=50, stretch=(col == "name"))

        diff_tree.tag_configure('diff',    background='#fff3cd')
        diff_tree.tag_configure('only_a',  background='#fce8e8')
        diff_tree.tag_configure('only_b',  background='#e8f5e8')

        dsb_y = ttk.Scrollbar(diff_frame, orient=tk.VERTICAL,   command=diff_tree.yview)
        dsb_x = ttk.Scrollbar(diff_frame, orient=tk.HORIZONTAL, command=diff_tree.xview)
        dsb_y.pack(side=tk.RIGHT,  fill=tk.Y)
        dsb_x.pack(side=tk.BOTTOM, fill=tk.X)
        diff_tree.configure(yscrollcommand=dsb_y.set, xscrollcommand=dsb_x.set)
        diff_tree.pack(fill=tk.BOTH, expand=True)

        # Build word-index -> C-Spec name lookup from loaded excel_data
        word_names: dict[int, str] = {}
        for row in (self.excel_data or []):
            idx = row.get('word_idx')
            name_s = row.get('name') or row.get('rtl_name') or ''
            if idx is not None and name_s and name_s not in ('N/A', 'None', ''):
                if idx not in word_names:
                    word_names[idx] = str(name_s)

        def _compare():
            a_name = sel_a[0]
            b_name = sel_b[0]
            if not a_name or not b_name:
                messagebox.showwarning("Version Diff", "Set both A and B versions first.", parent=win)
                return
            if a_name == b_name:
                messagebox.showwarning("Version Diff", "A and B are the same version.", parent=win)
                return

            raw_a = bin_data.get(a_name, b'')
            raw_b = bin_data.get(b_name, b'')
            words_a = [_struct.unpack_from('<H', raw_a, i)[0] for i in range(0, len(raw_a)-1, 2)]
            words_b = [_struct.unpack_from('<H', raw_b, i)[0] for i in range(0, len(raw_b)-1, 2)]
            max_len = max(len(words_a), len(words_b))

            diff_tree.delete(*diff_tree.get_children())
            diff_count = 0

            for idx in range(max_len):
                va = words_a[idx] if idx < len(words_a) else None
                vb = words_b[idx] if idx < len(words_b) else None
                if va == vb:
                    continue
                diff_count += 1
                offset_hex = f"0x{idx:04X}"
                name_str   = word_names.get(idx, "")
                val_a_str  = f"0x{va:04X}" if va is not None else "—"
                val_b_str  = f"0x{vb:04X}" if vb is not None else "—"

                # Which bits changed
                if va is not None and vb is not None:
                    xor = va ^ vb
                    changed = [str(b) for b in range(15, -1, -1) if xor & (1 << b)]
                    bits_str = "bits " + ",".join(changed) if changed else ""
                    tag = 'diff'
                elif va is None:
                    bits_str = "only in B"
                    tag = 'only_b'
                else:
                    bits_str = "only in A"
                    tag = 'only_a'

                diff_tree.insert("", tk.END,
                                  values=(idx, offset_hex, name_str, val_a_str, val_b_str, bits_str),
                                  tags=(tag,))

            summary_var.set(
                f"A: {a_name}    vs    B: {b_name}    →    {diff_count} word(s) differ"
                + ("  ✓  identical" if diff_count == 0 else "")
            )

        ttk.Button(ctrl, text="Compare  ▶", command=_compare, style='Build.TButton').pack(side=tk.RIGHT, padx=8)

        # Auto-compare if only 2 entries
        if len(entries) == 2:
            sel_a[0] = entries[0][0]; lbl_a.set(f"A: {entries[0][0]}")
            sel_b[0] = entries[1][0]; lbl_b.set(f"B: {entries[1][0]}")
            win.after(100, _compare)


if __name__ == "__main__":
    app = GbeImageEditor()
    app.mainloop()
