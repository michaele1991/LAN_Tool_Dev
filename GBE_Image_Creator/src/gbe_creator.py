"""
GBE Image Creator - Python implementation of Excel VBA NVM tool
Replicates the functionality of the Excel macro system for generating GBE NVM images.
"""

import json
import struct
from pathlib import Path
from typing import Optional
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import win32com.client
    EXCEL_COM_AVAILABLE = True
except ImportError:
    EXCEL_COM_AVAILABLE = False


class GBEImageCreator(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Intel® GBE NVM Image Generator - v1.0")
        self.geometry("1000x900")
        
        # Set window icon and styling
        try:
            self.iconbitmap(default='')  # Use default if no icon
        except:
            pass
        
        self.current_folder: Optional[Path] = None
        self.current_xlsm: Optional[Path] = None
        self.gbe_folders: list[Path] = []
        self.register_data: list[dict] = []  # Store register information from Excel
        
        self._build_ui()
        self._load_gbe_folders()
    
    def _build_ui(self):
        # Configure style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Header with title
        header_frame = tk.Frame(self, bg='#0071C5', height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text="Intel® GBE NVM Image Generator", 
                              font=('Segoe UI', 16, 'bold'), fg='white', bg='#0071C5')
        title_label.pack(pady=15)
        
        # Top frame - Project selection (always visible)
        top_frame = ttk.LabelFrame(self, text="Project Selection", padding=10)
        top_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.folder_var = tk.StringVar()
        ttk.Label(top_frame, text="Project:").pack(side=tk.LEFT, padx=(0, 5))
        self.folder_combo = ttk.Combobox(top_frame, textvariable=self.folder_var, width=40, state="readonly")
        self.folder_combo.pack(side=tk.LEFT, padx=(0, 5))
        self.folder_combo.bind("<<ComboboxSelected>>", lambda _: self.on_project_select())
        ttk.Button(top_frame, text="Refresh", command=self._load_gbe_folders).pack(side=tk.LEFT)
        
        # Notebook for tabs
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Tab 1: NVM Parameters
        params_tab = ttk.Frame(self.notebook)
        self.notebook.add(params_tab, text="NVM Parameters")
        self._build_params_tab(params_tab)
        
        # Tab 2: Register Editor
        editor_tab = ttk.Frame(self.notebook)
        self.notebook.add(editor_tab, text="Register Editor")
        self._build_editor_tab(editor_tab)
        
        # Action buttons at bottom
        btn_frame = ttk.Frame(self, padding=10)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Define button style for primary actions
        style = ttk.Style()
        style.configure('Primary.TButton', font=('Segoe UI', 9, 'bold'))
        
        ttk.Button(btn_frame, text="Generate NVM Images", command=self.generate_nvm, 
                  width=25, style='Primary.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Read Existing NVM", command=self.read_nvm, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Load from Excel", command=self.load_from_excel, width=20).pack(side=tk.LEFT, padx=5)
        
        # Log frame at bottom
        log_frame = ttk.LabelFrame(self, text="Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, width=80)
        self.log_text.pack(fill=tk.BOTH, expand=True)
    
    def _build_params_tab(self, parent):
        """Build the NVM Parameters tab."""
        # Create a canvas with scrollbar for the parameters
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Parameters frame inside scrollable frame
        param_frame = ttk.LabelFrame(scrollable_frame, text="NVM Parameters", padding=10)
        param_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Variables
        self.project_var = tk.StringVar()
        self.silicon_var = tk.StringVar()
        self.step_var = tk.StringVar()
        self.major_var = tk.StringVar()
        self.minor_var = tk.StringVar()
        self.device_id_var = tk.StringVar()
        self.sku_device_id_var = tk.StringVar()
        self.nvm_output_var = tk.StringVar()
        self.lan_sw_var = tk.StringVar(value="non_lan")
        self.lm_v_var = tk.StringVar(value="both")
        self.hide_values_var = tk.BooleanVar(value=True)
        self.create_with_read_var = tk.BooleanVar(value=False)
        
        # Grid layout for parameters
        row = 0
        ttk.Label(param_frame, text="Project name:").grid(row=row, column=0, sticky=tk.W, pady=2)
        ttk.Entry(param_frame, textvariable=self.project_var, width=40).grid(row=row, column=1, pady=2, sticky=tk.W)
        
        row += 1
        ttk.Label(param_frame, text="Silicon:").grid(row=row, column=0, sticky=tk.W, pady=2)
        ttk.Entry(param_frame, textvariable=self.silicon_var, width=40).grid(row=row, column=1, pady=2, sticky=tk.W)
        
        row += 1
        ttk.Label(param_frame, text="Step:").grid(row=row, column=0, sticky=tk.W, pady=2)
        ttk.Entry(param_frame, textvariable=self.step_var, width=40).grid(row=row, column=1, pady=2, sticky=tk.W)
        
        row += 1
        ttk.Label(param_frame, text="Insert major version:").grid(row=row, column=0, sticky=tk.W, pady=2)
        ttk.Entry(param_frame, textvariable=self.major_var, width=40).grid(row=row, column=1, pady=2, sticky=tk.W)
        
        row += 1
        ttk.Label(param_frame, text="Insert minor version:").grid(row=row, column=0, sticky=tk.W, pady=2)
        ttk.Entry(param_frame, textvariable=self.minor_var, width=40).grid(row=row, column=1, pady=2, sticky=tk.W)
        
        row += 1
        ttk.Label(param_frame, text="Insert device ID:").grid(row=row, column=0, sticky=tk.W, pady=2)
        ttk.Entry(param_frame, textvariable=self.device_id_var, width=40).grid(row=row, column=1, pady=2, sticky=tk.W)
        
        row += 1
        ttk.Label(param_frame, text="Insert SKU device ID:").grid(row=row, column=0, sticky=tk.W, pady=2)
        ttk.Entry(param_frame, textvariable=self.sku_device_id_var, width=40).grid(row=row, column=1, pady=2, sticky=tk.W)
        
        row += 1
        ttk.Label(param_frame, text="NVM output files:").grid(row=row, column=0, sticky=tk.W, pady=2)
        ttk.Entry(param_frame, textvariable=self.nvm_output_var, width=40).grid(row=row, column=1, pady=2, sticky=tk.W)
        
        row += 1
        ttk.Label(param_frame, text="LAN / non LAN SW:").grid(row=row, column=0, sticky=tk.W, pady=2)
        lan_frame = ttk.Frame(param_frame)
        lan_frame.grid(row=row, column=1, sticky=tk.W, pady=2)
        ttk.Radiobutton(lan_frame, text="LAN SW", value="lan", variable=self.lan_sw_var).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(lan_frame, text="Non LAN SW", value="non_lan", variable=self.lan_sw_var).pack(side=tk.LEFT, padx=5)
        
        row += 1
        ttk.Label(param_frame, text="LM/V select:").grid(row=row, column=0, sticky=tk.W, pady=2)
        lm_frame = ttk.Frame(param_frame)
        lm_frame.grid(row=row, column=1, sticky=tk.W, pady=2)
        ttk.Radiobutton(lm_frame, text="LM", value="LM", variable=self.lm_v_var).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(lm_frame, text="V", value="V", variable=self.lm_v_var).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(lm_frame, text="Both", value="both", variable=self.lm_v_var).pack(side=tk.LEFT, padx=5)
        
        row += 1
        ttk.Checkbutton(param_frame, text="Hide values read", variable=self.hide_values_var).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
        
        row += 1
        ttk.Checkbutton(param_frame, text="Create w/ read values", variable=self.create_with_read_var).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=2)
    
    def _build_editor_tab(self, parent):
        """Build the Register Editor tab."""
        # Left panel: Register list
        left_frame = ttk.Frame(parent)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        ttk.Label(left_frame, text="Select Register Offset:", font=('TkDefaultFont', 10, 'bold')).pack(pady=5)
        
        # Treeview for register list
        tree_frame = ttk.Frame(left_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.register_tree = ttk.Treeview(tree_frame, columns=('Offset', 'Bits', 'SpecName'), show='headings', yscrollcommand=tree_scroll.set)
        self.register_tree.heading('Offset', text='Offset')
        self.register_tree.heading('Bits', text='Bits')
        self.register_tree.heading('SpecName', text='Spec name')
        self.register_tree.column('Offset', width=80)
        self.register_tree.column('Bits', width=80)
        self.register_tree.column('SpecName', width=200)
        self.register_tree.pack(fill=tk.BOTH, expand=True)
        tree_scroll.config(command=self.register_tree.yview)
        
        self.register_tree.bind('<<TreeviewSelect>>', self.on_register_select)
        
        # Right panel: Register details and editor
        right_frame = ttk.Frame(parent)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        ttk.Label(right_frame, text="Register Details:", font=('Segoe UI', 10, 'bold')).pack(pady=5)
        
        # Details frame
        details_frame = ttk.LabelFrame(right_frame, text="Selected Register Information", padding=10)
        details_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.detail_text = scrolledtext.ScrolledText(details_frame, height=12, width=60, wrap=tk.WORD, 
                                                     state='disabled', font=('Consolas', 9))
        self.detail_text.pack(fill=tk.BOTH, expand=True)
        
        # Editor frame
        editor_frame = ttk.LabelFrame(right_frame, text="Edit Value", padding=10)
        editor_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(editor_frame, text="Current Value (hex):").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.reg_value_var = tk.StringVar()
        self.reg_value_entry = ttk.Entry(editor_frame, textvariable=self.reg_value_var, width=15)
        self.reg_value_entry.grid(row=0, column=1, pady=2, sticky=tk.W)
        
        ttk.Label(editor_frame, text="Variant:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.reg_variant_var = tk.StringVar(value="LM")
        ttk.Radiobutton(editor_frame, text="LM", value="LM", variable=self.reg_variant_var).grid(row=1, column=1, sticky=tk.W)
        ttk.Radiobutton(editor_frame, text="V", value="V", variable=self.reg_variant_var).grid(row=1, column=2, sticky=tk.W)
        
        ttk.Button(editor_frame, text="Update Register", command=self.update_register_value).grid(row=2, column=0, columnspan=3, pady=10)
    
    def on_register_select(self, event=None):
        """Handle register selection in treeview."""
        selection = self.register_tree.selection()
        if not selection:
            return
        
        item = self.register_tree.item(selection[0])
        values = item['values']
        
        # Find register data - convert to strings for comparison
        offset = str(values[0])
        bit_range = str(values[1])
        
        # Debug: log what we're searching for
        self.log(f"Looking for register: offset={offset}, bits={bit_range}")
        
        for reg in self.register_data:
            if str(reg['offset']) == offset and str(reg['bits']) == bit_range:
                # Debug: log what we found
                desc_preview = str(reg['description'])[:60]
                self.log(f"  Found register at row {reg['row']}: {desc_preview}...")
                
                # Update details
                self.detail_text.config(state='normal')
                self.detail_text.delete('1.0', tk.END)
                
                # Format the details nicely
                self.detail_text.insert('1.0', "═" * 60 + "\n", 'header')
                self.detail_text.insert(tk.END, f"REGISTER INFORMATION\n", 'header')
                self.detail_text.insert(tk.END, "═" * 60 + "\n\n", 'header')
                
                self.detail_text.insert(tk.END, f"Word Offset: {reg['offset']}\n", 'bold')
                self.detail_text.insert(tk.END, f"Bit Range:   {reg['bits']}\n", 'bold')
                self.detail_text.insert(tk.END, f"\nSpec Name:\n", 'section')
                self.detail_text.insert(tk.END, f"  {reg['c_spec']}\n\n")
                
                self.detail_text.insert(tk.END, f"RTL Symbol:\n", 'section')
                self.detail_text.insert(tk.END, f"  {reg['symbol']}\n\n")
                
                self.detail_text.insert(tk.END, f"Description:\n", 'section')
                desc = reg.get('description', 'N/A')
                if desc and desc != 'None' and str(desc).strip():
                    self.detail_text.insert(tk.END, f"  {desc}\n\n")
                else:
                    self.detail_text.insert(tk.END, f"  (No description available)\n\n")
                
                self.detail_text.insert(tk.END, f"Owner: {reg['owner']}\n\n", 'section')
                
                self.detail_text.insert(tk.END, "─" * 60 + "\n")
                self.detail_text.insert(tk.END, f"Current Values:\n", 'section')
                self.detail_text.insert(tk.END, f"  V Variant (Consumer):  {reg['v_value']}\n")
                self.detail_text.insert(tk.END, f"  LM Variant (Corporate): {reg['lm_value']}\n")
                self.detail_text.insert(tk.END, "─" * 60 + "\n")
                
                # Configure text tags for styling
                self.detail_text.tag_config('header', font=('Consolas', 9, 'bold'))
                self.detail_text.tag_config('section', font=('Consolas', 9, 'bold'), foreground='#0071C5')
                self.detail_text.tag_config('bold', font=('Consolas', 9, 'bold'))
                
                self.detail_text.config(state='disabled')
                
                # Update value entry based on selected variant
                if self.reg_variant_var.get() == "LM":
                    self.reg_value_var.set(str(reg['lm_value']))
                else:
                    self.reg_value_var.set(str(reg['v_value']))
                
                break
    
    def update_register_value(self):
        """Update the register value in Excel."""
        selection = self.register_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a register first.")
            return
        
        item = self.register_tree.item(selection[0])
        values = item['values']
        offset = str(values[0])
        bit_range = str(values[1])
        
        new_value = self.reg_value_var.get().strip()
        if not new_value:
            messagebox.showwarning("Warning", "Please enter a value.")
            return
        
        # Find the register row in register_data
        for reg in self.register_data:
            if str(reg['offset']) == offset and str(reg['bits']) == bit_range:
                try:
                    # Update Excel
                    wb = load_workbook(str(self.current_xlsm), keep_vba=True)
                    ws = wb["full nvm map"]
                    
                    row_num = reg['row']
                    variant = self.reg_variant_var.get()
                    col = 7 if variant == "V" else 8  # Column G or H
                    
                    ws.cell(row_num, col).value = new_value
                    wb.save(str(self.current_xlsm))
                    wb.close()
                    
                    # Update local data
                    if variant == "V":
                        reg['v_value'] = new_value
                    else:
                        reg['lm_value'] = new_value
                    
                    self.log(f"Updated {offset}[{bit_range}] {variant} variant to: {new_value}")
                    messagebox.showinfo("Success", f"Register updated successfully!")
                    
                    # Refresh the details display
                    self.on_register_select()
                    
                except Exception as exc:
                    messagebox.showerror("Error", f"Failed to update register: {exc}")
                    self.log(f"Error updating register: {exc}")
                
                break
    
    def load_register_data(self):
        """Load all register data from Excel into memory."""
        if not self.current_xlsm or not self.current_xlsm.exists():
            return
        
        try:
            wb = load_workbook(str(self.current_xlsm), data_only=True)
            ws = wb["full nvm map"]
            
            self.register_data = []
            
            # Read from row 6 onwards
            for row in range(6, 500):
                offset = ws.cell(row, 1).value
                if not offset:
                    break
                
                bits = ws.cell(row, 2).value
                symbol = ws.cell(row, 3).value or "N/A"
                c_spec = ws.cell(row, 4).value or "N/A"
                owner = ws.cell(row, 5).value or "N/A"
                description = ws.cell(row, 6).value or "No description"
                v_value = ws.cell(row, 7).value
                lm_value = ws.cell(row, 8).value
                
                self.register_data.append({
                    'row': row,
                    'offset': str(offset),
                    'bits': str(bits),
                    'symbol': str(symbol),
                    'c_spec': str(c_spec),
                    'owner': str(owner),
                    'description': str(description),
                    'v_value': str(v_value),
                    'lm_value': str(lm_value)
                })
            
            wb.close()
            
            # Populate treeview
            self.register_tree.delete(*self.register_tree.get_children())
            for reg in self.register_data:
                self.register_tree.insert('', 'end', values=(reg['offset'], reg['bits'], reg['c_spec']))
            
            # Debug: Log first few and last few descriptions to verify they're different
            if len(self.register_data) > 0:
                self.log(f"Loaded {len(self.register_data)} registers from Excel")
                self.log(f"  Sample descriptions:")
                for i in [0, 1, 10, 20, 30]:
                    if i < len(self.register_data):
                        reg = self.register_data[i]
                        desc_preview = str(reg['description'])[:50]
                        self.log(f"    [{i}] {reg['offset']}[{reg['bits']}]: {desc_preview}...")
            
        except Exception as exc:
            self.log(f"Error loading register data: {exc}")
    
    def log(self, message: str):
        """Add message to log window."""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.update_idletasks()
    
    def _load_gbe_folders(self):
        """Scan GBE_Image folder for project subfolders."""
        workspace_root = Path(__file__).parent.parent
        gbe_image_root = workspace_root / "GBE_Image"
        
        if not gbe_image_root.exists():
            gbe_image_root.mkdir(parents=True, exist_ok=True)
        
        self.gbe_folders = [f for f in gbe_image_root.iterdir() if f.is_dir() and not f.name.startswith(".")]
        self.folder_combo["values"] = [f.name for f in self.gbe_folders]
        
        if self.gbe_folders and not self.folder_var.get():
            self.folder_combo.current(0)
            self.on_project_select()
    
    def on_project_select(self):
        """Load the selected project folder."""
        if not self.folder_var.get():
            return
        
        folder_name = self.folder_var.get()
        folder_path = None
        for f in self.gbe_folders:
            if f.name == folder_name:
                folder_path = f
                break
        
        if not folder_path or not folder_path.exists():
            messagebox.showerror("Error", f"Folder not found: {folder_name}")
            return
        
        self.current_folder = folder_path
        
        # Find .xlsm file
        xlsm_files = list(self.current_folder.glob("*.xlsm"))
        if xlsm_files:
            self.current_xlsm = xlsm_files[0]
            self.log(f"Loaded project: {folder_name}")
            self.log(f"Excel workbook: {self.current_xlsm.name}")
            # Auto-load parameters from Excel
            self.load_from_excel()
        else:
            self.current_xlsm = None
            self.log(f"Warning: No Excel workbook found in {folder_name}")
    
    def load_from_excel(self):
        """Load parameters from Excel workbook."""
        if not self.current_xlsm or not self.current_xlsm.exists():
            messagebox.showerror("Error", "No Excel workbook found. Please select a project first.")
            return
        
        if load_workbook is None:
            messagebox.showerror("Error", "openpyxl is not installed.")
            return
        
        try:
            wb = load_workbook(str(self.current_xlsm), keep_vba=False, data_only=True)
            
            if "general variable" not in wb.sheetnames:
                messagebox.showerror("Error", "'general variable' sheet not found.")
                wb.close()
                return
            
            ws = wb["general variable"]
            
            # Load parameters from Excel
            self.project_var.set(str(ws["B2"].value or ""))
            self.silicon_var.set(str(ws["B3"].value or ""))
            self.step_var.set(str(ws["B4"].value or ""))
            self.nvm_output_var.set(str(ws["B5"].value or ""))
            
            # Parse version (no auto-increment on load - user can manually change)
            version_str = str(ws["B7"].value or "101")
            if len(version_str) >= 3:
                major = int(version_str[0])
                minor = int(version_str[1:])
                
                self.major_var.set(str(major))
                self.minor_var.set(str(minor))
            
            # Load device IDs (preserving 0x prefix)
            device_id = str(ws["B8"].value or "")
            sku_device_id = str(ws["B9"].value or "")
            self.device_id_var.set(device_id)
            self.sku_device_id_var.set(sku_device_id)
            
            wb.close()
            
            self.log(f"Parameters loaded from Excel successfully.")
            self.log(f"Version: {major}.{minor:02d}, Device ID: {device_id}, SKU: {sku_device_id}")
            
            # Load register data for editor tab
            self.load_register_data()
            
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to load from Excel: {exc}")
            self.log(f"Error loading from Excel: {exc}")
    
    def generate_nvm(self):
        """Generate NVM files (main entry point, replicates genNvmCMDline)."""
        if not self.current_xlsm or not self.current_xlsm.exists():
            messagebox.showerror("Error", "No Excel workbook found. Please select a project first.")
            return
        
        if load_workbook is None:
            messagebox.showerror("Error", "openpyxl is not installed.")
            return
        
        # Validate inputs
        if not self.project_var.get().strip():
            messagebox.showerror("Error", "Project name is required.")
            return
        if not self.major_var.get().strip() or not self.minor_var.get().strip():
            messagebox.showerror("Error", "Major and minor version are required.")
            return
        if not self.device_id_var.get().strip() or not self.sku_device_id_var.get().strip():
            messagebox.showerror("Error", "Device ID and SKU Device ID are required.")
            return
        
        self.log("="*60)
        self.log("Starting NVM generation...")
        
        try:
            # Write parameters to Excel
            wb = load_workbook(str(self.current_xlsm), keep_vba=True)
            if "general variable" not in wb.sheetnames:
                messagebox.showerror("Error", "'general variable' sheet not found.")
                wb.close()
                return
            
            ws = wb["general variable"]
            ws["B2"] = self.project_var.get().strip()
            ws["B3"] = self.silicon_var.get().strip()
            ws["B4"] = self.step_var.get().strip()
            
            image_file_name = self.nvm_output_var.get().strip() or self.project_var.get().strip()
            ws["B5"] = image_file_name
            
            # Build version string (e.g., 104 for v1.04)
            major = self.major_var.get().strip()
            minor = self.minor_var.get().strip().zfill(2)
            nvm_image_version = int(major + minor)
            ws["B7"] = nvm_image_version
            
            # Ensure device IDs have 0x prefix
            device_id = self.device_id_var.get().strip()
            if not device_id.startswith("0x") and not device_id.startswith("0X"):
                device_id = "0x" + device_id
            sku_device_id = self.sku_device_id_var.get().strip()
            if not sku_device_id.startswith("0x") and not sku_device_id.startswith("0X"):
                sku_device_id = "0x" + sku_device_id
            
            ws["B8"] = device_id
            ws["B9"] = sku_device_id
            
            wb.save(str(self.current_xlsm))
            wb.close()
            
            self.log(f"Parameters written to Excel: {self.current_xlsm.name}")
            
            # Generate based on selection
            # VBA calls incEEtracID twice per variant to match the commented-out LAN_SW pattern
            lm_v_mode = self.lm_v_var.get()
            
            if lm_v_mode in ["LM", "both"]:
                self._increment_eetrack_id()  # First increment (matches VBA)
                self._increment_eetrack_id()  # Second increment (matches VBA)
                self._generate_variant("LM", image_file_name, nvm_image_version)
            
            if lm_v_mode in ["V", "both"]:
                self._increment_eetrack_id()  # First increment (matches VBA)
                self._increment_eetrack_id()  # Second increment (matches VBA)
                self._generate_variant("V", image_file_name, nvm_image_version)
            
            self.log("="*60)
            self.log("NVM generation completed successfully!")
            messagebox.showinfo("Success", "NVM files generated successfully!")
            
        except Exception as exc:
            self.log(f"ERROR: {exc}")
            messagebox.showerror("Error", f"Failed to generate NVM: {exc}")
    
    def _increment_eetrack_id(self):
        """Increment EEtrack ID in Excel (replicates incEEtracID)."""
        try:
            wb = load_workbook(str(self.current_xlsm), keep_vba=True)
            ws = wb["full nvm map"]
            
            # Read EEtrack ID from cells G12 (high) and G13 (low)
            high = str(ws["G12"].value or "0x0000").replace("0x", "")
            low = str(ws["G13"].value or "0x0000").replace("0x", "")
            
            eetrack_id = int(high + low, 16)
            eetrack_id += 2
            
            eetrack_hex = f"{eetrack_id:08X}"
            high_new = f"0x{eetrack_hex[:4]}"
            low_new = f"0x{eetrack_hex[4:]}"
            
            ws["G12"] = high_new
            ws["H12"] = high_new
            ws["G13"] = low_new
            ws["H13"] = low_new
            
            wb.save(str(self.current_xlsm))
            wb.close()
            
            self.log(f"Incremented EEtrack ID to: {high_new} {low_new}")
            
        except Exception as exc:
            self.log(f"Warning: Failed to increment EEtrack ID: {exc}")
    
    def _recalculate_excel_formulas(self, xlsm_path: Path):
        """Use Excel COM to recalculate formulas in the workbook."""
        if not EXCEL_COM_AVAILABLE:
            self.log("Warning: win32com not available, formulas may not be recalculated")
            return False
        
        try:
            import pythoncom
            pythoncom.CoInitialize()
            
            # Use DispatchEx to create a new Excel instance (avoids conflicts)
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            
            wb = excel.Workbooks.Open(str(xlsm_path.absolute()), UpdateLinks=0, ReadOnly=False)
            
            # Force full calculation of all worksheets
            excel.CalculateFull()  # Calculate all cells in all open workbooks
            
            # Wait for calculation to complete
            import time
            time.sleep(2)  # Give Excel time to finish calculations
            
            # Save and close
            wb.Save()
            wb.Close(SaveChanges=True)
            
            excel.Quit()
            
            # Clean up COM objects
            del wb
            del excel
            pythoncom.CoUninitialize()
            
            self.log("  Excel formulas recalculated successfully")
            return True
            
        except Exception as exc:
            self.log(f"Warning: Failed to recalculate Excel formulas: {exc}")
            return False
    
    def _write_and_calculate_excel(self, variant: str, nvm_image_version: int):
        """Use Excel COM to write values and calculate all formulas."""
        if not EXCEL_COM_AVAILABLE:
            self.log("Warning: win32com not available, using openpyxl fallback")
            return False
        
        try:
            import pythoncom
            import time
            pythoncom.CoInitialize()
            
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            wb = excel.Workbooks.Open(str(self.current_xlsm.absolute()), UpdateLinks=0, ReadOnly=False)
            
            ws_var = wb.Worksheets("general variable")
            ws_map = wb.Worksheets("full nvm map")
            
            # Read device IDs
            dev_id = ws_var.Range("B8").Value
            sku_dev_id = ws_var.Range("B9").Value
            
            # Write device IDs based on variant
            if variant == "LM":
                ws_map.Range("H30").Value = sku_dev_id
                ws_map.Range("H150").Value = sku_dev_id
                ws_map.Range("H152").Value = sku_dev_id
                ws_map.Range("H155").Value = sku_dev_id
                ws_map.Range("H157").Value = sku_dev_id
                ws_map.Range("H153").Value = dev_id
                ws_var.Range("B10").Value = True
                ws_var.Range("B11").Value = False
            else:  # V
                ws_map.Range("G30").Value = dev_id
                ws_map.Range("G150").Value = sku_dev_id
                ws_map.Range("G152").Value = sku_dev_id
                ws_map.Range("G155").Value = sku_dev_id
                ws_map.Range("G157").Value = sku_dev_id
                ws_map.Range("G153").Value = dev_id
                ws_var.Range("B10").Value = False
                ws_var.Range("B11").Value = True
            
            # NON_LAN_SW_Init logic
            lan_sw_nvm = (self.lan_sw_var.get() == "lan")
            ws_var.Range("B12").Value = lan_sw_nvm
            
            # Calculate LCD extension pointer
            extention_sheet = "LCD extention LAN SW" if lan_sw_nvm else "LCD extention non LAN SW"
            ws_ext = wb.Worksheets(extention_sheet)
            lcd_row_count = ws_ext.UsedRange.Rows.Count
            lcd_pointer = (lcd_row_count - 1) // 2
            lcd_pointer_hex = f"0x{lcd_pointer:02X}"
            ws_map.Range("G84").Value = lcd_pointer_hex
            ws_map.Range("H84").Value = lcd_pointer_hex
            
            # Format version
            version_high = nvm_image_version // 100
            version_low = nvm_image_version % 10
            version_hex = f"0x{version_high:X}0{version_low:X}4"
            ws_map.Range("G11").Value = version_hex
            ws_map.Range("H11").Value = version_hex
            
            # Force full calculation
            excel.CalculateFull()
            time.sleep(2)  # Wait for calculation
            
            # Save and close
            wb.Save()
            wb.Close(SaveChanges=True)
            excel.Quit()
            
            del wb
            del excel
            pythoncom.CoUninitialize()
            
            self.log("  Excel values written and formulas calculated")
            return True
            
        except Exception as exc:
            self.log(f"ERROR: Failed to write/calculate Excel: {exc}")
            try:
                excel.Quit()
            except:
                pass
            pythoncom.CoUninitialize()
            return False
    
    def _generate_variant(self, variant: str, image_file_name: str, nvm_image_version: int):
        """Generate a single variant (LM or V). Replicates genNVM()."""
        self.log(f"\nGenerating {variant} variant...")
        
        try:
            # Use Excel COM to write ALL values and calculate formulas
            # This ensures formulas are properly calculated before we read them
            self._write_and_calculate_excel(variant, nvm_image_version)
            
            # Now read calculated values with openpyxl
            col_to_use = 8 if variant == "LM" else 7
            lan_sw_nvm = (self.lan_sw_var.get() == "lan")  # Recalculate for use below
            
            # Reload workbook with data_only=True to get calculated formula values
            wb = load_workbook(str(self.current_xlsm), data_only=True)
            ws_var = wb["general variable"]
            ws_map = wb["full nvm map"]
            
            # Calculate file version string (matches VBA logic)
            # VBA: If Mid(version, 2, 1) = 0 Then "x.y" Else "x.yy"
            if str(nvm_image_version)[1] == "0":
                file_version = f"{nvm_image_version // 100}.{nvm_image_version % 10}"
            else:
                file_version = f"{nvm_image_version // 100}.{nvm_image_version % 100}"
            
            sku_suffix = "_Release_Corp_Prod_NA" if variant == "LM" else "_Release_Cons_Prod_NA"
            
            # Create output directory
            output_dir_name = f"{image_file_name}_{file_version}{sku_suffix}"
            output_dir = self.current_folder / output_dir_name
            output_dir.mkdir(exist_ok=True)
            
            bin_file_path = output_dir / f"{image_file_name}_{file_version}{sku_suffix}.bin"
            txt_file_path = output_dir / f"{image_file_name}_{file_version}{sku_suffix}.txt"
            json_file_path = output_dir / "parameters.json"
            
            self.log(f"Output directory: {output_dir_name}")
            
            # Generate binary and text files
            with open(bin_file_path, "wb") as bin_file, open(txt_file_path, "w") as txt_file:
                txt_word_counter = 0
                txt_line = ""
                
                # Process "full nvm map" sheet
                self.log("  Processing 'full nvm map' sheet...")
                current_row = 6
                word_data_bin = ""
                
                while ws_map.cell(current_row, 1).value is not None:
                    range_data = ws_map.cell(current_row, col_to_use).value
                    bit_range = ws_map.cell(current_row, 2).value
                    
                    if not range_data or not bit_range:
                        current_row += 1
                        continue
                    
                    # Parse bit range to get size
                    try:
                        if ":" in str(bit_range):
                            parts = str(bit_range).split(":")
                            msb = int(parts[0])
                            lsb = int(parts[1])
                            places = msb - lsb + 1
                        else:
                            places = 16
                    except (ValueError, IndexError) as e:
                        self.log(f"Warning: Invalid bit range '{bit_range}' at row {current_row}, skipping")
                        current_row += 1
                        continue
                    
                    # Convert range_data to binary
                    range_data_str = str(range_data).replace("0x", "").replace("0X", "")
                    try:
                        range_data_int = int(range_data_str, 16)
                        range_data_bin = format(range_data_int, f"0{places}b")
                        word_data_bin += range_data_bin
                    except ValueError as e:
                        self.log(f"Warning: Invalid hex value '{range_data_str}' at row {current_row}, skipping")
                        current_row += 1
                        continue
                    
                    # Check if this is the last range of the word
                    next_word = ws_map.cell(current_row + 1, 1).value
                    current_word = ws_map.cell(current_row, 1).value
                    
                    if current_word != next_word:
                        # Write word to files
                        # Ensure binary string is max 16 bits
                        if len(word_data_bin) > 16:
                            self.log(f"Warning: Word at row {current_row} has {len(word_data_bin)} bits, truncating to 16 bits")
                            word_data_bin = word_data_bin[-16:]  # Take last 16 bits
                        
                        word_value = int(word_data_bin, 2) if word_data_bin else 0
                        
                        # Validate range
                        if word_value < 0 or word_value > 65535:
                            self.log(f"ERROR: Word value {word_value} at row {current_row} (word {current_word}) is out of range")
                            word_value = word_value & 0xFFFF  # Mask to 16 bits
                        
                        # Write to bin file (little-endian)
                        bin_file.write(struct.pack("<H", word_value))
                        
                        # Write to txt file
                        word_hex = f"{word_value:04X}"
                        txt_word_counter += 1
                        if txt_word_counter == 8:
                            txt_line += word_hex
                            txt_file.write(txt_line + "\n")
                            txt_word_counter = 0
                            txt_line = ""
                        else:
                            txt_line += word_hex + " "
                        
                        word_data_bin = ""
                    
                    current_row += 1
                
                # Process "LCD extention non LAN SW" or "LCD extention LAN SW" sheet
                extention_sheet = "LCD extention LAN SW" if lan_sw_nvm else "LCD extention non LAN SW"
                self.log(f"  Processing '{extention_sheet}' sheet...")
                
                if extention_sheet in wb.sheetnames:
                    ws_ext = wb[extention_sheet]
                    current_row = 2
                    word_data_bin = ""
                    last_valid_word_address = None
                    
                    while ws_ext.cell(current_row, 1).value is not None:
                        # Get values
                        word_addr = ws_ext.cell(current_row, 1).value
                        range_data = ws_ext.cell(current_row, 3).value
                        bit_range = ws_ext.cell(current_row, 2).value
                        
                        # Skip rows with missing data or non-hex word addresses
                        if not range_data or not bit_range or not word_addr:
                            current_row += 1
                            continue
                        
                        # Validate word address is hex
                        word_addr_str = str(word_addr).replace("h", "").replace("0x", "")
                        try:
                            int(word_addr_str, 16)
                            last_valid_word_address = word_addr
                        except ValueError:
                            # Skip header rows or invalid addresses
                            current_row += 1
                            continue
                        
                        # Parse bit range to get size
                        try:
                            if ":" in str(bit_range):
                                parts = str(bit_range).split(":")
                                msb = int(parts[0])
                                lsb = int(parts[1])
                                places = msb - lsb + 1
                            else:
                                places = 16
                        except (ValueError, IndexError) as e:
                            self.log(f"Warning: Invalid bit range '{bit_range}' at LCD row {current_row}, skipping")
                            current_row += 1
                            continue
                        
                        # Convert range_data to binary
                        range_data_str = str(range_data).replace("0x", "").replace("0X", "")
                        try:
                            range_data_int = int(range_data_str, 16)
                            range_data_bin = format(range_data_int, f"0{places}b")
                            word_data_bin += range_data_bin
                        except ValueError as e:
                            self.log(f"Warning: Invalid hex value '{range_data_str}' at LCD row {current_row}, skipping")
                            current_row += 1
                            continue
                        
                        next_word = ws_ext.cell(current_row + 1, 1).value
                        current_word = ws_ext.cell(current_row, 1).value
                        
                        # Check if next row has different word address (or is end)
                        is_last_range = (current_word != next_word)
                        
                        if is_last_range:
                            # Ensure binary string is max 16 bits
                            if len(word_data_bin) > 16:
                                self.log(f"Warning: LCD word at row {current_row} has {len(word_data_bin)} bits, truncating to 16 bits")
                                word_data_bin = word_data_bin[-16:]  # Take last 16 bits
                            
                            word_value = int(word_data_bin, 2) if word_data_bin else 0
                            
                            # Validate range
                            if word_value < 0 or word_value > 65535:
                                self.log(f"ERROR: LCD word value {word_value} at row {current_row} (word {current_word}) is out of range")
                                word_value = word_value & 0xFFFF  # Mask to 16 bits
                            
                            bin_file.write(struct.pack("<H", word_value))
                            
                            word_hex = f"{word_value:04X}"
                            txt_word_counter += 1
                            if txt_word_counter == 8:
                                txt_line += word_hex
                                txt_file.write(txt_line + "\n")
                                txt_word_counter = 0
                                txt_line = ""
                            else:
                                txt_line += word_hex + " "
                            
                            word_data_bin = ""
                        
                        current_row += 1
                    
                    # Fill with FFFF from end of LCD extention to start of ISCSI
                    if last_valid_word_address:
                        last_word_str = str(last_valid_word_address).replace("h", "").replace("0x", "")
                        try:
                            start_fill = int(last_word_str, 16) + 1
                        except ValueError:
                            self.log(f"Warning: Could not parse last LCD word address '{last_valid_word_address}', skipping fill")
                            start_fill = None
                    else:
                        start_fill = None
                    
                    if start_fill is not None:
                    
                        # Get end address from full nvm map G308
                        end_fill_str = str(ws_map["G308"].value or "0x0").replace("0x", "")
                        try:
                            end_fill = int(end_fill_str, 16) - 1
                        except ValueError:
                            self.log(f"Warning: Could not parse end fill address from G308, skipping fill")
                            end_fill = None
                        
                        if end_fill is not None and end_fill >= start_fill:
                            self.log(f"  Filling 0xFFFF from {start_fill:04X}h to {end_fill:04X}h...")
                            
                            for i in range(start_fill, end_fill + 1):
                                bin_file.write(struct.pack("<H", 0xFFFF))
                                txt_word_counter += 1
                                if txt_word_counter == 8:
                                    txt_line += "FFFF"
                                    txt_file.write(txt_line + "\n")
                                    txt_word_counter = 0
                                    txt_line = ""
                                else:
                                    txt_line += "FFFF "
                
                # Process "ISCSI_MODULE" sheet
                self.log("  Processing 'ISCSI_MODULE' sheet...")
                
                if "ISCSI_MODULE" in wb.sheetnames:
                    ws_iscsi = wb["ISCSI_MODULE"]
                    current_row = 2
                    word_data_bin = ""
                    last_valid_word_address = None
                    
                    while ws_iscsi.cell(current_row, 1).value is not None:
                        # Get values  
                        word_addr = ws_iscsi.cell(current_row, 1).value
                        range_data = ws_iscsi.cell(current_row, 3).value
                        bit_range = ws_iscsi.cell(current_row, 2).value
                        
                        # Skip rows with missing data or non-hex word addresses
                        if not range_data or not bit_range or not word_addr:
                            current_row += 1
                            continue
                        
                        # Validate word address is hex
                        word_addr_str = str(word_addr).replace("h", "").replace("0x", "")
                        try:
                            int(word_addr_str, 16)
                            last_valid_word_address = word_addr
                        except ValueError:
                            # Skip header rows or invalid addresses
                            current_row += 1
                            continue
                        
                        # Parse bit range to get size
                        try:
                            if ":" in str(bit_range):
                                parts = str(bit_range).split(":")
                                msb = int(parts[0])
                                lsb = int(parts[1])
                                places = msb - lsb + 1
                            else:
                                places = 16
                        except (ValueError, IndexError) as e:
                            self.log(f"Warning: Invalid bit range '{bit_range}' at ISCSI row {current_row}, skipping")
                            current_row += 1
                            continue
                        
                        # Convert range_data to binary
                        range_data_str = str(range_data).replace("0x", "").replace("0X", "")
                        try:
                            range_data_int = int(range_data_str, 16)
                            range_data_bin = format(range_data_int, f"0{places}b")
                            word_data_bin += range_data_bin
                        except ValueError as e:
                            self.log(f"Warning: Invalid hex value '{range_data_str}' at ISCSI row {current_row}, skipping")
                            current_row += 1
                            continue
                        
                        next_word = ws_iscsi.cell(current_row + 1, 1).value
                        current_word = ws_iscsi.cell(current_row, 1).value
                        
                        if current_word != next_word:
                            # Ensure binary string is max 16 bits
                            if len(word_data_bin) > 16:
                                self.log(f"Warning: ISCSI word at row {current_row} has {len(word_data_bin)} bits, truncating to 16 bits")
                                word_data_bin = word_data_bin[-16:]  # Take last 16 bits
                            
                            word_value = int(word_data_bin, 2) if word_data_bin else 0
                            
                            # Validate range
                            if word_value < 0 or word_value > 65535:
                                self.log(f"ERROR: ISCSI word value {word_value} at row {current_row} (word {current_word}) is out of range")
                                word_value = word_value & 0xFFFF  # Mask to 16 bits
                            
                            bin_file.write(struct.pack("<H", word_value))
                            
                            word_hex = f"{word_value:04X}"
                            txt_word_counter += 1
                            if txt_word_counter == 8:
                                txt_line += word_hex
                                txt_file.write(txt_line + "\n")
                                txt_word_counter = 0
                                txt_line = ""
                            else:
                                txt_line += word_hex + " "
                            
                            word_data_bin = ""
                        
                        current_row += 1
                    
                    # Fill with FFFF from end of ISCSI to 4087 (matches VBA)
                    if last_valid_word_address:
                        last_word_str = str(last_valid_word_address).replace("h", "").replace("0x", "")
                        try:
                            start_fill = int(last_word_str, 16) + 1
                            end_fill = 4087  # VBA fills to 4087, not 4095
                            
                            self.log(f"  Filling 0xFFFF from {start_fill:04X}h to {end_fill:04X}h...")
                            
                            for i in range(start_fill, end_fill + 1):
                                bin_file.write(struct.pack("<H", 0xFFFF))
                                txt_word_counter += 1
                                if txt_word_counter == 8:
                                    txt_line += "FFFF"
                                    txt_file.write(txt_line + "\n")
                                    txt_word_counter = 0
                                    txt_line = ""
                                else:
                                    txt_line += "FFFF "
                        except ValueError:
                            self.log(f"Warning: Could not parse last ISCSI word address '{last_valid_word_address}', skipping fill")
                
                # Final safety: ensure we have exactly 4096 words (8192 bytes)
                current_pos = bin_file.tell()
                expected_size = 4096 * 2  # 4096 words * 2 bytes per word
                if current_pos < expected_size:
                    words_remaining = (expected_size - current_pos) // 2
                    self.log(f"  Final fill: Adding {words_remaining} more 0xFFFF words to reach 4096 total")
                    for i in range(words_remaining):
                        bin_file.write(struct.pack("<H", 0xFFFF))
                        txt_word_counter += 1
                        if txt_word_counter == 8:
                            txt_line += "FFFF"
                            txt_file.write(txt_line + "\n")
                            txt_word_counter = 0
                            txt_line = ""
                        else:
                            txt_line += "FFFF "
                
                # Flush any remaining txt line
                if txt_line.strip():
                    txt_file.write(txt_line.strip() + "\n")
            
            # Generate JSON parameters file
            params = {
                "projectName": self.project_var.get().strip(),
                "silicon": self.silicon_var.get().strip(),
                "step": self.step_var.get().strip(),
                "version": {
                    "major": self.major_var.get().strip(),
                    "minor": self.minor_var.get().strip(),
                    "combined": file_version
                },
                "deviceId": self.device_id_var.get().strip(),
                "skuDeviceId": self.sku_device_id_var.get().strip(),
                "nvmOutput": image_file_name,
                "variant": variant,
                "lanMode": self.lan_sw_var.get(),
                "hideValuesRead": self.hide_values_var.get(),
                "createWithReadValues": self.create_with_read_var.get()
            }
            
            json_file_path.write_text(json.dumps(params, indent=2), encoding="utf-8")
            
            wb.save(str(self.current_xlsm))
            wb.close()
            
            self.log(f"  Generated: {bin_file_path.name}")
            self.log(f"  Generated: {txt_file_path.name}")
            self.log(f"  Generated: {json_file_path.name}")
            
            # Calculate checksum (replicates calc_csum.py)
            self._update_checksum(bin_file_path, txt_file_path)
            
        except Exception as exc:
            self.log(f"ERROR generating {variant} variant: {exc}")
            raise
    
    def _update_checksum(self, bin_path: Path, txt_path: Path):
        """Update checksum in bin and txt files (replicates calc_csum.py)."""
        CHECKSUM_WORD = 0x3F
        CHECKSUM_VALUE = 0xBABA
        CHECKSUM_LINE = 7
        
        try:
            # Read binary file and calculate checksum
            data = bin_path.read_bytes()
            words = [int.from_bytes(data[i:i + 2], "little") for i in range(0, len(data), 2)]
            
            if len(words) <= CHECKSUM_WORD:
                self.log("Warning: Binary file too small for checksum.")
                return
            
            total = sum(words[0:CHECKSUM_WORD])
            checksum = ((-total) + (2 ** 32) + CHECKSUM_VALUE) & 0xFFFF
            
            # Update binary file
            words[CHECKSUM_WORD] = checksum
            new_data = b"".join(w.to_bytes(2, "little") for w in words)
            bin_path.write_bytes(new_data)
            
            # Update txt file
            lines = txt_path.read_text(encoding="utf-8").splitlines()
            if len(lines) > CHECKSUM_LINE:
                parts = lines[CHECKSUM_LINE].split()
                if parts:
                    parts[-1] = f"{checksum:04X}"
                    lines[CHECKSUM_LINE] = " ".join(parts)
                    txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
            
            self.log(f"  Updated checksum: 0x{checksum:04X}")
            
        except Exception as exc:
            self.log(f"Warning: Failed to update checksum: {exc}")
    
    def read_nvm(self):
        """Read existing NVM file into Excel (replicates readNVM)."""
        if not self.current_xlsm or not self.current_xlsm.exists():
            messagebox.showerror("Error", "No Excel workbook found. Please select a project first.")
            return
        
        bin_file = filedialog.askopenfilename(
            title="Select NVM bin file to read",
            filetypes=[("Binary files", "*.bin"), ("All files", "*.*")]
        )
        
        if not bin_file:
            return
        
        self.log(f"Reading NVM file: {Path(bin_file).name}")
        
        try:
            wb = load_workbook(str(self.current_xlsm), keep_vba=True)
            ws = wb["full nvm map"]
            
            # Read binary file
            data = Path(bin_file).read_bytes()
            words = [int.from_bytes(data[i:i + 2], "little") for i in range(0, len(data), 2)]
            
            # Populate column I
            current_row = 6
            word_idx = 0
            
            while ws.cell(current_row, 1).value is not None and word_idx < len(words):
                bit_range = ws.cell(current_row, 2).value
                
                if bit_range and ":" in str(bit_range):
                    try:
                        parts = str(bit_range).split(":")
                        msb = int(parts[0])
                        lsb = int(parts[1])
                        
                        # Extract value from word
                        mask = (1 << (msb - lsb + 1)) - 1
                        value = (words[word_idx] >> lsb) & mask
                        
                        ws.cell(current_row, 9).value = f"0x{value:X}"
                    except (ValueError, IndexError) as e:
                        self.log(f"Warning: Could not parse bit range '{bit_range}' at row {current_row}")
                
                # Check if next row is a different word
                next_word = ws.cell(current_row + 1, 1).value
                current_word = ws.cell(current_row, 1).value
                if current_word != next_word:
                    word_idx += 1
                
                current_row += 1
            
            # Show column I
            ws.column_dimensions["I"].hidden = False
            
            wb.save(str(self.current_xlsm))
            wb.close()
            
            self.log(f"Successfully read {len(words)} words into Excel.")
            messagebox.showinfo("Success", f"NVM file read successfully. {len(words)} words loaded.")
            
        except Exception as exc:
            self.log(f"ERROR reading NVM: {exc}")
            messagebox.showerror("Error", f"Failed to read NVM: {exc}")


if __name__ == "__main__":
    app = GBEImageCreator()
    app.mainloop()
